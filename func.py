from calendar import day_abbr
from queue import Empty
import re
import math
import random
from time import time
import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta

def getAllZoneMapping():
    # Load the Excel file
    workbook = openpyxl.load_workbook("Database/Zone Mapping.xlsx")
    sheet = workbook.active

    # Initialize data_dict
    data_dict = {}

    # Get the column headers (zone names) from the first row
    column_headers = [sheet.cell(row=1, column=col).value for col in range(2, sheet.max_column + 1)]

    # Read data from the second row onwards and populate the dictionary
    for row_num in range(2, sheet.max_row + 1):
        key = str(sheet.cell(row=row_num, column=1).value)
        data_dict[key] = {}
        for col_num in range(2, sheet.max_column + 1):
            zone_name = column_headers[col_num - 2]
            value = sheet.cell(row=row_num, column=col_num).value
            data_dict[key][zone_name] = value

    return data_dict

def changeFormatData(data):
       returnDataString = ''

       if data >= 1048576:
              data = round(data/1048576)
              returnDataString = str(data)+' GB'
       elif data >= 1024:
              data = round(data/1024)
              returnDataString = str(data)+' MB'
       else:
              data = round(data)
              returnDataString = str(data)+' KB'
              
       return returnDataString

allZoneMapping = getAllZoneMapping()

def exportExcelFlatCallSLI(eventName, params=None, neededParams = None):
       # Export Test Cases to Excel File
       wb = Workbook()
       ws = wb.active
       
       if params is not None:
              countryData = params['Country']
              prepaidPrice = params['Price Pre Paid']
              postpaidPrice = params['Price Post Paid']
       else:
              countryData = ''
              prepaidPrice = params['Price Pre Paid']
              postpaidPrice = params['Price Post Paid']
       
       for country in countryData:
              country = country.replace('\n', '')
              #Prepaid Section
              stepsPrePaid = [
                     "Create & Activate new subscriber PP Telkomsel Prepaid 10K",
                     "Create event update expired date to "+datetime.now().strftime("%Y")+"-12-31",
                     ["Consume Preload 150Mb","Consume Bonus"],
                     "Create event update balance 5000000",
                     ["Create event voice idd using access code 007 to "+str(country)+" 1200s, 9.30PM","Charged Not IDR "+str(prepaidPrice)],
                     ["Check Pricing Item ID event above","Checked"],
                     ["Create event voice idd using access code 007 to "+str(country)+" 1200s, 10PM","Charged IDR "+str(prepaidPrice)],
                     ["Check Pricing Item ID event above","Checked"],
                     ["Create event voice international to "+str(country)+" 60s, 11PM","Charged Not IDR "+str(prepaidPrice)],
                     "Create event voice initial international to "+str(country)+" using access code 007, 11PM | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 11PM | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 11PM | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 11PM | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 11PM | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 11PM | GSU should be 180s",
                     ["Create event voice intermediate international to "+str(country)+" using access code 007 180s, 11PM | GSU should be 120s","Intermediate Success | Final Unit Ind 0 | FUI 1"],
                     ["Create event voice terminate international to "+str(country)+" using access code 007 120s, 11PM","Charged IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 600s, 11.30PM","Charged Not IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 1s, 1AM D+1 | Rounded should be 1200s","Charged IDR "+str(prepaidPrice)],
                     ["Create event voice onnet 60s, 3AM D+1","Charged Not IDR "+str(prepaidPrice)],
                     ["Create event voice roaming MO Home from "+str(country)+" 60s, 5AM D+1","Charged Not IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 60s, 7AM D+1","Charged Not IDR "+str(prepaidPrice)],
                     "Create event voice initial international to "+str(country)+" using access code 007, 10.10PM D+1 | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 10.10PM D+1 | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 10.10PM D+1 | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 10.10PM D+1 | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 10.10PM D+1 | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 10.10PM D+1 | GSU should be 180s",
                     ["Create event voice intermediate international to "+str(country)+" using access code 007 180s, 10.10PM D+1 | GSU should be 120s","Intermediate Success | Final Unit Ind 0 | FUI 1"],
                     ["Create event voice terminate international to "+str(country)+" using access code 007 120s, 10.10PM D+1","Charged IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 600s, 11PM D+1 | Rounded should be 600s","Charged Not IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 001 to "+str(country)+" 60s, 1AM Before next bc","Charged Not IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 008 to "+str(country)+" 60s, 2AM Before next bc","Charged Not IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 01018 to "+str(country)+" 60s, 3AM Before next bc","Charged Not IDR "+str(prepaidPrice)],
                     ["Create event voice offnet 60s, 3AM Before next bc","Charged Not IDR "+str(prepaidPrice)],
                     ["Create event voice roaming MT Home in "+str(country)+" 60s, 5AM Before next bc","Charged Not IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 60s, 6AM Before next bc","Charged IDR "+str(prepaidPrice)],
                     ["Create event 1 sms onnet, 6AM Before next bc","Charged Not IDR "+str(prepaidPrice)],
                     ["Create event gprs roaming using tapcode HKGHT (MCCMNC : 45404) RG11 with apn telkomsel.r, 6AM Before next bc","Charged Not IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to Brunei (+673) 60s, 6.50AM Before next bc | Rounded should be 1200s","Charged IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 60s, 10PM Before next bc","Charged Not IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 60s, 11PM Before next bc","Charged Not IDR "+str(prepaidPrice)],
                     "Create event voice initial international to "+str(country)+" using access code 007, 12AM  next bc | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 12AM  next bc | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 12AM  next bc | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 12AM  next bc | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 12AM Next bc | GSU should be 180s",
                     "Create event voice intermediate international to "+str(country)+" using access code 007 180s, 12AM Next bc | GSU should be 180s",
                     ["Create event voice intermediate international to "+str(country)+" using access code 007 180s, 12AM Next bc | GSU should be 120s","Intermediate Success | Final Unit Ind 0 | FUI 1"],
                     ["Create event voice terminate international to "+str(country)+" using access code 007 120s, 12AM Next bc","Charged IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 1200s, 0.20AM Next bc","Charged IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 1200s, 11PM Next bc","Charged Not IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 1200s, 10PM on "+datetime.now().strftime("%Y")+"-05-03","Charged IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 1200s, 11PM on "+datetime.now().strftime("%Y")+"-05-03","Charged IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 1200s, 2AM on "+datetime.now().strftime("%Y")+"-12-31","Charged IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 1200s, 11PM on "+datetime.now().strftime("%Y")+"-12-31","Charged IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 1200s, 2AM on "+str(int(datetime.now().strftime("%Y"))+1)+"-01-01","Charged Not IDR "+str(prepaidPrice)],
                     ["Create event voice idd using access code 007 to "+str(country)+" 1200s, 11PM on "+str(int(datetime.now().strftime("%Y"))+1)+"-01-01","Charged Not IDR "+str(prepaidPrice)]
              ]

              # Write Header Row
              headerPrepaid = [f'{eventName} | {country} | Prepaid']
              ws.append(headerPrepaid)

              startColumnIndex = 3  # Example of a dynamic column index
              startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
              endColumn = "E"
              startRow = 1
              endRow = 1
              cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
              ws.merge_cells(cellRange)

              headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
              ws.append(headerRow)

              for no, step in enumerate(stepsPrePaid):
                     no = no+1
                     if isinstance(step, str):
                            row = [
                                   no,
                                   step,
                                   "Success",
                                   "No Bonus",
                                   "XYZ"
                            ]
                     else:
                            row = [
                                   no,
                                   step[0],
                                   step[1],
                                   "No Bonus",
                                   "XYZ"
                            ]
                     ws.append(row)

              #Postpaid Section
              stepsPostPaid = [
                     "Create and actived new KartuHALO Bebas Abonemen",
                     "Create event update parameter Init activation date",
                     "Create event attach offer with param Spending limit offer level and set CLS 10000000 3669354 Roaming",
                     "Create event attach offer international roaming 36327",
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 9.30PM","Charged Not IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 10PM","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice international to  "+str(country)+" 60s, 11PM","Charged not IDR "+str(postpaidPrice)],
                     "Create event voice initial international to  "+str(country)+" using access code 007, 11PM | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 11PM | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 11PM | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 11PM | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 11PM | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 11PM | GSU should be 180s",
                     ["Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 11PM | GSU should be 120s","Intermediate Success | Final Unit Ind 0 | FUI 1"],
                     ["Create event voice terminate international to  "+str(country)+" using access code 007 120s, 11PM","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 600s, 11.30PM","Charged Not IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1s, 1AM D+1 | Rounded should be 1200s","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice onnet 60s, 3AM D+1","Charged not IDR "+str(postpaidPrice)],
                     ["Create event voice roaming MO Home from  "+str(country)+" 60s, 5AM D+1","Charged not IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 60s, 7AM D+1","Charged not IDR "+str(postpaidPrice)],
                     "Create event voice initial international to  "+str(country)+" using access code 007, 10.10PM D+1 | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 10.10PM D+1 | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 10.10PM D+1 | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 10.10PM D+1 | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 10.10PM D+1 | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 10.10PM D+1 | GSU should be 180s",
                     ["Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 10.10PM D+1 | GSU should be 120s","Intermediate Success | Final Unit Ind 0 | FUI 1"],
                     ["Create event voice terminate international to  "+str(country)+" using access code 007 120s, 10.10PM D+1","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 600s, 11PM D+1 | Rounded should be 600s","Charged Not IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 001 to  "+str(country)+" 60s, 1AM Before next bc","Charged not IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 008 to  "+str(country)+" 60s, 2AM Before next bc","Charged not IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 01018 to  "+str(country)+" 60s, 3AM Before next bc","Charged not IDR "+str(postpaidPrice)],
                     ["Create event voice offnet 60s, 3AM Before next bc","Charged not IDR "+str(postpaidPrice)],
                     ["Create event voice roaming MT Home in  "+str(country)+" 60s, 5AM Before next bc","Charged not IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 60s, 6AM Before next bc","Charged IDR "+str(postpaidPrice)],
                     ["Create event 1 sms onnet, 6AM Before next bc","Charged not IDR "+str(postpaidPrice)],
                     ["Create event gprs roaming using tapcode HKGHT (MCCMNC : 45404) RG11 with apn telkomsel.r, 6AM Before next bc","Charged not IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to Brunei (+673) 60s, 6.50AM Before next bc | Rounded should be 1200s","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 60s, 10PM Before next bc","Charged not IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 60s, 11PM Before next bc","Charged not IDR "+str(postpaidPrice)],
                     "Create event voice initial international to  "+str(country)+" using access code 007, 12AM  next bc | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 12AM  next bc | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 12AM  next bc | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 12AM  next bc | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 12AM Next bc | GSU should be 180s",
                     "Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 12AM Next bc | GSU should be 180s",
                     ["Create event voice intermediate international to  "+str(country)+" using access code 007 180s, 12AM Next bc | GSU should be 120s","Intermediate Success | Final Unit Ind 0 | FUI 1"],
                     ["Create event voice terminate international to  "+str(country)+" using access code 007 120s, 12AM Next bc","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 0.20AM Next bc","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 11PM Next bc","Charged Not IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 10PM on "+datetime.now().strftime("%Y")+"-05-03","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 11PM on "+datetime.now().strftime("%Y")+"-05-03","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 10PM on "+datetime.now().strftime("%Y")+"-06-03","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 11PM on "+datetime.now().strftime("%Y")+"-06-03","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 2AM on "+datetime.now().strftime("%Y")+"-07-03","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 11PM on "+datetime.now().strftime("%Y")+"-07-03","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 2AM on "+datetime.now().strftime("%Y")+"-08-03","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 11PM on "+datetime.now().strftime("%Y")+"-08-03","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 2AM on "+datetime.now().strftime("%Y")+"-09-03","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 11PM on "+datetime.now().strftime("%Y")+"-09-03","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 2AM on "+datetime.now().strftime("%Y")+"-10-03","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 11PM on "+datetime.now().strftime("%Y")+"-10-03","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 2AM on "+datetime.now().strftime("%Y")+"-11-03","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 11PM on "+datetime.now().strftime("%Y")+"-11-03","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 2AM on "+datetime.now().strftime("%Y")+"-12-31","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 11PM on "+datetime.now().strftime("%Y")+"-12-31","Charged IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 2AM on "+str(int(datetime.now().strftime("%Y"))+1)+"-01-01","Charged not IDR "+str(postpaidPrice)],
                     ["Create event voice idd using access code 007 to  "+str(country)+" 1200s, 11PM on "+str(int(datetime.now().strftime("%Y"))+1)+"-01-01","Charged not IDR "+str(postpaidPrice)],
                     ["Check INDIRA PRE","Checked"],
                     ["Check INDIRA POST","Checked"],

              ]

              # Write Header Row
              headerPostpaid = [f'{eventName} | {country} | Postpaid']
              ws.append(headerPostpaid)

              # Merge Header Cells
              startColumnIndex = 3  # Example of a dynamic column index
              startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
              endColumn = "E"
              startRow = 1
              endRow = 1
              cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
              ws.merge_cells(cellRange)
                     

              headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
              ws.append(headerRow)

              for no, step in enumerate(stepsPostPaid):
                     no = no+1
                     if isinstance(step, str):
                            row = [
                                   no,
                                   step,
                                   "Success",
                                   "No Bonus",
                                   "XYZ"
                            ]
                     else:
                            row = [
                                   no,
                                   step[0],
                                   step[1],
                                   "No Bonus",
                                   "XYZ"
                            ]
                     ws.append(row)

       print("Testing Scenario Successfully Generated")
       
       # Save Excel File
       wb.save('Result/Scenario '+str(eventName)+'.xlsx')

def exportExcelEmptyOffer(eventName, params=None, neededParams = None):
       # Export Test Cases to Excel File
       wb = Workbook()
       ws = wb.active
       for params in params:
              if "OfferName" in params:
                     offerName = params['OfferName']
              else:
                     offerName = ''
              
              if offerName == '':
                     continue
              
              if "OfferDesc" in params:
                     offerDesc = params['OfferDesc']
              else:
                     offerDesc = ''

              steps = [
                     ["Create & Activate new subscriber PP KartuHalo Bebas Abonemen","Check active period"],
                     "Update Parameter (Init activation date)",
                     "Set New Credit Limit Service (offer id : 3669334) as 10000000",
                     ["Attach Offer "+str(offerName)+"","Offer Attached"],
                     ["Check Offer Name & Description",""+str(offerName)+"|"+str(offerDesc)+""],
                     ["Check 888","Checked"],
                     ["Check 889*1","Checked"],
                     ["Check 889*2","Checked"],
                     ["Check 889*3","Checked"],
                     ["Check 889*4","Checked"],
                     ["Check Bonus Info","Checked"],
                     ["Create event 10 SMS Onnet, 1PM","Charged"],
                     ["Create event 10 SMS Offnet, 11PM","Charged"],
                     ["Create event 10 SMS FWA, D+1 6AM","Charged"],
                     ["Create event SMS International to Malaysia (+61), D+2 11AM","Charged"],
                     ["Create event voice onnet 600s, D+2 11AM","Charged"],
                     ["Create event voice offnet 600s, D+2 11AM","Charged"],
                     ["Create event GPRS 10MB using RG55, D+2 11AM","Charged"],
                     ["Create event direct debit using vascode google with charge 50K, D+2 11AM","Charged"],
                     ["Ceate event voice PSTN 600s before next bc 11AM","Charged"],
                     ["Create event voice FWA 600 before next bc 10PM","Charged"],
                     ["Check bonus before next bc","Checked"],
                     ["Check cycle month","Checked"],
                     ["Check bonus after next bc","Çhecked"],
                     ["Create event 10 SMS Onnet after next bc","Charged"],
                     ["Check cycle month","Checked"],
                     ["Create event GPRS 1MB using RG55 after next bc","Charged"],
                     ["Create event direct debit using vascode google with charge 50K after next bc","Charged"],
                     ["Check INDIRA","Checked"],
                     ["Create & Activate new subscriber PP Hybrid","Check active period"],
                     "Update Parameter (Init activation date)",
                     "Set New Credit Limit Service (offer id : 3669334) as 10000000",
                     ["Attach Offer "+str(offerName)+"","Offer Attached"],
                     ["Attach Offer "+str(offerName)+"","Offer Attached"],
                     ["Attach Offer "+str(offerName)+"","Offer Attached"],
                     ["Attach Offer "+str(offerName)+"","Offer Attached"],
                     ["Attach Offer "+str(offerName)+"","Offer Attached"],
                     ["Attach Offer "+str(offerName)+"","Offer Attached"],
                     ["Check All Offer","Should be 6 offers populated"],
                     ["Check 888","Checked"],
                     ["Check 889","Checked"],
                     ["Create event 100 SMS Onnet, 1PM","Charged"],
                     ["Create event 100 SMS Offnet, 11PM","Charged"],
                     ["Create event voice onnet 6000s D+1 6AM","Charged"],
                     ["Create event voice PSTN 6000s D+1 6AM","Charged"],
                     ["Create event SMS International to Malaysia (+61), D+2 11AM","Charged"],
                     ["Create event direct debit using vascode google with charge 50K, D+2 11AM","Charged"],
                     ["Check cycle month","Checked"],
                     ["Check INDIRA","Checked"],
              ]

              # Write Header Row
              header = [f'{eventName} | {offerName} | {offerDesc}']
              ws.append(header)

              # Merge Header Cells
              startColumnIndex = 3  # Example of a dynamic column index
              startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
              endColumn = "E"
              startRow = 1
              endRow = 1
              cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
              ws.merge_cells(cellRange)

              headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
              ws.append(headerRow)

              for no, step in enumerate(steps):
                     no = no+1
                     if isinstance(step, str):
                            row = [
                                   no,
                                   step,
                                   "Success",
                                   "No Bonus",
                                   "XYZ"
                            ]
                     else:
                            row = [
                                   no,
                                   step[0],
                                   step[1],
                                   "No Bonus",
                                   "XYZ"
                            ]
                     ws.append(row)

       print("Testing Scenario Successfully Generated")
       
       # Save Excel File
       wb.save('Result/Scenario '+str(eventName)+'.xlsx')

def exportExcelRcOffer(eventName, params=None, neededParams = None):
       RCType = ''
       for data in params:
              if "RC Type" in data:
                     RCType = data["RC Type"][0]
       # Export Test Cases to Excel File
       wb = Workbook()
       ws = wb.active
       Quota         = []
       pritName      = ''
       proration     = ''
       bonusDescription = ''
       for params in params:
              if "OfferName" in params:
                     offerName = params['OfferName']
              else:
                     offerName = ''
              
              if offerName == '':
                     continue
              
              if "OfferDesc" in params:
                     offerDesc = params['OfferDesc']
              else:
                     offerDesc = ''
              
              if "Rate" in params:
                     rate = params['Rate']
              else:
                     rate = ''
              
              if "Proration" in params:
                     proration = params['Proration']
                     if re.search('Non Prorate|Non Proration', proration):
                            strProration = ''
                            strResultProration = "| "+str(rate)
                     elif re.search('Prorate|Proration', proration):
                            strProration = "| Proration"
                            strResultProration = "| "+str(rate)+" (Prorate)"
                     else:
                            strProration = ''
                            strResultProration = ''
              else:
                     proration = ''
              
              if "AMDD Charge Code" in params:
                     chargeCode = params['AMDD Charge Code']
              else:
                     chargeCode = ''
              
              if "Prit Name" in params:
                     pritName = params["Prit Name"]
              
              if "Quota" in params:
                     Quota = params["Quota"]
              
              if "Bonus Description" in params:
                     bonusDescription = params["Bonus Description"][0]

              if RCType == 'Charge + Allowance':
                     steps = getStepsForRCOffer(RCType, offerName, offerDesc, rate, chargeCode, strProration, strResultProration, pritName, Quota, bonusDescription)
              else:
                     steps = [
                            ["Create and actived new subscriber PP KartuHALO Bebas Abonemen","Check active period"],
                            ["Update parameter Init activation date","Success"],
                            ["Set New Credit Limit Service (offer id : 3669334) as 10.000.000","Success"],
                            ["Attach Offer "+str(offerName)+" | 3919479","Offer Attached"],
                            ["Check Offer Name & Description",""+str(offerName)+"|"+str(offerDesc)+""],
                            ["Check 888 "+strProration,"Checked "+strResultProration],
                            ["Check RC AMDD Charge Code","Checked | "+str(chargeCode)],
                            ["Check Bonus 889 ","Checked"],
                            ["Check I9getBonusInfo","Checked"],
                            ["Create event voice onnet 60s 11AM","Charged"],
                            ["Create event voice offnet 60s 1PM","Charged"],
                            ["Create event voice pstn 60s 3PM","Charged"],
                            ["Create event voice fwa 60s 5PM D+1","Charged"],
                            ["Create event 1 sms onnet 7PM D+1","Charged"],
                            ["Create event 1 sms offnet 9PM D+3","Charged"],
                            ["Create event 1 sms fwa 11PM D+3","Charged"],
                            ["Create event GPRS 1MB RG 55 1AM D+7","Charged"],
                            ["Create event MMS 5AM D+7","Charged"],
                            ["Create event Direct Debit using Vascode bank_digi_250 8AM D+7","Charged"],
                            ["Check Bonus next BC ","Checked"],
                            ["Check Cycle Month","Checked"],
                            ["Check table TRB1_Subs_Errs","Should be no errors"],
                            ["Check Indira","Checked"],
                            ["Invoicing","Success"],
                            ["Create and actived new KartuHalo Hybrid Instant V2","Check active period"],
                            ["Update parameter Init activation date","Success"],
                            ["Set New Credit Limit Service (offer id : 3669334) as 10.000.000","Success"],
                            ["Attach Offer "+str(offerName)+" | 3919479","Offer Attached"],
                            ["Attach Offer "+str(offerName)+" | 3919479","Offer Attached"],
                            ["Attach Offer "+str(offerName)+" | 3919479","Offer Attached"],
                            ["Attach Offer "+str(offerName)+" | 3919479","Offer Attached"],
                            ["Attach Offer "+str(offerName)+" | 3919479","Offer Attached"],
                            ["Attach Offer "+str(offerName)+" | 3919479","Offer Attached"],
                            ["Check 888 "+strProration,"Checked "+strResultProration],
                            ["Check All Offer on DB","Should be 6 offers populated"],
                            ["Create event Direct Debit using Vascode bank_digi_250 8AM D+7","Charged"],
                            ["Check Cycle Month","Checked"],
                            ["Check table TRB1_Subs_Errs","Should be no errors"],
                            ["Check Indira","Checked"],
                            ["Invoicing","Success"],
                     ]

              # Write Header Row
              header = [f'{eventName} | {offerName} | {offerDesc}']
              ws.append(header)

              # Merge Header Cells
              startColumnIndex = 3  # Example of a dynamic column index
              startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
              endColumn = "E"
              startRow = 1
              endRow = 1
              cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
              ws.merge_cells(cellRange)

              headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
              ws.append(headerRow)

              for no, step in enumerate(steps):
                     no = no+1
                     if isinstance(step, str):
                            row = [
                                   no,
                                   step,
                                   "Success",
                                   "No Bonus",
                                   "XYZ"
                            ]
                     else:
                            if step is None:
                                   continue
                            else:
                                   if len(step) == 5:
                                          row = [
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 step[3],
                                                 step[4]
                                          ]
                                   
                                   elif len(step) == 4:
                                          row = [
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 step[3],
                                                 "XYZ"
                                          ] 
                                   elif len(step) == 3:
                                          row = [
                                                 no,
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 "XYZ"
                                          ]
                                          no = no+1
                                   else:
                                          row = [
                                                 no,
                                                 step[0],
                                                 step[1],
                                                 "No Bonus",
                                                 "XYZ"
                                          ]
                                          no = no+1
                     ws.append(row)

       print("Testing Scenario Successfully Generated")
       
       # Save Excel File
       wb.save('Result/Scenario '+str(eventName)+' '+str(RCType)+'.xlsx')

def exportExcelOCOffer(eventName, params=None, neededParams = None):
       # Export Test Cases to Excel File
       OCType = ''
       for data in params:
              if "OC Type" in data:
                     OCType = data["OC Type"][0]

       wb = Workbook()
       ws = wb.active
       Quota         = []
       pritName      = ''
       proration     = ''
       bonusDescription = ''
       for params in params:
              if "OfferName" in params:
                     offerName = params['OfferName']
              else:
                     offerName = ''

              if offerName == '':
                     continue
              
              if "OfferDesc" in params:
                     offerDesc = params['OfferDesc']
              else:
                     offerDesc = ''

              if "Rate" in params:
                     rate = params['Rate']
              else:
                     rate = ''

              if "AMDD Charge Code" in params:
                     chargeCode = params['AMDD Charge Code']
              else:
                     chargeCode = ''

              if "Commitment Period" in params:
                     commitmentPeriod = f" with Commitment period {params['Commitment Period']} months"
              else:
                     commitmentPeriod = " without Commitment period"
              
              if "Prit Name" in params:
                     pritName = params["Prit Name"]
              
              if "Quota" in params:
                     Quota = params["Quota"]
              
              if "Proration" in params:
                     proration = params['Proration']
                     if re.search('Non Prorate|Non Proration', proration):
                            strProration = ''
                            strResultProration = "| "+str(rate)
                     elif re.search('Prorate|Proration', proration):
                            strProration = "| Proration"
                            strResultProration = "| "+str(rate)+" (Prorate)"
                     else:
                            strProration = ''
                            strResultProration = ''
              
              if "Bonus Description" in params:
                     bonusDescription = params["Bonus Description"][0]
              
              if OCType == "Penalty + Allowance" or OCType == "Charge + Allowance" or OCType == "Charge Installment + Allowance":
                     steps = getStepsForOCOffer(OCType, offerName, offerDesc, rate, chargeCode, strProration, strResultProration, pritName, Quota, commitmentPeriod, bonusDescription)
              else:
                     steps = [
                            ["Create & Activate PP Kartu Halo Bebas Abonemen V2","Number Activated"],
                            ["Update Parameter","Success"],
                            ["Attach Offer CLS (Credit Limit Service) 2000000 IDR","Offer Attached"],
                            ["Attach Offer "+str(offerName)+str(commitmentPeriod),"Offer not Attached"],
                            ["Check Offer name and offer description",""+str(offerName)+"|"+str(offerDesc)+""],
                            ["Check 888",f'Charged | {rate}'],
                            ["Create event voice onnet 60s","Charged"],
                            ["Create event 10 SMS onnet ","Charged"],
                            ["Create event GPRS 1MB RG50","Charged"],
                            ["Check RB Log","Checked"],
                            ["Check Rated Event","Checked"],
                            ["Check AMDD Charge code OC",chargeCode],
                            ["Check error from table TRB1_Subs_errs","No errors"],
                            "Invoicing",
                            ["Create & Activate PP Kartu Halo Bebas Abonemen V2","Number Activated"],
                            ["Update Parameter","Success"],
                            ["Attach Offer CLS (Credit Limit Service) 2000000 IDR","Offer Attached"],
                            ["Attach Offer "+str(offerName)+str(commitmentPeriod),"Offer Attached"],
                            ["Attach Offer "+str(offerName)+str(commitmentPeriod),"Offer Attached"],
                            ["Attach Offer "+str(offerName)+str(commitmentPeriod),"Offer Attached"],
                            ["Attach Offer "+str(offerName)+str(commitmentPeriod),"Offer Attached"],
                            ["Attach Offer "+str(offerName)+str(commitmentPeriod),"Offer Attached"],
                            ["Attach Offer "+str(offerName)+str(commitmentPeriod),"Offer Attached"],
                            ["Check Offer name and offer description",""+str(offerName)+"|"+str(offerDesc)+""],
                            ["Check 888",f"Checked | {rate}"],
                            ["Create event voice onnet 60s","Charged"],
                            ["Create event 10 SMS onnet ","Charged"],
                            ["Create event GPRS 1MB RG50","Charged"],
                            ["Check RB Log","Checked"],
                            ["Check Rated Event","Checked"],
                            ["Check AMDD Charge code OC",chargeCode],
                            ["Check error from table TRB1_Subs_errs","No errors"],
                            "Invoicing",
                     ]

                     if OCType == 'Penalty':
                            if "AMDD Charge Code Remove Offer" in params:
                                   chargeCodeRemove = params['AMDD Charge Code Remove Offer']
                            else:
                                   chargeCodeRemove = chargeCode

                            additionalStep = [
                                   [f"Remove offer {offerName} on D+1","Offer Removed"],
                                   ["Check AMDD Charge code OC",f"{chargeCodeRemove}"]
                            ]

                            insert_index = 12

                            for addStep in additionalStep[::-1]:
                                   steps.insert(insert_index, addStep)

              # Write Header Row
              header = [f'{eventName} | {offerName} | {offerDesc}']
              ws.append(header)

              # Merge Header Cells
              startColumnIndex = 3  # Example of a dynamic column index
              startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
              endColumn = "E"
              startRow = 1
              endRow = 1
              cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
              ws.merge_cells(cellRange)

              headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
              ws.append(headerRow)

              for no, step in enumerate(steps):
                     no = no+1
                     if isinstance(step, str):
                            row = [
                                   no,
                                   step,
                                   "Success",
                                   "No Bonus",
                                   "XYZ"
                            ]
                     else:
                            if step is None:
                                   continue
                            else:
                                   if len(step) == 5:
                                          row = [
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 step[3],
                                                 step[4]
                                          ]
                                   
                                   elif len(step) == 4:
                                          row = [
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 step[3],
                                                 "XYZ"
                                          ] 
                                   elif len(step) == 3:
                                          row = [
                                                 no,
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 "XYZ"
                                          ]
                                          no = no+1
                                   else:
                                          row = [
                                                 no,
                                                 step[0],
                                                 step[1],
                                                 "No Bonus",
                                                 "XYZ"
                                          ]
                                          no = no+1
                     ws.append(row)

       print("Testing Scenario Successfully Generated")
       
       # Save Excel File
       wb.save('Result/Scenario '+str(eventName)+' '+str(OCType)+'.xlsx')

def getStepsForOCOffer(OCType, offerName, offerDesc, rate, chargeCode, strProration, strResultProration, pritName, Quota, commitmentPeriod, bonusDescription):
       QuotaSplitString     = Quota.split(';')
       QuotaSMS             = 0
       firstQuotaSMS        = 0
       QuotaVoice           = int(QuotaSplitString[0])
       firstQuotaVoice      = int(QuotaSplitString[0])
       QuotaString          = ''
       lastQuota            = ''
       stepNextBCSMS        = None
       stepNextBCVoice      = None

       if bonusDescription == 'All Opr':
              stringBonus = "All Opr"
       elif bonusDescription == 'Tsel (Onnet, Onbrand for Loop)':
              stringBonus = "Tsel"
       else:
              stringBonus = "Opr Lain"

       if OCType == "Penalty + Allowance" or OCType == "Charge Installment + Allowance":
              stringInstallment = commitmentPeriod
       else:
              stringInstallment = ''
       

       if len(QuotaSplitString) > 1:
              QuotaSMS             = int(QuotaSplitString[1])
              firstQuotaSMS        = int(QuotaSplitString[1])
       
       if QuotaVoice > 0:
              tenPercentVoice = round(QuotaVoice*0.1)
              useTenPercentVoice = round(QuotaVoice-tenPercentVoice)
              QuotaString += stringBonus + ' ' + str(QuotaVoice) + " minutes"

              if QuotaSMS > 0:
                     stringQuotaSMS = ", " + stringBonus + ' ' + str(QuotaSMS) + " sms"
              else:
                     stringQuotaSMS = ''

              if bonusDescription == 'All Opr':
                     event_descriptions = [
                            [
                                   "voice onnet paramVoice minutes 11am",
                                   0
                            ],
                            [
                                   "voice offnet paramVoice minutes 1pm",
                                   0
                            ],
                            [
                                   "voice fwa paramVoice minutes 8am",
                                   0
                            ],
                            [
                                   "voice pstn paramVoice minutes 11am",
                                   0
                            ],
                            [
                                   "voice paramVoice minutes voice onnet 3am",
                                   0
                            ],
                            [
                                   "voice paramVoice minutes voice pstn 10am",
                                   0
                            ],
                            [
                                   "voice offnet paramVoice minutes",
                                   0
                            ],
                            [
                                   "voice onnet paramVoice minutes using ServiceSubType 21",
                                   0
                            ]
                     ]
              elif bonusDescription == 'Tsel (Onnet, Onbrand for Loop)':
                     event_descriptions = [
                            [
                                   "voice onnet paramVoice minutes 11am",
                                   0
                            ],
                            [
                                   "voice offnet paramVoice minutes 1pm",
                                   1
                            ],
                            [
                                   "voice fwa paramVoice minutes 8am",
                                   1
                            ],
                            [
                                   "voice pstn paramVoice minutes 11am",
                                   1
                            ],
                            [
                                   "voice paramVoice minutes voice onnet 3am",
                                   0
                            ],
                            [
                                   "voice paramVoice minutes voice pstn 10am",
                                   1
                            ],
                            [
                                   "voice offnet paramVoice minutes",
                                   1
                            ],
                            [
                                   "voice onnet paramVoice minutes using ServiceSubType 21",
                                   0
                            ]
                     ]
              elif bonusDescription == 'Opr Lain (Include fwa,pstn)':
                     event_descriptions = [
                            [
                                   "voice onnet paramVoice minutes 11am",
                                   1
                            ],
                            [
                                   "voice offnet paramVoice minutes 1pm",
                                   0
                            ],
                            [
                                   "voice fwa paramVoice minutes 8am",
                                   0
                            ],
                            [
                                   "voice pstn paramVoice minutes 11am",
                                   0
                            ],
                            [
                                   "voice paramVoice minutes voice onnet 3am",
                                   1
                            ],
                            [
                                   "voice paramVoice minutes voice pstn 10am",
                                   0
                            ],
                            [
                                   "voice offnet paramVoice minutes",
                                   0
                            ],
                            [
                                   "voice onnet paramVoice minutes using ServiceSubType 21",
                                   1
                            ]
                     ]
              else:
                     event_descriptions = [
                            [
                                   "voice onnet paramVoice minutes 11am",
                                   1
                            ],
                            [
                                   "voice offnet paramVoice minutes 1pm",
                                   0
                            ],
                            [
                                   "voice fwa paramVoice minutes 8am",
                                   1
                            ],
                            [
                                   "voice pstn paramVoice minutes 11am",
                                   1
                            ],
                            [
                                   "voice paramVoice minutes voice onnet 3am",
                                   1
                            ],
                            [
                                   "voice paramVoice minutes voice pstn 10am",
                                   1
                            ],
                            [
                                   "voice offnet paramVoice minutes",
                                   0
                            ],
                            [
                                   "voice onnet paramVoice minutes using ServiceSubType 21",
                                   1
                            ]
                     ]
              
              stepsConsumeVoice = []
              for desc in event_descriptions:
                     description   = desc[0]
                     validating    = desc[1]
                     if validating == 0:
                            if QuotaVoice >= 0:
                                   decreasingQuotaVoice = round((QuotaVoice*0.5)/4)
                                   QuotaVoice -= decreasingQuotaVoice
                                   step = [
                                   f"Create event {description.replace('paramVoice', str(decreasingQuotaVoice))}",
                                   "Consume Bonus",
                                   f"{stringBonus} {QuotaVoice} minutes{stringQuotaSMS}"
                                   ]
                                   lastQuota = f"{stringBonus} {QuotaVoice} minutes{stringQuotaSMS}"
                                   stepsConsumeVoice.append(step)
                     else:
                            step = [
                                   f"Create event {description.replace('paramVoice', '1')}",
                                   "Charged",
                                   f"{stringBonus} {QuotaVoice} minutes{stringQuotaSMS}"
                            ]
                            lastQuota = f"{stringBonus} {QuotaVoice} minutes{stringQuotaSMS}"
                            stepsConsumeVoice.append(step)
              stepNextBCVoice = ["Create event voice onnet "+str(tenPercentVoice)+" minutes after next bc","Consume Bonus",stringBonus + ' ' + str(useTenPercentVoice)+" minutes, "+ stringBonus + ' ' + str(firstQuotaSMS)+" sms"]

       if QuotaSMS > 0:
              tenPercentSMS = round(QuotaSMS*0.1)
              useTenPercentSMS = round(QuotaSMS-tenPercentSMS)
              QuotaString += ", "+ stringBonus + ' ' + str(QuotaSMS) +" sms"

              if QuotaVoice > 0:
                     stringQuotaVoice = stringBonus + ' ' + str(QuotaVoice) + " minutes,"
              else:
                     stringQuotaVoice = ''

              if bonusDescription == 'All Opr':
                     event_descriptions = [
                            [
                                   "paramSMS sms onnet 5pm",
                                   0
                            ],
                            [
                                   "paramSMS sms offnet 7pm",
                                   0
                            ],
                            [
                                   "paramSMS sms onnet 1am",
                                   0
                            ],
                            [
                                   "paramSMS sms offnet 1pm",
                                   0
                            ],
                            [
                                   "paramSMS sms onnet",
                                   0
                            ]
                     ]
              elif bonusDescription == 'Tsel (Onnet, Onbrand for Loop)':
                     event_descriptions = [
                            [
                                   "paramSMS sms onnet 5pm",
                                   0
                            ],
                            [
                                   "paramSMS sms offnet 7pm",
                                   1
                            ],
                            [
                                   "paramSMS sms onnet 1am",
                                   0
                            ],
                            [
                                   "paramSMS sms offnet 1pm",
                                   1
                            ],
                            [
                                   "paramSMS sms onnet",
                                   0
                            ]
                     ]
              else:
                    event_descriptions = [
                            [
                                   "paramSMS sms onnet 5pm",
                                   1
                            ],
                            [
                                   "paramSMS sms offnet 7pm",
                                   0
                            ],
                            [
                                   "paramSMS sms onnet 1am",
                                   1
                            ],
                            [
                                   "paramSMS sms offnet 1pm",
                                   0
                            ],
                            [
                                   "paramSMS sms onnet",
                                   1
                            ]
                     ] 

              stepsConsumeSMS = []
              for desc in event_descriptions:
                     description   = desc[0]
                     validating    = desc[1]

                     if validating == 0:
                            if QuotaSMS >= 0:
                                   decreasingQuotaSMS = round((QuotaSMS*0.5)/4)
                                   QuotaSMS -= decreasingQuotaSMS
                                   step = [
                                   f"Create event {description.replace('paramSMS', str(decreasingQuotaSMS))}",
                                   "Consume Bonus",
                                   f"{stringQuotaVoice} {stringBonus} {QuotaSMS} sms"
                                   ]
                                   lastQuota = f"{stringQuotaVoice} {stringBonus} {QuotaSMS} sms"
                                   stepsConsumeSMS.append(step)
                     else:
                            step = [
                                   f"Create event {description.replace('paramSMS', '1')}",
                                   "Charged",
                                   f"{stringQuotaVoice} {stringBonus} {QuotaSMS} sms"
                            ]
                            lastQuota = f"{stringQuotaVoice} {stringBonus} {QuotaSMS} sms"
                            stepsConsumeSMS.append(step)
              stepNextBCSMS = ["Create event "+str(tenPercentSMS)+" sms offnet after next bc","Consume Bonus",stringBonus + ' ' + str(useTenPercentVoice)+" minutes, "+ stringBonus + ' ' + str(useTenPercentSMS)+" sms"]

       steps = []
       firstStep = [
              ["Create & Activate new subscriber PP KartuHalo Bebas Abonemen","Check active period","No Bonus"],
              ["Update Parameter (Init activation date)","SUCCESS","No Bonus"],
              ["Set New Credit Limit Service (offer id : 3669334) as 50.000.000","Offer Attached","No Bonus"],
              ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString],
              ["Check Offer Name & Offer Description",""+str(offerName)+"|"+str(offerDesc)+"",""+str(offerName)+"|"+str(offerDesc)+""],
              ["Check 888",""+str(rate)+" "+str(strProration),QuotaString],
              ["Check 889","Checked",QuotaString],
              ["Check 889*1","Checked","No Bonus"],
              ["Check 889*2","Checked",stringBonus + ' ' + str(firstQuotaVoice) +" minutes"],
              ["Check 889*3","Checked",stringBonus + ' ' + str(firstQuotaSMS)+" sms"],
              ["Check 889*4","Checked","No Bonus"],
              ["Check AMDD and Charge ",""+ str(chargeCode) +" | "+ str(rate)+" "+str(strProration),"No Bonus"],
              ["Check PRIT Name",pritName,"No Bonus"]
       ]
       #Step Consume Quota (Voice & SMS)
       
       secondStep = [
              ["Create event direct debit with vascode google (5500k) 1pm","Charged",lastQuota],
              ["Create event voice international to taiwan 60s 2pm","Charged",lastQuota],
              ["Create event gprs 1Mb rg 55 5pm","Charged",lastQuota],
              ["Check bonus before next bc","Checked",lastQuota],
              ["Check bonus after next bc","Checked",QuotaString],
              stepNextBCVoice,
              stepNextBCSMS,
              ["Check table TRB1_SUB Errs","Checked","No Bonus"],
              ["INVOICING","Checked","No Bonus"],
              ["Create & Activate new subscriber PP KartuHalo Bebas Abonemen","Check active period","No Bonus"],
              ["Update Parameter (Init activation date)","SUCCESS","No Bonus"],
              ["Set New Credit Limit Service (offer id : 3669334) as 50.000.000","Offer Attached","No Bonus"],
              ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString],
              ["Check Offer Name & Offer Description",""+str(offerName)+"|"+str(offerDesc)+"",""+str(offerName)+"|"+str(offerDesc)+""],
              ["Check 888",""+str(rate)+" "+str(strProration)+"|"+str(chargeCode)+"",QuotaString],
              ["Check 889","Checked",QuotaString]
       ]
       #Step Consume Quota (Voice & SMS)
       if OCType == "Penalty + Allowance":
              thirdStep = [
                     ["Remove offer D+1","Offer removed","No Bonus"],
                     ["Check AMDD and Charge ",""+str(chargeCode)+" |"+str(rate)+" "+str(strProration),"No Bonus"],
                     ["Check 888 charge","248417","No Bonus"],
                     ["Check table TRB1_SUB Errs","Checked","No Bonus"],
                     ["INVOICING","Checked","No Bonus"],
                     ["Create & Activate new subscriber PP KartuHalo Bebas Abonemen","Check active period","No Bonus"],
                     ["Update Parameter (Init activation date)","SUCCESS","No Bonus"],
                     ["Set New Credit Limit Service (offer id : 3669334) as 50.000.000","Offer Attached","No Bonus"],
                     ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString],
                     ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString],
                     ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["Check 888",""+str(rate)+" "+str(strProration)+" x 6",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["Check 889","Checked",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["Check on Database","Should be 6 Offers because Postpaid",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["Check table TRB1_SUB Errs","Checked",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["INVOICING","Checked",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
              ]
       else:
              thirdStep = [
                     ["Check AMDD and Charge ",""+str(chargeCode)+" |"+str(rate)+" "+str(strProration),"No Bonus"],
                     ["Check 888 charge","Usage + ("+rate+" "+str(strProration)+") + Penalty","No Bonus"],
                     ["Check table TRB1_SUB Errs","Checked","No Bonus"],
                     ["INVOICING","Checked","No Bonus"],
                     ["Create & Activate new subscriber PP KartuHalo Bebas Abonemen","Check active period","No Bonus"],
                     ["Update Parameter (Init activation date)","SUCCESS","No Bonus"],
                     ["Set New Credit Limit Service (offer id : 3669334) as 50.000.000","Offer Attached","No Bonus"],
                     ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString],
                     ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString],
                     ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["Check 888",""+str(rate)+" "+str(strProration)+" x 6",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["Check 889","Checked",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["Check on Database","Should be 6 Offers because Postpaid",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["Check table TRB1_SUB Errs","Checked",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
                     ["INVOICING","Checked",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
              ]  
       steps.extend(firstStep)
       
       if firstQuotaVoice > 0:
              steps.extend(stepsConsumeVoice)
       
       if firstQuotaSMS > 0:
              steps.extend(stepsConsumeSMS)
       
       steps.extend(secondStep)
       
       if firstQuotaVoice > 0:
              steps.extend(stepsConsumeVoice)
       
       if firstQuotaSMS > 0:
              steps.extend(stepsConsumeSMS)
       
       steps.extend(thirdStep)
       return steps

def getStepsForRCOffer(RCType, offerName, offerDesc, rate, chargeCode, strProration, strResultProration, pritName, Quota, bonusDescription):
       QuotaSplitString     = Quota.split(';')
       QuotaSMS             = 0
       firstQuotaSMS        = 0
       QuotaVoice           = int(QuotaSplitString[0])
       firstQuotaVoice      = int(QuotaSplitString[0])
       QuotaString          = ''
       lastQuota            = ''
       stepNextBCSMS        = None
       stepNextBCVoice      = None
       stringInstallment    = ''

       if bonusDescription == 'All Opr':
              stringBonus = "All Opr"
       elif bonusDescription == 'Tsel (Onnet, Onbrand for Loop)':
              stringBonus = "Tsel"
       else:
              stringBonus = "Opr Lain"

       if len(QuotaSplitString) > 1:
              QuotaSMS             = int(QuotaSplitString[1])
              firstQuotaSMS        = int(QuotaSplitString[1])
       
       if QuotaVoice > 0:
              tenPercentVoice = round(QuotaVoice*0.1)
              useTenPercentVoice = round(QuotaVoice-tenPercentVoice)
              QuotaString += stringBonus + ' ' + str(QuotaVoice) + " minutes"

              if QuotaSMS > 0:
                     stringQuotaSMS = ", " + stringBonus + ' ' + str(QuotaSMS) + " sms"
              else:
                     stringQuotaSMS = ''

              if bonusDescription == 'All Opr':
                     event_descriptions = [
                            [
                                   "voice onnet paramVoice minutes 11am",
                                   0
                            ],
                            [
                                   "voice offnet paramVoice minutes 1pm",
                                   0
                            ],
                            [
                                   "voice fwa paramVoice minutes 8am",
                                   0
                            ],
                            [
                                   "voice pstn paramVoice minutes 11am",
                                   0
                            ],
                            [
                                   "voice paramVoice minutes voice onnet 3am",
                                   0
                            ],
                            [
                                   "voice paramVoice minutes voice pstn 10am",
                                   0
                            ],
                            [
                                   "voice offnet paramVoice minutes",
                                   0
                            ],
                            [
                                   "voice onnet paramVoice minutes using ServiceSubType 21",
                                   0
                            ]
                     ]
              elif bonusDescription == 'Tsel (Onnet, Onbrand for Loop)':
                     event_descriptions = [
                            [
                                   "voice onnet paramVoice minutes 11am",
                                   0
                            ],
                            [
                                   "voice offnet paramVoice minutes 1pm",
                                   1
                            ],
                            [
                                   "voice fwa paramVoice minutes 8am",
                                   1
                            ],
                            [
                                   "voice pstn paramVoice minutes 11am",
                                   1
                            ],
                            [
                                   "voice paramVoice minutes voice onnet 3am",
                                   0
                            ],
                            [
                                   "voice paramVoice minutes voice pstn 10am",
                                   1
                            ],
                            [
                                   "voice offnet paramVoice minutes",
                                   1
                            ],
                            [
                                   "voice onnet paramVoice minutes using ServiceSubType 21",
                                   0
                            ]
                     ]
              elif bonusDescription == 'Opr Lain (Include fwa,pstn)':
                     event_descriptions = [
                            [
                                   "voice onnet paramVoice minutes 11am",
                                   1
                            ],
                            [
                                   "voice offnet paramVoice minutes 1pm",
                                   0
                            ],
                            [
                                   "voice fwa paramVoice minutes 8am",
                                   0
                            ],
                            [
                                   "voice pstn paramVoice minutes 11am",
                                   0
                            ],
                            [
                                   "voice paramVoice minutes voice onnet 3am",
                                   1
                            ],
                            [
                                   "voice paramVoice minutes voice pstn 10am",
                                   0
                            ],
                            [
                                   "voice offnet paramVoice minutes",
                                   0
                            ],
                            [
                                   "voice onnet paramVoice minutes using ServiceSubType 21",
                                   1
                            ]
                     ]
              else:
                     event_descriptions = [
                            [
                                   "voice onnet paramVoice minutes 11am",
                                   1
                            ],
                            [
                                   "voice offnet paramVoice minutes 1pm",
                                   0
                            ],
                            [
                                   "voice fwa paramVoice minutes 8am",
                                   1
                            ],
                            [
                                   "voice pstn paramVoice minutes 11am",
                                   1
                            ],
                            [
                                   "voice paramVoice minutes voice onnet 3am",
                                   1
                            ],
                            [
                                   "voice paramVoice minutes voice pstn 10am",
                                   1
                            ],
                            [
                                   "voice offnet paramVoice minutes",
                                   0
                            ],
                            [
                                   "voice onnet paramVoice minutes using ServiceSubType 21",
                                   1
                            ]
                     ]
              
              stepsConsumeVoice = []
              for desc in event_descriptions:
                     description   = desc[0]
                     validating    = desc[1]
                     if validating == 0:
                            if QuotaVoice >= 0:
                                   decreasingQuotaVoice = round((QuotaVoice*0.5)/4)
                                   QuotaVoice -= decreasingQuotaVoice
                                   step = [
                                   f"Create event {description.replace('paramVoice', str(decreasingQuotaVoice))}",
                                   "Consume Bonus",
                                   f"{stringBonus} {QuotaVoice} minutes{stringQuotaSMS}"
                                   ]
                                   lastQuota = f"{stringBonus} {QuotaVoice} minutes{stringQuotaSMS}"
                                   stepsConsumeVoice.append(step)
                     else:
                            step = [
                                   f"Create event {description.replace('paramVoice', '1')}",
                                   "Charged",
                                   f"{stringBonus} {QuotaVoice} minutes{stringQuotaSMS}"
                            ]
                            lastQuota = f"{stringBonus} {QuotaVoice} minutes{stringQuotaSMS}"
                            stepsConsumeVoice.append(step)
              stepNextBCVoice = ["Create event voice onnet "+str(tenPercentVoice)+" minutes after next bc","Consume Bonus",stringBonus + ' ' + str(useTenPercentVoice)+" minutes, "+ stringBonus + ' ' + str(firstQuotaSMS)+" sms"]

       if QuotaSMS > 0:
              tenPercentSMS = round(QuotaSMS*0.1)
              useTenPercentSMS = round(QuotaSMS-tenPercentSMS)
              QuotaString += ", "+ stringBonus + ' ' + str(QuotaSMS) +" sms"

              if QuotaVoice > 0:
                     stringQuotaVoice = stringBonus + ' ' + str(QuotaVoice) + " minutes,"
              else:
                     stringQuotaVoice = ''

              if bonusDescription == 'All Opr':
                     event_descriptions = [
                            [
                                   "paramSMS sms onnet 5pm",
                                   0
                            ],
                            [
                                   "paramSMS sms offnet 7pm",
                                   0
                            ],
                            [
                                   "paramSMS sms onnet 1am",
                                   0
                            ],
                            [
                                   "paramSMS sms offnet 1pm",
                                   0
                            ],
                            [
                                   "paramSMS sms onnet",
                                   0
                            ]
                     ]
              elif bonusDescription == 'Tsel (Onnet, Onbrand for Loop)':
                     event_descriptions = [
                            [
                                   "paramSMS sms onnet 5pm",
                                   0
                            ],
                            [
                                   "paramSMS sms offnet 7pm",
                                   1
                            ],
                            [
                                   "paramSMS sms onnet 1am",
                                   0
                            ],
                            [
                                   "paramSMS sms offnet 1pm",
                                   1
                            ],
                            [
                                   "paramSMS sms onnet",
                                   0
                            ]
                     ]
              else:
                    event_descriptions = [
                            [
                                   "paramSMS sms onnet 5pm",
                                   1
                            ],
                            [
                                   "paramSMS sms offnet 7pm",
                                   0
                            ],
                            [
                                   "paramSMS sms onnet 1am",
                                   1
                            ],
                            [
                                   "paramSMS sms offnet 1pm",
                                   0
                            ],
                            [
                                   "paramSMS sms onnet",
                                   1
                            ]
                     ] 

              stepsConsumeSMS = []
              for desc in event_descriptions:
                     description   = desc[0]
                     validating    = desc[1]

                     if validating == 0:
                            if QuotaSMS >= 0:
                                   decreasingQuotaSMS = round((QuotaSMS*0.5)/4)
                                   QuotaSMS -= decreasingQuotaSMS
                                   step = [
                                   f"Create event {description.replace('paramSMS', str(decreasingQuotaSMS))}",
                                   "Consume Bonus",
                                   f"{stringQuotaVoice} {stringBonus} {QuotaSMS} sms"
                                   ]
                                   lastQuota = f"{stringQuotaVoice} {stringBonus} {QuotaSMS} sms"
                                   stepsConsumeSMS.append(step)
                     else:
                            step = [
                                   f"Create event {description.replace('paramSMS', '1')}",
                                   "Charged",
                                   f"{stringQuotaVoice} {stringBonus} {QuotaSMS} sms"
                            ]
                            lastQuota = f"{stringQuotaVoice} {stringBonus} {QuotaSMS} sms"
                            stepsConsumeSMS.append(step)
              stepNextBCSMS = ["Create event "+str(tenPercentSMS)+" sms offnet after next bc","Consume Bonus",stringBonus + ' ' + str(useTenPercentVoice)+" minutes, "+ stringBonus + ' ' + str(useTenPercentSMS)+" sms"]

       steps = []
       firstStep = [
              ["Create & Activate new subscriber PP KartuHalo Bebas Abonemen","Check active period","No Bonus"],
              ["Update Parameter (Init activation date)","SUCCESS","No Bonus"],
              ["Set New Credit Limit Service (offer id : 3669334) as 50.000.000","Offer Attached","No Bonus"],
              ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString],
              ["Check Offer Name & Offer Description",""+str(offerName)+"|"+str(offerDesc)+"",""+str(offerName)+"|"+str(offerDesc)+""],
              ["Check 888",""+str(rate)+" "+str(strProration),QuotaString],
              ["Check 889","Checked",QuotaString],
              ["Check 889*1","Checked","No Bonus"],
              ["Check 889*2","Checked",stringBonus + ' ' + str(firstQuotaVoice) +" minutes"],
              ["Check 889*3","Checked",stringBonus + ' ' + str(firstQuotaSMS)+" sms"],
              ["Check 889*4","Checked","No Bonus"],
              ["Check AMDD and Charge ",""+ str(chargeCode) +" | "+ str(rate)+" "+str(strProration),"No Bonus"],
              ["Check PRIT Name",pritName,"No Bonus"]
       ]
       #Step Consume Quota (Voice & SMS)
       
       secondStep = [
              ["Create event direct debit with vascode google (5500k) 1pm","Charged",lastQuota],
              ["Create event voice international to taiwan 60s 2pm","Charged",lastQuota],
              ["Create event gprs 1Mb rg 55 5pm","Charged",lastQuota],
              ["Check bonus before next bc","Checked",lastQuota],
              ["Check bonus after next bc","Checked",QuotaString],
              stepNextBCVoice,
              stepNextBCSMS,
              ["Check table TRB1_SUB Errs","Checked","No Bonus"],
              ["INVOICING","Checked","No Bonus"],
              ["Create & Activate new subscriber PP KartuHalo Bebas Abonemen","Check active period","No Bonus"],
              ["Update Parameter (Init activation date)","SUCCESS","No Bonus"],
              ["Set New Credit Limit Service (offer id : 3669334) as 50.000.000","Offer Attached","No Bonus"],
              ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString],
              ["Check Offer Name & Offer Description",""+str(offerName)+"|"+str(offerDesc)+"",""+str(offerName)+"|"+str(offerDesc)+""],
              ["Check 888",""+str(rate)+" "+str(strProration)+"|"+str(chargeCode)+"",QuotaString],
              ["Check 889","Checked",QuotaString]
       ]
       #Step Consume Quota (Voice & SMS)
       thirdStep = [
              ["Check AMDD and Charge ",""+str(chargeCode)+" |"+str(rate)+" "+str(strProration),"No Bonus"],
              ["Check 888 charge","Usage + ("+rate+" "+str(strProration)+") + Penalty","No Bonus"],
              ["Check table TRB1_SUB Errs","Checked","No Bonus"],
              ["INVOICING","Checked","No Bonus"],
              ["Create & Activate new subscriber PP KartuHalo Bebas Abonemen","Check active period","No Bonus"],
              ["Update Parameter (Init activation date)","SUCCESS","No Bonus"],
              ["Set New Credit Limit Service (offer id : 3669334) as 50.000.000","Offer Attached","No Bonus"],
              ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString],
              ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString],
              ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString+"; "+QuotaString],
              ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
              ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
              ["Attach Offer with "+str(offerName)+" "+stringInstallment,"Offer Attached",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
              ["Check 888",""+str(rate)+" "+str(strProration)+" x 6",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
              ["Check 889","Checked",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
              ["Check on Database","Should be 6 Offers because Postpaid",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
              ["Check table TRB1_SUB Errs","Checked",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
              ["INVOICING","Checked",QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString+"; "+QuotaString],
       ]  
       steps.extend(firstStep)
       
       if firstQuotaVoice > 0:
              steps.extend(stepsConsumeVoice)
       
       if firstQuotaSMS > 0:
              steps.extend(stepsConsumeSMS)
       
       steps.extend(secondStep)
       
       if firstQuotaVoice > 0:
              steps.extend(stepsConsumeVoice)
       
       if firstQuotaSMS > 0:
              steps.extend(stepsConsumeSMS)
       
       steps.extend(thirdStep)
       return steps

def exportExcelDiscountOffer(eventName, params=None, neededParams = None):
       # Export Test Cases to Excel File
       wb = Workbook()
       ws = wb.active
       for params in params:
              if "Offer Name" in params:
                     offerName = params['Offer Name']
              else:
                     offerName = ''

              if "Offer Description" in params:
                     offerDescription = params['Offer Description']
              else:
                     offerDescription = ''
              
              if "Rate/Discount Value" in params:
                     rate = params['Rate/Discount Value']
              else:
                     rate = ''

              if "Rate Offer OC" in params:
                     offerNameOC = params['Rate Offer OC'].split(",")
                     if len(offerNameOC) > 1:
                            offerNameOCCase1 = offerNameOC[0]
                            offerNameOCCase2 = offerNameOC[1]
                            offerNameOCCase3 = offerNameOC[2]
                     else:
                            offerNameOCCase1 = offerNameOC
                            offerNameOCCase2 = offerNameOC
                            offerNameOCCase3 = offerNameOC
              else:
                     offerNameOC = ""
                     offerNameOCCase1 = rate
                     offerNameOCCase2 = float(rate) * 0.05
                     offerNameOCCase3 = rate

              if "Rate Offer RC" in params:
                     offerNameRC = params['Rate Offer RC'].split(",")
                     if len(offerNameRC) > 1:
                            offerNameRCCase1 = offerNameRC[0]
                            offerNameRCCase2 = offerNameRC[1]
                            offerNameRCCase3 = offerNameRC[2]
                     else:
                            offerNameRCCase1 = offerNameRC
                            offerNameRCCase2 = offerNameRC
                            offerNameRCCase3 = offerNameRC
              else:
                     offerNameRC = ""
                     offerNameRCCase1 = rate
                     offerNameRCCase2 = float(rate) * 0.05
                     offerNameRCCase3 = rate

              rateVascode = formatted_amount = "{:,.0f}".format((float(rate) * 0.05)+((float(rate) * 0.05)*0.11)).replace(",", ".")

              steps = [
                     #Case 1
                     ["", "Case 1 = UC+RC+OC > "+str(rate)+""," ", "", ""],
                     ["Create and actived new subscriber PP KartuHALO Bebas Abonemen","Check active period"],
                     ["Update parameter Init activation date","Success"],
                     ["Set New Credit Limit Service (offer id : 3669354) as 1.000.000 ","Success"],
                     ["Create Event Voice Onnet 60s 11am","Charged"],
                     ["Create Event Voice Offnet 60s 1pm","Charged"],
                     ["Create Event 1 SMS Onnet 3pm","Charged"],
                     ["Create Event 1 SMS Offnet 5pm","Charged"],
                     ["Attach Offer "+str(offerName),"Offer Attached"],
                     ["Check Offer Name Disc & Offer Description",""+str(offerName)+"|"+str(offerDescription)+""],
                     ["Check 888","Checked"],
                     [f"Attach Offer B2C Flexible Abonemen Charge with param name : Commitment period|Rate|Proration|Override RC amount|Invoice description|Quotation reference|External product id|Penalty Remaining|Penalty Flat|Penalty ind; Value : 12|{offerNameOCCase1}|N|0|Invoice description|Quotation reference|External product id|0|0|No","Offer Attached"],
                     [f"Attach Offer B2C Flexible Abonemen Charge with param name : Commitment period|Rate|Proration|Override RC amount|Invoice description|Quotation reference|External product id|Penalty Remaining|Penalty Flat|Penalty ind; Value : 0|{offerNameRCCase1}|N|-1|Invoice description|Quotation reference|External product id|0|0|No","Offer Attached"],
                     ["Create Event Voice International to Malaysia (60) 60s D+1 7pm ","Charged"],
                     ["Create Event GPRS 1MB RG 50 D+5 8pm","Charged"],
                     ["Create Event 1 MMS Onnet D+10 9pm","Charged"],
                     ["Create Event Direct Debit using Vascode google with Charge 11.100 before next BC","Charged"],
                     ["Check 888 (Total Usage > "+str(rate)+")","Checked"],
                     ["Check AMDD charge code for RC","Checked"],
                     ["Check AMDD charge code for OC","Checked"],
                     ["Check Cycle Month","Checked"],
                     ["Check in trb1_sub_errs","No Error"],
                     ["Check Indira (CHG)","Checked"],
                     ["Check Rated Event","Checked"],
                     ["Check RB Log","Checked"],
                     "Invoicing",
                     #Case 2
                     ["", "Case 2 = UC+RC+OC < "+str(rate)+""," ", "", ""],
                     ["Create and actived new subscriber PP KartuHALO Bebas Abonemen","Check active period"],
                     ["Update parameter Init activation date","Success"],
                     ["Set New Credit Limit Service (offer id : 3669354) as 1.000.000 ","Success"],
                     ["Create Event Voice Onnet 60s 11am","Charged"],
                     ["Create Event Voice Offnet 60s 1pm","Charged"],
                     ["Create Event 1 SMS Onnet 3pm","Charged"],
                     ["Create Event 1 SMS Offnet 5pm","Charged"],
                     ["Attach Offer "+str(offerName),"Offer Attached"],
                     ["Check Offer Name Disc & Offer Description",""+str(offerName)+"|"+str(offerDescription)+""],
                     ["Check 888","Checked"],
                     [f"Attach Offer B2C Flexible Abonemen Charge with param name : Commitment period|Rate|Proration|Override RC amount|Invoice description|Quotation reference|External product id|Penalty Remaining|Penalty Flat|Penalty ind; Value : 12|{offerNameOCCase2}|N|0|Invoice description|Quotation reference|External product id|0|0|No","Offer Attached"],
                     [f"Attach Offer B2C Flexible Abonemen Charge with param name : Commitment period|Rate|Proration|Override RC amount|Invoice description|Quotation reference|External product id|Penalty Remaining|Penalty Flat|Penalty ind; Value : 0|{offerNameRCCase2}|N|-1|Invoice description|Quotation reference|External product id|0|0|No","Offer Attached"],
                     ["Create Event Voice International to Malaysia (60) 60s D+1 7pm ","Charged"],
                     ["Create Event GPRS 1MB RG 50 D+5 8pm","Charged"],
                     ["Create Event 1 MMS Onnet D+10 9pm","Charged"],
                     [f"Create Event Direct Debit using Vascode google with Charge {rateVascode} before next BC","Charged"],
                     ["Check 888 (Total Usage > "+str(rate)+")","Checked"],
                     ["Check AMDD charge code for RC","Checked"],
                     ["Check AMDD charge code for OC","Checked"],
                     ["Check Cycle Month","Checked"],
                     ["Check in trb1_sub_errs","No Error"],
                     ["Check Indira (CHG)","Checked"],
                     ["Check Rated Event","Checked"],
                     ["Check RB Log","Checked"],
                     "Invoicing",
                     #Case 3
                     ["", "Case 3 = Remove "+str(rate)+""," ", "", ""],
                     ["Create and actived new subscriber PP KartuHALO Bebas Abonemen","Check active period"],
                     ["Update parameter Init activation date","Success"],
                     ["Set New Credit Limit Service (offer id : 3669354) as 1.000.000 ","Success"],
                     ["Attach Offer "+str(offerName),"Offer Attached"],
                     ["Check Offer Name Disc & Offer Description",""+str(offerName)+"|"+str(offerName)+""],
                     ["Check 888","Checked"],
                     [f"Attach Offer B2C Flexible Abonemen Charge with param name : Commitment period|Rate|Proration|Override RC amount|Invoice description|Quotation reference|External product id|Penalty Remaining|Penalty Flat|Penalty ind; Value : 0|{offerNameRCCase3}|N|-1|Invoice description|Quotation reference|External product id|0|0|No","Offer Attached"],
                     ["Check 888","Checked"],
                     [f"Attach Offer B2C Flexible Abonemen Charge with param name : Commitment period|Rate|Proration|Override RC amount|Invoice description|Quotation reference|External product id|Penalty Remaining|Penalty Flat|Penalty ind; Value : 12|{offerNameOCCase3}|N|0|Invoice description|Quotation reference|External product id|0|0|No","Offer Attached"],
                     ["Check 888","Checked"],
                     ["Create Event Voice Onnet 60s 11am","Charged"],
                     ["Create Event Voice Offnet 60s 1pm","Charged"],
                     ["Create Event 1 SMS Onnet 3pm","Charged"],
                     ["Create Event 1 SMS Offnet 5pm","Charged"],
                     ["Create Event Voice International to Malaysia (60) 60s 7pm ","Charged"],
                     ["Create Event GPRS 1MB RG 50 8pm","Charged"],
                     ["Create Event 1 MMS Onnet 9pm","Charged"],
                     ["Create Event Direct Debit using Vascode google with Charge 1100 11pm","Charged"],    
                     ["Remove offer "+str(offerName)+" on last BC","Offer Removed"],
                     ["Check 888 (Total Usage > "+str(rate)+")","Checked"],
                     ["Check AMDD charge code for RC","Checked"],
                     ["Check AMDD charge code for OC","Checked"],
                     ["Check Cycle Month","Checked"],
                     ["Check in trb1_sub_errs","No Error"],
                     ["Check Indira (CHG)","Checked"],
                     ["Check Rated Event","Checked"],
                     ["Check RB Log","Checked"],
                     "Invoicing"
              ]

              # Write Header Row
              header = [f'{eventName} | {offerName} ']
              ws.append(header)

              # Merge Header Cells
              startColumnIndex = 3  # Example of a dynamic column index
              startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
              endColumn = "E"
              startRow = 1
              endRow = 1
              cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
              ws.merge_cells(cellRange)

              headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
              ws.append(headerRow)
              no = 1
              for num, step in enumerate(steps, start=1):
                     
                     if isinstance(step, str):
                            row = [
                                   no,
                                   step,
                                   "Success",
                                   "No Bonus",
                                   "XYZ"
                            ]
                            no = no+1
                     else:
                            if len(step) == 5:
                                   row = [
                                          step[0],
                                          step[1],
                                          step[2],
                                          step[3],
                                          step[4]
                                   ]
                            else:
                                   row = [
                                          no,
                                          step[0],
                                          step[1],
                                          "No Bonus",
                                          "XYZ"
                                   ]
                                   no = no+1
                     ws.append(row)

       print("Testing Scenario Successfully Generated")
       
       # Save Excel File
       wb.save('Result/Scenario '+str(eventName)+'.xlsx')

def exportExcelDiscountOfferOLD(eventName, params=None, neededParams = None):
       if params is not None:
              type = params['Discount Type']
       else:
              type = ''
       
       
       #UC > Disc
       if type == 1:
              discountType = 'UC > Disc'
              steps = [
                     'Create & Activate new subscriber PP KartuHalo Bebas Abonemen ( 1 primary and 2 child)',
                     'Attach offer Diskon OU SVC 200,000',
                     'MSISDN 1 Update Parameter (Init activation date)',
                     'MSISDN 1 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 1 Check Prit Name',
                     'MSISDN 1 Check Offer Name & Offer Description',
                     'MSISDN 1 Create Event Voice 60s Onnet 11am',
                     'MSISDN 1 Create Event SMS Offnet 11am',
                     'MSISDN 1 Create Event GPRS 1MB RG 50 1am D+1',
                     'MSISDN 1 Create Event MMS Onnet 1am D+1',
                     'MSISDN 1 Create Event Direct Debit using Vascode google with Charge 100000 11pm D+1',
                     'MSISDN 2 Update Parameter (Init activation date)',
                     'MSISDN 2 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 2 Check Prit Name',
                     'MSISDN 2 Create Event Direct Debit using Vascode google with Charge 100000 11pm D+1',
                     'MSISDN 3 Update Parameter (Init activation date)',
                     'MSISDN 3 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 3 Check Prit Name',
                     'MSISDN 3 Create Event Direct Debit using Vascode google with Charge 50000 11pm D+1',
                     'Check in trb1_sub_errs',
                     'Invoicing',
              ]
       #UC < Disc
       elif type == 2:
              discountType = 'UC < Disc'
              steps = [
                     'Create & Activate new subscriber PP KartuHalo Bebas Abonemen ( 1 primary and 2 child)',
                     'Attach offer Diskon OU SVC 200,000',
                     'MSISDN 1 Update Parameter (Init activation date)',
                     'MSISDN 1 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 1 Check Prit Name',
                     'MSISDN 1 Check Offer Name & Offer Description',
                     'MSISDN 1 Create Event Voice 60s Onnet 11am',
                     'MSISDN 1 Create Event SMS Offnet 11am',
                     'MSISDN 1 Create Event GPRS 1MB RG 50 1am D+1',
                     'MSISDN 1 Create Event MMS Onnet 1am D+1',
                     'MSISDN 1 Create Event Direct Debit using Vascode google with Charge 30000 11pm D+1',
                     'MSISDN 2 Update Parameter (Init activation date)',
                     'MSISDN 2 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 2 Check Prit Name',
                     'MSISDN 2 Check Offer Name & Offer Description',
                     'MSISDN 2 Create Event Direct Debit using Vascode google with Charge 10000 11pm D+1',
                     'MSISDN 3 Update Parameter (Init activation date)',
                     'MSISDN 3 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 3 Check Prit Name',
                     'MSISDN 3 Check Offer Name & Offer Description',
                     'MSISDN 3 Create Event Direct Debit using Vascode google with Charge 10000 11pm D+1',
                     'Check in trb1_sub_errs',
                     'Invoicing',
              ]
       #UC = Disc
       elif type == 3:
              discountType = 'UC = Disc'
              steps = [
                     'Create & Activate new subscriber PP KartuHalo Bebas Abonemen ( 1 primary and 2 child)',
                     'Attach offer Diskon OU SVC 200,000',
                     'MSISDN 1 Update Parameter (Init activation date)',
                     'MSISDN 1 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 1 Check Prit Name',
                     'MSISDN 1 Check Offer Name & Offer Description',
                     'MSISDN 1 Create Event Direct Debit using Vascode google with Charge 100000 11pm D+1',
                     'MSISDN 2 Update Parameter (Init activation date)',
                     'MSISDN 2 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 2 Check Prit Name',
                     'MSISDN 2 Create Event Direct Debit using Vascode google with Charge 82000 11pm D+1',
                     'MSISDN 3 Update Parameter (Init activation date)',
                     'MSISDN 3 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 3 Check Prit Name',
                     'MSISDN 3 Create Event Direct Debit using Vascode google with Charge 40000 11pm D+1',
                     'Check in trb1_sub_errs',
                     'Invoicing',
              ]
       #RC > Disc
       elif type == 4:
              discountType = 'RC > Disc'
              steps = [
                     "Create & Activate new subscriber PP KartuHalo Bebas Abonemen ( 1 primary and 2 child)",
                     "Attach offer Diskon OU SVC 200,000",
                     "MSISDN 1 Update Parameter (Init activation date)",
                     "MSISDN 1 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 1 Attach Offer Flexible Charge Subscription Fee (set Charge Offer 600000) (3759586)",
                     "MSISDN 1 Check Prit Name",
                     "MSISDN 1 Check Offer Name & Offer Description",
                     "MSISDN 1 Check in Charge and AMDD Charge Code in bl1_rc_rates",
                     "MSISDN 2 Update Parameter (Init activation date)",
                     "MSISDN 2 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 2 Attach Offer RC Paket DPI Dana 2GB | 3868529 -- Amount 50.000",
                     "MSISDN 2 Check Prit Name",
                     "MSISDN 2 Check Offer Name & Offer Description",
                     "MSISDN 2 Check in Charge and AMDD Charge Code in bl1_rc_rates",
                     "MSISDN 3 Update Parameter (Init activation date)",
                     "MSISDN 3 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 3 Attach Offer RC Paket DPI Dana 2GB | 3868529 -- Amount 50.000",
                     "MSISDN 3 Check Prit Name",
                     "MSISDN 3 Check Offer Name & Offer Description",
                     "MSISDN 3 Check in Charge and AMDD Charge Code in bl1_rc_rates",
                     "Check in trb1_sub_errs",
                     "Invoicing",
              ]
       #RC < Disc
       elif type == 5:
              discountType = 'RC < Disc'
              steps = [
                     "Create & Activate new subscriber PP KartuHalo Bebas Abonemen ( 1 primary and 2 child)",
                     "Attach offer Diskon OU SVC 200,000",
                     "MSISDN 1 Update Parameter (Init activation date)",
                     "MSISDN 1 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 3 Attach Offer RC Paket DPI Dana 2GB | 3868529 -- Amount 50.000",
                     "MSISDN 1 Check Prit Name",
                     "MSISDN 1 Check Offer Name & Offer Description",
                     "MSISDN 1 Check in Charge and AMDD Charge Code in bl1_rc_rates",
                     "MSISDN 2 Update Parameter (Init activation date)",
                     "MSISDN 2 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 3 Attach Offer RC Paket DPI Dana 2GB | 3868529 -- Amount 50.000",
                     "MSISDN 2 Check Prit Name",
                     "MSISDN 2 Check Offer Name & Offer Description",
                     "MSISDN 2 Check in Charge and AMDD Charge Code in bl1_rc_rates",
                     "MSISDN 3 Update Parameter (Init activation date)",
                     "MSISDN 3 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 3 Attach Offer RC Paket DPI Dana 2GB | 3868529 -- Amount 50.000",
                     "MSISDN 3 Check Prit Name",
                     "MSISDN 3 Check Offer Name & Offer Description",
                     "MSISDN 3 Check in Charge and AMDD Charge Code in bl1_rc_rates",
                     "Check in trb1_sub_errs",
                     "Invoicing",
              ]
       #RC = Disc
       elif type == 6:
              discountType = 'RC = Disc'
              steps = [
                     "Create & Activate new subscriber PP KartuHalo Bebas Abonemen ( 1 primary and 2 child)",
                     "Attach offer Diskon OU SVC 200,000",
                     "MSISDN 1 Update Parameter (Init activation date)",
                     "MSISDN 1 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 1 Attach Offer Flexible Charge Subscription Fee (set Charge Offer 100mio) (3759586)",
                     "MSISDN 1 Check Prit Name",
                     "MSISDN 1 Check Offer Name & Offer Description",
                     "MSISDN 1 Check in Charge and AMDD Charge Code in bl1_rc_rates",
                     "MSISDN 2 Update Parameter (Init activation date)",
                     "MSISDN 2 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 3 Update Parameter (Init activation date)",
                     "MSISDN 3 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
              ]
       #OC < Disc
       elif type == 7:
              discountType = 'OC < Disc'
              steps = [
                     "Create & Activate new subscriber PP KartuHalo Bebas Abonemen ( 1 primary and 2 child)",
                     "Attach offer Diskon OU SVC 200,000",
                     "MSISDN 1 Update Parameter (Init activation date)",
                     "MSISDN 1 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 1 Attach Offer B2C Flexible Charge Subscription Fee (3875629) with param (Commitment period|Rate|Penalty Remaining|Invoice description|Quotation reference|External product id|Penalty Flat|Penalty ind|Proration) and value (12|14000000|0|Invoice description|Quotation reference|External product id|0|No|N)",
                     "MSISDN 1 Check Prit Name",
                     "MSISDN 1 Check Offer Name & Offer Description",
                     "MSISDN 1 Check in Charge and AMDD Charge Code in bl1_rc_rates",
                     "MSISDN 2 Update Parameter (Init activation date)",
                     "MSISDN 2 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 2 Attach Offer B2C Flexible Charge Subscription Fee (3875629) with param (Commitment period|Rate|Penalty Remaining|Invoice description|Quotation reference|External product id|Penalty Flat|Penalty ind|Proration) and value (12|14000000|0|Invoice description|Quotation reference|External product id|0|No|N)",
                     "MSISDN 2 Check Prit Name",
                     "MSISDN 2 Check Offer Name & Offer Description",
                     "MSISDN 2 Check in Charge and AMDD Charge Code in bl1_rc_rates",
                     "MSISDN 3 Update Parameter (Init activation date)",
                     "MSISDN 3 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 3 Attach Offer B2C Flexible Charge Subscription Fee (3875629) with param (Commitment period|Rate|Penalty Remaining|Invoice description|Quotation reference|External product id|Penalty Flat|Penalty ind|Proration) and value (12|14000000|0|Invoice description|Quotation reference|External product id|0|No|N)",
                     "MSISDN 3 Check Prit Name",
                     "MSISDN 3 Check Offer Name & Offer Description",
                     "MSISDN 3 Check in Charge and AMDD Charge Code in bl1_rc_rates",
                     "MSISDN 3 Check Prit Name",
                     "Check in trb1_sub_errs",
                     "Invoicing",
              ]
       #OC > Disc
       elif type == 8:
              discountType = 'OC > Disc'
              steps = [
                     'Create & Activate new subscriber PP KartuHalo Bebas Abonemen ( 1 primary and 2 child)',
                     'Attach offer Diskon OU SVC 200,000',
                     'MSISDN 1 Update Parameter (Init activation date)',
                     'MSISDN 1 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 1 attach Samsung Knox Configure Dynamic - GAMA (charge 140000) (3884149)',
                     'MSISDN 1 Check Prit Name',
                     'MSISDN 1 Check Offer Name & Offer Description',
                     'MSISDN 1 Check in Charge and AMDD Charge Code in bl1_rc_rates',
                     'MSISDN 2 Update Parameter (Init activation date)',
                     'MSISDN 2 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 2 attach Samsung Knox Configure Dynamic - GAMA (charge 140000) (3884149)',
                     'MSISDN 2 Check Prit Name',
                     'MSISDN 2 Check Offer Name & Offer Description',
                     'MSISDN 2 Check in Charge and AMDD Charge Code in bl1_rc_rates',
                     'MSISDN 3 Update Parameter (Init activation date)',
                     'MSISDN 3 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 3 attach Samsung Knox Configure Dynamic - GAMA (charge 140000) (3884149)',
                     'MSISDN 3 Check Prit Name',
                     'MSISDN 3 Check Offer Name & Offer Description',
                     'MSISDN 3 Check in Charge and AMDD Charge Code in bl1_rc_rates',
                     'Check in trb1_sub_errs',
                     'Invoicing',
              ]
       #OC = Disc
       elif type == 9:
              discountType = 'OC = Disc'
              steps = [
                     'Create & Activate new subscriber PP KartuHalo Bebas Abonemen ( 1 primary and 2 child)',
                     'Attach offer Diskon OU SVC 200,000',
                     'MSISDN 1 Update Parameter (Init activation date)',
                     'MSISDN 1 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 1 Attach Offer B2C Flexible Charge Subscription Fee (3875629) with param (Commitment period|Rate|Penalty Remaining|Invoice description|Quotation reference|External product id|Penalty Flat|Penalty ind|Proration) and value (12|14000000|0|Invoice description|Quotation reference|External product id|0|No|N)',
                     'MSISDN 1 Check Prit Name',
                     'MSISDN 1 Check Offer Name & Offer Description',
                     'MSISDN 1 Check in Charge and AMDD Charge Code in bl1_rc_rates',
                     'MSISDN 2 Update Parameter (Init activation date)',
                     'MSISDN 2 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 2 Attach Offer B2C Flexible Charge Subscription Fee (3875629) with param (Commitment period|Rate|Penalty Remaining|Invoice description|Quotation reference|External product id|Penalty Flat|Penalty ind|Proration) and value (12|14000000|0|Invoice description|Quotation reference|External product id|0|No|N)',
                     'MSISDN 2 Check Prit Name',
                     'MSISDN 2 Check Offer Name & Offer Description',
                     'MSISDN 2 Check in Charge and AMDD Charge Code in bl1_rc_rates',
                     'MSISDN 3 Update Parameter (Init activation date)',
                     'MSISDN 3 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 3 Attach Offer B2C Flexible Charge Subscription Fee (3875629) with param (Commitment period|Rate|Penalty Remaining|Invoice description|Quotation reference|External product id|Penalty Flat|Penalty ind|Proration) and value (12|14000000|0|Invoice description|Quotation reference|External product id|0|No|N)',
                     'MSISDN 3 Check Prit Name',
                     'MSISDN 3 Check Offer Name & Offer Description',
                     'MSISDN 3 Check in Charge and AMDD Charge Code in bl1_rc_rates',
                     'Check in trb1_sub_errs',
                     'Invoicing',
              ]
       #UC RC OC > disc
       elif type == 10:
              discountType = 'UC RC OC > Disc'
              steps = [
                     'Create & Activate new subscriber PP KartuHalo Bebas Abonemen ( 1 primary and 2 child)',
                     'Attach offer Diskon OU SVC 200,000',
                     'MSISDN 1 Update Parameter (Init activation date)',
                     'MSISDN 1 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 1 attach Samsung Knox Configure Dynamic - GAMA (charge 140000) (OC)',
                     'MSISDN 1 Check prit name ',
                     'MSISDN 1 Check Offer Name & Offer Description',
                     'MSISDN 2 Update Parameter (Init activation date)',
                     'MSISDN 2 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 2 Attach Offer RC Add On Berlangganan VIU - 3935794  (RC 22727)',
                     'MSISDN 2 Check prit name ',
                     'MSISDN 2 Check Offer Name & Offer Description',
                     'MSISDN 3 Update Parameter (Init activation date)',
                     'MSISDN 3 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR',
                     'MSISDN 3 Create Event Voice 60s Onnet 11am',
                     'MSISDN 3 Create Event SMS Offnet 11am',
                     'MSISDN 3 Create Event Direct Debit using Vascode google with Charge 1000 11pm ',
                     'Check in trb1_sub_errs',
                     'Invoicing',
              ]
       #UC RC OC < disc
       elif type == 11:
              discountType = 'UC RC OC < Disc'
              steps = [
                     "Create & Activate new subscriber PP KartuHalo Bebas Abonemen ( 1 primary and 2 child)",
                     "Attach offer Diskon OU SVC 200,000",
                     "MSISDN 1 Update Parameter (Init activation date)",
                     "MSISDN 1 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 1 Attach Offer OC 3884149 Samsung Knox Configure Dynamic - GAMA ( OC 140000 )",
                     "MSISDN 1 Check Prit Name",
                     "MSISDN 1 Check Offer Name & Offer Description",
                     "MSISDN 1 Create Event Voice 60s Onnet 11am",
                     "MSISDN 1 Create Event SMS Offnet 11am",
                     "MSISDN 1 Create Event GPRS 1MB RG 50 1am D+1",
                     "MSISDN 1 Create Event MMS Onnet 1am",
                     "MSISDN 2 Update Parameter (Init activation date)",
                     "MSISDN 2 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 2 Attach Offer RC Attach Offer RC Paket DPI Dana 2GB | 3868529 -- Amount 50.000",
                     "MSISDN 2 Video Conference MF 210K (charged 210000)",
                     "MSISDN 2 Create Event Direct Debit using Vascode google with Charge 10000 11pm",
                     "MSISDN 3 Update Parameter (Init activation date)",
                     "MSISDN 3 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 3 Attach Offer RC Add On Berlangganan VIU - 3935794  (RC 22727)",
                     "MSISDN 3 Video Conference MF 210K (charged 210000)",
                     "MSISDN 3 Create Event Direct Debit using Vascode google with Charge 10000 11pm",
                     "Check in trb1_sub_errs",
                     "Invoicing",
              ]
       #UC RC OC = disc
       elif type == 12:
              discountType = 'UC RC OC = Disc'
              steps = [
                     "Create & Activate new subscriber PP KartuHalo Bebas Abonemen ( 1 primary and 2 child)",
                     "Attach offer Diskon OU SVC 200,000",
                     "MSISDN 1 Update Parameter (Init activation date)",
                     "MSISDN 1 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 1 Attach Offer OC 3884149 Samsung Knox Configure Dynamic - GAMA ( OC 140000 )",
                     "MSISDN 1 Check Prit Name",
                     "MSISDN 1 Check Offer Name & Offer Description",
                     "MSISDN 2 Update Parameter (Init activation date)",
                     "MSISDN 2 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 2 Attach Offer RC Attach Offer RC Paket DPI Dana 2GB | 3868529 -- Amount 50.000 (prorate 23333)",
                     "MSISDN 2 Check Prit Name",
                     "MSISDN 2 Check Offer Name & Offer Description",
                     "MSISDN 3 Update Parameter (Init activation date)",
                     "MSISDN 3 Set New Credit Limit Service (offer id : 3669354) as 1.000.000 IDR",
                     "MSISDN 3 Create Event Direct Debit using Vascode google with Charge 40737 11pm ",
                     "Check in trb1_sub_errs",
                     "Invoicing",
              ]

       # Export Test Cases to Excel File
       wb = Workbook()
       ws = wb.active

       # Write Header Row
       header = [f'{eventName} | {discountType}']
       ws.append(header)

       # Merge Header Cells
       startColumn = "A"
       endColumn = "E"
       startRow = 1
       endRow = 1
       cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
       ws.merge_cells(cellRange)
              

       headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
       ws.append(headerRow)

       for no, step in enumerate(steps):
              no = no+1
              row = [
                     no,
                     step,
                     "Success",
                     "No Bonus",
                     "XYZ"
              ]
              ws.append(row)

       print("Testing Scenario Successfully Generated")
       
       # Save Excel File
       wb.save('Result/Scenario '+str(eventName)+' '+str(offerName)+'.xlsx')

def exportExcelNewServiceFilter(eventName, params=None, neededParams = None):
       # Export Test Cases to Excel File
       wb = Workbook()
       ws = wb.active
       for params in params:
              if "Product Name" in params:
                     productName = params['Product Name']
              else:
                     productName = ''

              if "Service Filter" in params:
                     serviceFilter = params['Service Filter'].split(",")
                     if len(serviceFilter) > 1:
                            serviceFilterON = serviceFilter[0]
                            serviceFilterOFF = serviceFilter[1]
                     else:
                            serviceFilterON = serviceFilter
                            serviceFilterOFF = ''
              else:
                     serviceFilter = ""
                     serviceFilterON = ""
                     serviceFilterOFF = ""

              if serviceFilterON != '':
                     strServiceFilterON = "| Check Service filter (should be "+str(serviceFilterON)+")"
              else:
                     strServiceFilterON = ''

              if serviceFilterOFF != '':
                     strServiceFilterOFF = "| Check Service filter (should be "+str(serviceFilterOFF)+")"
              else:
                     strServiceFilterOFF = ''

              if "Charge Code" in params:
                     chargeCode = params['Charge Code'].split(",")
                     if len(chargeCode) > 1:
                            chargeCodeON = chargeCode[0]
                            chargeCodeOFF = chargeCode[1]
                     else:
                            chargeCodeON = chargeCode
                            chargeCodeOFF = ''
              else:
                     chargeCode = ""
                     chargeCodeON = ""
                     chargeCodeOFF = ""

              if chargeCodeON != '':
                     strchargeCodeON = "| Check rounded event and Charge Code (should be "+str(chargeCodeON)+")"
              else:
                     strchargeCodeON = ''

              if chargeCodeOFF != '':
                     strchargeCodeOFF = "| Check rounded event and Charge Code (should be "+str(chargeCodeOFF)+")"
                     strCheckChargeCodeOFF = "| Check Charge Code (should be "+str(chargeCodeOFF)+")"
              else:
                     strchargeCodeON = ''
              

              steps = [
                     ["Create & Activate new subscriber PP KartuHALO Corporate Executive","Check active period"],
                     ["Update Parameter","Parameter Updated"],
                     ["Set New Credit Limit Service (offer id : 3669334) as 1.000.000 | 3669334","Offer Attached"],
                     ["Set CLS Roaming (offer id : 3669354) as 20.000.000 | 3669354","Offer Attached"],
                     ["Attach offer international roaming","Offer Attached"],
                     ["Attach offer 3908759-"+str(productName)+"","Offer Attached"],
                     ["Check 888","Checked"],
                     ["Check offer name and offer description",""+str(productName)+"|"+str(productName)+""],
                     ["Create event voice onnet 8s, 3PM "+strServiceFilterON,"Charged"],
                     ["Create event voice onnet 190s, 3PM "+strServiceFilterON,"Charged"],
                     ["Create event voice onnet 5s, 3PM "+strchargeCodeON,"Charged"],
                     ["Create event voice offnet 8s, 3PM "+strServiceFilterOFF,"Charged"],
                     ["Create event voice offnet 185s, 3PM "+strServiceFilterOFF,"Charged"],
                     ["Create event voice offnet 5s, 3PM "+strchargeCodeOFF,"Charged"],
                     ["Create event voice PSTN 5s, 3PM "+strServiceFilterOFF,"Charged"],
                     ["Create event voice PSTN 210s, 3PM "+strCheckChargeCodeOFF,"Charged"],
                     ["Create event voice FWA 8s, 3PM "+strServiceFilterOFF,"Charged"],
                     ["Create event voice FWA 180s, 3PM "+strCheckChargeCodeOFF,"Charged"],
                     ["Create event voice onnet 1s, 6PM D+1 | Check rounded event (Should be 1s) and Check Pricing Item ID","Charged"],
                     ["Create event voice offnet 60s, 7PM D+1 | Check rounded event (Should be 6s) and Check Pricing Item ID","Charged"],
                     ["Create event voice PSTN 1s, 8PM D+1 | Check Pricing Item ID","Charged"],
                     ["Create event voice FWA 181s, 8PM D+1 | Check Pricing Item ID","Charged"],
                     ["Create event voice International to Singapore (+65) 60s, 9PM D+1","Charged"],
                     ["Create event voice spesial number to 14045 60s, 9PM D+1","Charged"],
                     ["Create event voice roaming mo home from Malaysia (+60) 60s,  10PM D+1","Charged"],
                     ["Create event 1 sms onnet 10PM D+1","Charged"],
                     ["Create event 1 sms offnet 10PM D+1","Charged"],
                     ["Create event 1 sms international to SIngapore (+65) 11PM D+1","Charged"],
                     ["Create event 1MB GPRS RG17, 11PM D+1","Charged"],
                     ["Remove offer 3908759-"+str(productName)+"","Offer removed"],
                     ["Create event voice onnet 5s, 11AM | Check Pricing Item ID and Check Rate back to basic","Charged"],
                     ["Create event voice offnet 10s, 1PM | Check Pricing Item ID and Check Rate back to basic","Charged"],
                     ["Create event voice PSTN 240s, 4PM | Check Pricing Item ID and Check Rate back to basic","Charged"],
                     ["Create event voice FWA 9s, 5AM | Check Pricing Item ID and Check Rate back to basic","Charged"],
                     ["Check AMDD","Checked"],
                     ["Check Indira","Checked"],
                     ["Check RBLog","Checked"],
                     ["Check TRB_SUB_Errs","Checked|Should be empty"],
                     ["Check Rated Event","Checked"],
                     "INVOICING",
              ]

              # Write Header Row
              header = [f'{eventName} | {productName} ']
              ws.append(header)

              # Merge Header Cells
              startColumnIndex = 3  # Example of a dynamic column index
              startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
              endColumn = "E"
              startRow = 1
              endRow = 1
              cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
              ws.merge_cells(cellRange)

              headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
              ws.append(headerRow)

              for no, step in enumerate(steps):
                     no = no+1
                     if isinstance(step, str):
                            row = [
                                   no,
                                   step,
                                   "Success",
                                   "No Bonus",
                                   "XYZ"
                            ]
                     else:
                            row = [
                                   no,
                                   step[0],
                                   step[1],
                                   "No Bonus",
                                   "XYZ"
                            ]
                     ws.append(row)

       print("Testing Scenario Successfully Generated")
       
       # Save Excel File
       wb.save('Result/Scenario '+str(eventName)+'.xlsx')

def exportExcelMinimunUsage(eventName, params=None, neededParams = None):
       # Export Test Cases to Excel File
       wb = Workbook()
       ws = wb.active
       for params in params:
              if "Offer Name" in params:
                     offerName = params['Offer Name']
              else:
                     offerName = ''
              
              if "Offer Description" in params:
                     offerDesc = params['Offer Description']
              else:
                     offerDesc = ''
              
              if "Amount" in params:
                     amount = params['Amount']
                     amount = float(amount)
              else:
                     amount = ''

              if "Vascode" in params:
                     vascode = params['Vascode']
                     if (vascode == ''):
                            vascode = 'google'       
              else:
                     vascode = 'google'
              
              creditLimitService = "{:,.0f}".format(amount+(amount*0.5)).replace(",", ".")
              intFlexibleVascode = amount*0.1
              flexibleVascode = "{:,.0f}".format(intFlexibleVascode).replace(",", ".")
              flexibleVascodePPN = "{:,.0f}".format(intFlexibleVascode+(intFlexibleVascode*0.11)).replace(",", ".")
              amountFlexibleOC = amount*0.1


              stepsCase1 = [
                     ["","Case 1: Attach offer RC + MU Offer","","",""],
                     ["Create & Activate new subscriber KartuHALO Bebas Abonemen v2 (PP ID : 2427)","Check active period"],
                     ["Update Parameter (Init activation date)","Updated"],
                     [f"Set New Credit Limit Service (offer id : 3669334) as {creditLimitService}","Offer Attached"],
                     ["Attach Offer Video Conference MF1110K (Charge Offer 1100000) (3882799)","Offer Attached"],
                     ["Attach Offer Minimum Usage "+str(offerName),"Offer Attached"],
                     ["Check Offer Description",str(offerName)+" | "+str(offerDesc)],
                     ["Check 888","Checked"],
                     ["Check AMDD RC", "Checked"],
                     ["Check TRB1_Sub_Errs","Checked | Makesure no Error"],
                     ["Check Indira","Checked"],
                     "Invoicing"	 
              ]

              stepsCase2 = [
                     ["","Case 2: Attach Offer OC/OC Installment + MU Offer","","",""],
                     ["Create & Activate new subscriber KartuHALO Bebas Abonemen v2 (PP ID : 2427)","Check active period"],
                     ["Update Parameter (Init activation date)","Updated"],
                     [f"Set New Credit Limit Service (offer id : 3669334) as {creditLimitService}","Offer Attached"],
                     [f"Attach Offer B2C Flexible Abonemen Charge with param name : Commitment period|Rate|Proration|Override RC amount|Invoice description|Quotation reference|External product id|Penalty Remaining|Penalty Flat|Penalty ind; 12|{amountFlexibleOC}|N|0|Invoice description|Quotation reference|External product id|0|0|No","Offer Attached"],
                     ["Attach Offer Minimum Usage "+str(offerName),"Offer Attached"],
                     ["Check Offer Description",str(offerName)+" | "+str(offerDesc)],
                     ["Check 888","Checked"],
                     ["Check AMDD OC", "Checked"],
                     ["Check TRB1_Sub_Errs","Checked | Makesure no Error"],
                     ["Check Indira","Checked"],
                     "Invoicing"
              ]

              stepsCase3 = [
                     ["","Case 3: Attach Offer OC Installment w Penalty + MU Offer","","",""],
                     ["Create & Activate new subscriber KartuHALO Bebas Abonemen v2 (PP ID : 2427)","Check active period"],
                     ["Update Parameter (Init activation date)","Updated"],
                     [f"Set New Credit Limit Service (offer id : 3669334) as {creditLimitService}","Offer Attached"],
                     ["Attach Offer OC Ins penalty Internet SuperRoam Subscription Charge 150K | 3875659 -- Amount 150000 set Commitment period 12 month","Offer Attached"],
                     ["Attach Offer Minimum Usage "+str(offerName),"Offer Attached"],
                     ["Check Offer Description",str(offerName)+" | "+str(offerDesc)],
                     ["Check 888","Checked"],
                     ["Check TRB1_Sub_Errs","Checked | Makesure no Error"],
                     ["Remove offer D+1 === OC Ins penalty Internet SuperRoam Subscription Charge 150K | 3875659 -- Amount 150000","Removed"],
                     ["Check Indira","Checked"],
                     "Invoicing"	 
              ]

              stepsCase4 = [
                     ["","Case 4: Attach MU Offer + Create Usage < MU","","",""],
                     ["Create & Activate new subscriber KartuHALO Bebas Abonemen v2 (PP ID : 2427)","Check active period"],
                     ["Update Parameter (Init activation date)","Updated"],
                     [f"Set New Credit Limit Service (offer id : 3669334) as {creditLimitService}","Offer Attached"],
                     ["Attach Offer Minimum Usage "+str(offerName)+"","Offer Attached"],
                     ["Check Offer Description",str(offerName)+" | "+str(offerDesc)],
                     ["Check 888","Checked"],
                     ["5 PM, Create event voice onnet 600s","Charged"],
                     ["6 PM, Create event voice offnet 600s","Charged"],
                     ["7 PM D+1, Create event 1 SMS onnet","Charged"],
                     ["8 PM D+1, Create event 1 SMS offnet","Charged"],
                     ["11 PM D+3, Create event GPRS 100MB RG 55","Charged"],
                     [f"11 PM D+3, Create event Direct Debit using vascode {vascode} {flexibleVascodePPN}",f"Charge {flexibleVascode} IDR"],
                     ["Check TRB1_Sub_Errs","Checked | Makesure no Error"],
                     ["Check Indira","Checked"],
                     "Invoicing"	 
              ]

              if (amount % 3 == 0):
                     flexibleVascode1AmountInt   = amount // 3
                     flexibleVascode1Amount      = "{:,.0f}".format(flexibleVascode1AmountInt).replace(",", ".")
                     flexibleVascode2AmountInt   = amount // 3
                     flexibleVascode2Amount      = "{:,.0f}".format(flexibleVascode2AmountInt).replace(",", ".")
                     flexibleVascode3AmountInt   = amount - (flexibleVascode1AmountInt + flexibleVascode2AmountInt)
                     flexibleVascode3Amount      = "{:,.0f}".format(flexibleVascode3AmountInt).replace(",", ".")
                     
              elif (amount % 4 == 0):
                     flexibleVascode1AmountInt   = amount // 4
                     flexibleVascode1Amount      = "{:,.0f}".format(flexibleVascode1AmountInt).replace(",", ".")
                     flexibleVascode2AmountInt   = amount // 4
                     flexibleVascode2Amount      = "{:,.0f}".format(flexibleVascode2AmountInt).replace(",", ".")
                     flexibleVascode3AmountInt   = amount - (flexibleVascode1AmountInt + flexibleVascode2AmountInt)
                     flexibleVascode3Amount      = "{:,.0f}".format(flexibleVascode3AmountInt).replace(",", ".")
              else:
                     flexibleVascode1AmountInt   = round(amount // 2)
                     flexibleVascode1Amount      = "{:,.0f}".format(flexibleVascode1AmountInt).replace(",", ".")
                     flexibleVascode2AmountInt   = round((amount - flexibleVascode1AmountInt) // 2)
                     flexibleVascode2Amount      = "{:,.0f}".format(flexibleVascode2AmountInt).replace(",", ".")
                     flexibleVascode3AmountInt   = amount - (flexibleVascode1AmountInt + flexibleVascode2AmountInt)
                     flexibleVascode3Amount      = "{:,.0f}".format(flexibleVascode3AmountInt).replace(",", ".")
              

              flexibleVascode1PPN = "{:,.0f}".format(flexibleVascode1AmountInt+(flexibleVascode1AmountInt*0.11)).replace(",", ".")
              flexibleVascode2PPN = "{:,.0f}".format(flexibleVascode2AmountInt+(flexibleVascode2AmountInt*0.11)).replace(",", ".")
              flexibleVascode3PPN = "{:,.0f}".format(flexibleVascode3AmountInt+(flexibleVascode3AmountInt*0.11)).replace(",", ".")
              stepsCase5 = [
                     ["","Case 5: Attach MU Offer + Create Usage = MU","","",""],
                     ["Create & Activate new subscriber KartuHALO Bebas Abonemen v2 (PP ID : 2427)","Check active period"],
                     ["Update Parameter (Init activation date)","Updated"],
                     [f"Set New Credit Limit Service (offer id : 3669334) as {creditLimitService}","Offer Attached"],
                     ["Attach Offer Minimum Usage "+str(offerName)+"","Offer Attached"],
                     ["Check Offer Description",str(offerName)+" | "+str(offerDesc)],
                     ["Check 888","Checked"],
                     [f"12 AM, Create flexible vascode {vascode} {flexibleVascode1PPN}",f"Charge {flexibleVascode1Amount} IDR"],
                     [f"1 PM D+1, Create flexible vascode {vascode} {flexibleVascode2PPN}",f"Charge {flexibleVascode2Amount} IDR"],
                     [f"2 PM D+3, Create flexible vascode {vascode} {flexibleVascode3PPN}",f"Charge {flexibleVascode3Amount} IDR"],
                     ["Check TRB1_Sub_Errs","Checked | Makesure no Error"],
                     ["Check Indira","Checked"],
                     "Invoicing"	 
              ]

              stepsCase6 = [
                     ["","Case 6: Attach MU Offer + Create Usage > MU","","",""],
                     ["Create & Activate new subscriber KartuHALO Bebas Abonemen v2 (PP ID : 2427)","Check active period"],
                     ["Update Parameter (Init activation date)","Updated"],
                     [f"Set New Credit Limit Service (offer id : 3669334) as {creditLimitService}","Offer Attached"],
                     ["Attach Offer Minimum Usage "+str(offerName)+"","Offer Attached"],
                     ["Check Offer Description",str(offerName)+" | "+str(offerDesc)],
                     ["Check 888","Checked"],
                     ["D+0 5 PM, Create event voice Onnet 60s","Charged 1440 IDR"],
                     ["D+1 7 PM, Create event 10 SMS onnet","Charged 2300 IDR"],
                     ["D+2 11 PM, Create event GPRS 1MB RG 55","Charged 6144 IDR"],
                     ["D+3 11 PM, Create event Direct Debit using vascode bank_digi_250","Charged 250 IDR"],
                     [f"D+4 11 AM, Create event Direct Debit using vascode {vascode} {flexibleVascode1PPN}",f"Charge {flexibleVascode1Amount} IDR"],
                     ["D+5 5 PM, Create event voice Onnet 600s","Charged 14400 IDR"],
                     ["D+6 7 PM, Create event 10 SMS onnet","Charged 2300 IDR"],
                     [f"D+7 11 PM, Create event Direct Debit using vascode {vascode} {flexibleVascode2PPN}",f"Charge {flexibleVascode2Amount} IDR"],
                     [f"D+7 11 PM, Create event Direct Debit using vascode {vascode} {flexibleVascode3PPN}",f"Charge {flexibleVascode3Amount} IDR"],
                     ["Check TRB1_Sub_Errs","Checked | Makesure no Error"],
                     ["Check Indira","Checked"],
                     "Invoicing"	 
              ]

              stepsCase7 = [
                     ["","Case 7: Attach MU Offer < Create Usage Roaming","","",""],
                     ["Create & Activate new subscriber KartuHALO Bebas Abonemen v2 (PP ID : 2427)","Check active period"],
                     ["Update Parameter (Init activation date)","Updated"],
                     [f"Set New Credit Limit Service (offer id : 3669334) as {creditLimitService}","Offer Attached"],
                     [f"Set New Credit Limit Service International (offer id : 3669354) as {creditLimitService}","Offer Attached"],
                     ["Attach Offer International Roaming 36327","Offer Attached"],
                     ["Attach Offer Minimum Usage "+str(offerName)+"","Checked"],
                     ["Check Offer Description",str(offerName)+" | "+str(offerDesc)],
                     ["Check 888","Checked"],
                     ["3 PM, Create Event Voice Roaming MOC Home in Malaysia (60) 60s","Charged 35.000 IDR"],
                     ["6 PM, Create Event Voice Roaming MTC Local in Singapore (65) 60s","Charged 20.000 IDR"],
                     ["5 AM D+1, Create event vascode rw_asia_3in1_6000","Charged"],
                     ["8 AM D+1, Create event vascode rw_asia_3in1_7000","Charged"],
                     ["Check TRB1_Sub_Errs","Checked | Makesure no Error"],
                     ["Check Indira","Checked"],
                     "Invoicing"	 
              ]

              stepsCase8 = [
                     ["","Case 8: Attach RC + OC + Usage + MU Offer","","",""],
                     ["Create & Activate new subscriber KartuHALO Bebas Abonemen v2 (PP ID : 2427)","Check active period"],
                     ["Update Parameter (Init activation date)","Updated"],
                     [f"Set New Credit Limit Service (offer id : 3669334) as {creditLimitService}","Offer Attached"],
                     ["Attach Offer Minimum Usage "+str(offerName)+"","Offer Attached"],
                     ["Attach Offer RC Video Conference MF1110K (Charge Offer 1100000) (3882799)","Offer Attached"],
                     [f"Attach Offer B2C Flexible Abonemen Charge with param name : Commitment period|Rate|Proration|Override RC amount|Invoice description|Quotation reference|External product id|Penalty Remaining|Penalty Flat|Penalty ind; 12|{amountFlexibleOC}|N|0|Invoice description|Quotation reference|External product id|0|0|No","Offer Attached"],
                     ["Check Offer Description",str(offerName)+" | "+str(offerDesc)],
                     ["Check 888","Checked"],
                     [f"11 AM, Create flexible vascode {vascode} {flexibleVascodePPN}",f"Charged {flexibleVascode} IDR"],
                     [f"2 PM, Create flexible vascode {vascode} {flexibleVascodePPN}",f"Charged {flexibleVascode} IDR"],
                     ["5 PM, Create event voice onnet 6000s","Charged"],
                     ["6 PM D+1, Create event voice offnet 6000s","Charged"],
                     ["7 PM D+1, Create event 1 SMS onnet","Charged"],
                     ["Check AMDD RC", "Checked"],
                     ["Check AMDD OC", "Checked"],
                     ["Check TRB1_Sub_Errs","Checked | Makesure no Error"],
                     ["Check Indira","Checked"],
                     "Invoicing"	 
              ]

              allSteps = stepsCase1 + stepsCase2 + stepsCase3 + stepsCase4 + stepsCase5 + stepsCase6 + stepsCase7 + stepsCase8 

              # Write Header Row
              header = [f'{eventName} | {offerName} ']
              ws.append(header)

              # Merge Header Cells
              startColumnIndex = 3  # Example of a dynamic column index
              startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
              endColumn = "E"
              startRow = 1
              endRow = 1
              cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
              ws.merge_cells(cellRange)

              headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
              ws.append(headerRow)
              no = 1
              for num, step in enumerate(allSteps, start=1):
                     
                     if isinstance(step, str):
                            row = [
                                   no,
                                   step,
                                   "Success",
                                   "No Bonus",
                                   "XYZ"
                            ]
                            no = no+1
                     else:
                            if len(step) == 5:
                                   row = [
                                          step[0],
                                          step[1],
                                          step[2],
                                          step[3],
                                          step[4]
                                   ]
                            else:
                                   row = [
                                          no,
                                          step[0],
                                          step[1],
                                          "No Bonus",
                                          "XYZ"
                                   ]
                                   no = no+1
                     ws.append(row)

       print("Testing Scenario Successfully Generated")
       
       # Save Excel File
       wb.save('Result/Scenario '+str(eventName)+'.xlsx')

def exportExcelMUBALevel(eventName, params=None, neededParams = None):
       # Export Test Cases to Excel File
       wb = Workbook()
       ws = wb.active
       for params in params:
              if "Offer Name" in params:
                     offerName = params['Offer Name']
              else:
                     offerName = ''

              if "Offer Description" in params:
                     offerDesc = params['Offer Description']
              else:
                     offerDesc = ''

              if "Amount MU" in params:
                     amount = params['Amount MU']
                     amount = float(amount)
              else:
                     amount = ''

              if "Vascode" in params:
                     vascode = params['Vascode']
                     if (vascode == ''):
                            vascode = 'google'       
              else:
                     vascode = 'google'

              if "Amount For Flexible OC" in params:
                     rateOC = params['Amount For Flexible OC']
              else:
                     rateOC = amount * 0.05

              if "Amount For Flexible RC" in params:
                     rateRC = params['Amount For Flexible RC']
              else:
                     rateRC = amount * 0.05
                     

              formatted_amount = "{:,.0f}".format(amount).replace(",", ".")

              multiple = 10 ** (len(str(amount)) - 1)

              rounded_value = math.ceil(amount / multiple) * multiple

              formatted_value = "{:,.0f}".format(rounded_value).replace(",", ".")

              #For case Total Usage Subscriber 1 +2+ 3 = MU
              if (amount % 3 == 0):
                     MSISDN1Amount = amount // 3
                     MSISDN2Amount = amount // 3
                     MSISDN3Amount = amount - (MSISDN1Amount + MSISDN2Amount)
              elif (amount % 4 == 0):
                     MSISDN1Amount = amount // 4
                     MSISDN2Amount = amount // 4
                     MSISDN3Amount = amount - (MSISDN1Amount + MSISDN2Amount)
              

              MSISDN1PPN = MSISDN1Amount+(MSISDN1Amount*0.11)
              MSISDN2PPN = MSISDN2Amount+(MSISDN2Amount*0.11)
              MSISDN3PPN = MSISDN3Amount+(MSISDN3Amount*0.11)

              amountPPN = amount+(amount*0.11)


              steps = [
                     ["Create & Activate new 3 subscriber with OU on PP KartuHALO Corporate Business | 2437","Check active period"],
                     ["MSISDN 1 Update Parameter (Init activation date)","SUCCESS"],
                     ["MSISDN 1 Set New Credit Limit Service (offer id : 3669334) as "+str(formatted_value),"Offer Attached"],
                     ["MSISDN 1 Attach Offer "+str(offerName)+" with ParamValue "+str(formatted_amount)+" in ParamName PRIM MU Value","Offer Attached"],
                     [f"MSISDN 1 Attach Offer B2C Flexible Abonemen Charge with param name : Commitment period|Rate|Proration|Override RC amount|Invoice description|Quotation reference|External product id|Penalty Remaining|Penalty Flat|Penalty ind; Value : 0|{rateRC}|N|-1|Invoice description|Quotation reference|External product id|0|0|No","Offer Attached"],
                     ["MSISDN 1 Check 888","Checked"],
                     ["MSISDN 1 Check Offer Name & Offer Description",""+str(offerName)+"|"+str(offerDesc)+" (MU in BA Level)"],
                     ["MSISDN 1 D+0 5 PM, Create event voice onnet 60s","Charged 900 IDR"],
                     ["MSISDN 1 D+1 7 PM, Create event 1 SMS onnet","Charged 150 IDR"],
                     ["MSISDN 1 D+2 11 PM, Create event GPRS 1MB RG 55","Charged 6144 IDR"],
                     ["MSISDN 1 D+3 11 PM, Create event Direct Debit using vascode bank_digi_250","Charged 250 IDR"],
                     ["MSISDN 2 Update Parameter (Init activation date)","SUCCESS"],
                     ["MSISDN 2 Set New Credit Limit Service (offer id : 3669334) as "+str(formatted_value),"Offer Attached"],
                     ["MSISDN 2 Attach Offer New CLS International Roaming - 3669354 as 1000.000","Offer Attached"],
                     ["MSISDN 2 Attach Offer International Roaming 36327","Offer Attached"],
                     [f"MSISDN 2 Attach Offer B2C Flexible Abonemen Charge with param name : Commitment period|Rate|Proration|Override RC amount|Invoice description|Quotation reference|External product id|Penalty Remaining|Penalty Flat|Penalty ind; Value : 12|{rateOC}|N|0|Invoice description|Quotation reference|External product id|0|0|No","Offer Attached"],
                     ["MSISDN 2 Check charge 888","Checked"],
                     ["MSISDN 2 Create event 1 SMS onnet","Charged 150 IDR"],
                     ["MSISDN 2 D+1 11 PM,  Create Event Voice Roaming MOC Home in India (91) 120s","Charged 63636 IDR"],
                     ["MSISDN 2 D+2 9 PM, Create Event Voice Roaming MOC Local in Qatar (974) 120s","Charged 27272 IDR"],
                     ["MSISDN 3 Update Parameter (Init activation date)","SUCCESS"],
                     ["MSISDN 3 Set New Credit Limit Service (offer id : 3669334) as "+str(formatted_value),"Offer Attached"],
                     [f"MSISDN 3 Attach Offer B2C Flexible Abonemen Charge with param name : Commitment period|Rate|Proration|Override RC amount|Invoice description|Quotation reference|External product id|Penalty Remaining|Penalty Flat|Penalty ind; Value : 12|{rateOC}|N|0|Invoice description|Quotation reference|External product id|0|0|No","Offer Attached"],
                     ["MSISDN 3 Check charge 888","Checked"],
                     ["MSISDN 3 D+0 11 PM, Create event Direct Debit using vascode google 16.650.000","Charged 15000000 IDR"],
                     ["MSISDN 3 Remove Offer 3911924 MF Flash 22.5K D+1 ","Offer Removed"],
                     [f"Total Usage Subscriber 1 +2+ 3 < MU ({str(formatted_amount)})","Checked"],
                     ["MSISDN 1 Check MSISDN in Table TRB1_SUB_ERRS","Should be Empty"],
                     ["MSISDN 2 Check MSISDN in Table TRB1_SUB_ERRS","Should be Empty"],
                     ["MSISDN 3 Check MSISDN in Table TRB1_SUB_ERRS","Should be Empty"],
                     ["Check Indira","Checked"],
                     ["Check RBLog","Checked"],
                     ["Invoicing","OK"],
                     ["Create & Activate new 3 subscriber with OU on PP KartuHALO Corporate Business | 2437","Check active period"],
                     ["MSISDN 1 Update Parameter (Init activation date)","SUCCESS"],
                     ["MSISDN 1 Set New Credit Limit Service (offer id : 3669334) as "+str(formatted_value),"Offer Attached"],
                     ["MSISDN 1 Attach Offer International Roaming 36327","Offer Attached"],
                     [f"MSISDN 1 Attach Offer {offerName} with ParamValue {formatted_amount} in ParamName PRIM MU Value","Offer Attached"],
                     ["MSISDN 1 Check 888","Checked"],
                     [f"MSISDN 1 D+1 11 PM, Create event Direct Debit using vascode {vascode} {MSISDN1PPN}",f"Charged {MSISDN1Amount} IDR"],
                     ["MSISDN 2 Update Parameter (Init activation date)","SUCCESS"],
                     ["MSISDN 2 Set New Credit Limit Service (offer id : 3669334) as "+str(formatted_value),"Offer Attached"],
                     [f"MSISDN 2 D+1 11 PM, Create event Direct Debit using vascode {vascode} {MSISDN2PPN}",f"Charged {MSISDN2Amount} IDR"],
                     ["MSISDN 3 Update Parameter (Init activation date)","SUCCESS"],
                     ["MSISDN 3 Set New Credit Limit Service (offer id : 3669334) as "+str(formatted_value),"Offer Attached"],
                     [f"MSISDN 3 D+1 11 PM, Create event Direct Debit using vascode {vascode} {MSISDN3PPN}",f"Charged {MSISDN3Amount} IDR"],
                     [f"Total Usage Subscriber 1 +2+ 3 = MU ({str(formatted_amount)})","Checked"],
                     ["MSISDN 1 Check MSISDN in Table TRB1_SUB_ERRS","Should be Empty"],
                     ["MSISDN 2 Check MSISDN in Table TRB1_SUB_ERRS","Should be Empty"],
                     ["MSISDN 3 Check MSISDN in Table TRB1_SUB_ERRS","Should be Empty"],
                     ["Check Indira","Checked"],
                     ["Check RBLog","Checked"],
                     ["Invoicing","OK"],
                     ["Create & Activate new 3 subscriber with OU on PP KartuHALO Corporate Business | 2437","Check active period"],
                     ["MSISDN 1 Update Parameter (Init activation date)","SUCCESS"],
                     ["MSISDN 1 Set New Credit Limit Service (offer id : 3669334) as "+str(formatted_value),"Offer Attached"],
                     [f"MSISDN 1 Attach Offer {offerName} with ParamValue {formatted_amount} in ParamName PRIM MU Value","Offer Attached"],
                     ["MSISDN 1 Check 888","Checked"],
                     ["MSISDN 1 D+0 5 PM, Create event voice onnet 60s","Charged 900 IDR"],
                     ["MSISDN 1 D+1 7 PM, Create event 10 SMS onnet","Charged 1500 IDR"],
                     ["MSISDN 1 D+2 11 PM, Create event GPRS 1MB RG 55","Charged 6144 IDR"],
                     ["MSISDN 1 D+3 11 PM, Create event Direct Debit using vascode bank_digi_250","Charged 250 IDR"],
                     [f"MSISDN 1 D+3 11 PM, Create event Direct Debit using vascode {vascode} {amountPPN}",f"Charged {amount} IDR"],
                     ["MSISDN 2 Update Parameter (Init activation date)","SUCCESS"],
                     ["MSISDN 2 Set New Credit Limit Service (offer id : 3669334) as "+str(formatted_value),"Offer Attached"],
                     ["MSISDN 2 D+0 5 PM, Create event voice onnet 60s","Charged 900 IDR"],
                     ["MSISDN 2 D+1 7 PM, Create event 50 SMS onnet","Charged 7500 IDR"],
                     [f"MSISDN 2 D+2 11 PM, Create event Direct Debit using vascode {vascode} {amountPPN} ( total < MU )",f"Charged {amount} IDR"],
                     ["MSISDN 3 Update Parameter (Init activation date)","SUCCESS"],
                     ["MSISDN 3 Set New Credit Limit Service (offer id : 3669334) as "+str(formatted_value),"Offer Attached"],
                     [f"MSISDN 3 D+3 11 PM, Create event Direct Debit using vascode {vascode} {amountPPN} ( total > MU )",f"Charged {amount} IDR"],
                     [f"Total Usage Subscriber 1 +2+ 3 > MU ({str(formatted_amount)})","Checked"],
                     ["MSISDN 1 Check MSISDN in Table TRB1_SUB_ERRS","Should be Empty"],
                     ["MSISDN 2 Check MSISDN in Table TRB1_SUB_ERRS","Should be Empty"],
                     ["MSISDN 3 Check MSISDN in Table TRB1_SUB_ERRS","Should be Empty"],
                     ["Check Indira","Checked"],
                     ["Check RBLog","Checked"],
                     ["Invoicing","OK"]
              ]

              # Write Header Row
              header = [f'{eventName} | {offerName} ']
              ws.append(header)

              # Merge Header Cells
              startColumnIndex = 3  # Example of a dynamic column index
              startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
              endColumn = "E"
              startRow = 1
              endRow = 1
              cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
              ws.merge_cells(cellRange)

              headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
              ws.append(headerRow)

              for no, step in enumerate(steps):
                     no = no+1
                     if isinstance(step, str):
                            row = [
                                   no,
                                   step,
                                   "Success",
                                   "No Bonus",
                                   "XYZ"
                            ]
                     else:
                            row = [
                                   no,
                                   step[0],
                                   step[1],
                                   "No Bonus",
                                   "XYZ"
                            ]
                     ws.append(row)

       print("Testing Scenario Successfully Generated")
       
       # Save Excel File
       wb.save('Result/Scenario '+str(eventName)+'.xlsx')

def exportExcelUpdateAMDD(eventName, params=None, neededParams = None):
       # Export Test Cases to Excel File
       wb = Workbook()
       ws = wb.active
       for params in params:
              if "Offer Name" in params:
                     offerName = params['Offer Name']
              else:
                     offerName = ''

              if "Offer Description" in params:
                     offerDesc = params['Offer Description']
              else:
                     offerDesc = ''
              
              if "Offer ID" in params:
                     offerID = params['Offer ID']
              else:
                     offerID = ''

              if "Service Filter" in params:
                     serviceFilter = params['Service Filter']
              else:
                     serviceFilter = ''

              stepsCaseBefore = [
                     ["Create & Activate new subscriber PP KartuHALO Corporate Executive","Check active period"],
                     ["Update Parameter (Init activation date)","Parameter Updated"],
                     ["Set New Credit Limit Service (offer id : 3669334) as 1.000.000","Offer Attached"],
                     ["Set CLS Roaming (offer id : 3669354) as 20.000.000 ","Offer Attached"],
                     ["Attach offer SVC Outbound 1 ","Offer Attached"],
                     ["Check offer name and offer description","SVC Outbound 1|SVC Outbound"],
                     ["Check prit name","Checked"],
                     ["Create event voice onnet 8s, 11AM D+2 | Check rounded event should be 8s | Check Pricing Item ID | Check Service Filter","Charged IDR 60"],
                     ["Create event voice offnet 7s, 1PM D+2 | Check rounded event should be 7s | Check Pricing Item ID | Check Service Filter","Charged IDR 88"],
                     ["Create event voice PSTN 9s, 3PM D+2 | Check rounded event should be 9s | Check Pricing Item ID | Check Service Filter","Charged IDR 113"],
                     ["Create event voice FWA 8s, 5PM D+2 | Check rounded event should be 8s | Check Pricing Item ID | Check Service Filter","Charged IDR 100"],
                     ["Create event voice onnet 181s, 5PM D+2 | Check rounded event should be 181s | Check Pricing Item ID | Check Service Filter","Charged IDR 1358"],
                     ["Create event voice offnet 190s, 5PM D+2 | Check rounded event should be 190s | Check Pricing Item ID | Check Service Filter","Charged IDR 2375"],
                     ["Create event voice PSTN 187s, 5PM D+2 | Check rounded event should be 87s | Check Pricing Item ID | Check Service Filter","Charged IDR 2338"],
                     ["Create event voice FWA 191s, 5PM D+2 | Check rounded event should be 191s | Check Pricing Item ID | Check Service Filter","Charged IDR 2388"],
                     ["Create event voice International to Singapore (+65) 60s, 7AM D+3","Charged IDR 6364"],
                     ["Create event voice spesial number to 14045 60s, 9AM D+3","Charged IDR 1200"],
                     ["Create event 100KB GPRS RG17, 1PM D+4","Charged IDR 600"],
                     ["Create event 1 sms onnet 5AM D+4","Charged IDR 150"],
                     ["Remove offer SVC Outbound 1 ","Offer removed"],
                     ["Create event voice onnet 43s, D+4 1AM | Check rounded | Check Pricing Item ID | Check Service Filter","Charged IDR 516"],
                     ["Create event voice offnet 10s, D+4 2AM | Check rounded | Check Pricing Item ID | Check Service Filter","Charged IDR 200"],
                     ["Create event voice PSTN 60s, D+4 3AM | Check rounded | Check Pricing Item ID | Check Service Filter","Charged IDR 1200"],
                     ["Create event voice FWA 38s, D+4 4AM | Check rounded | Check Pricing Item ID | Check Service Filter","Charged IDR 760"],
                     ["Create event voice International to Singapore (+65) 60s, D+4 6AM","Charged IDR 6364"],
                     ["Check Cycle Month","Checked"],
                     ["Check INDIRA ","Checked"],
                     ["Check RB Log ","Checked"],
                     ["Check Rated event (Before tax 11%) ","Checked"],
                     ["Invoicing","Checked"],
 
              ]

              stepsCaseAfter = [
                     ["Create & Activate new subscriber PP KartuHALO Corporate Executive","Check active period"],
                     ["Update Parameter (Init activation date)","Parameter Updated"],
                     ["Set New Credit Limit Service (offer id : 3669334) as 1.000.000","Offer Attached"],
                     ["Set CLS Roaming (offer id : 3669354) as 20.000.000 ","Offer Attached"],
                     [f"Attach offer {offerName} ","Offer Attached"],
                     ["Check offer name and offer description",f"{offerName}|{offerDesc}"],
                     ["Check prit name","Checked"],
                     [f"Create event voice onnet 8s, 11AM D+2 | Check rounded event should be 8s | Check Pricing Item ID | Check Service Filter : {serviceFilter}","Charged IDR 60"],
                     [f"Create event voice offnet 7s, 1PM D+2 | Check rounded event should be 7s | Check Pricing Item ID | Check Service Filter : {serviceFilter}","Charged IDR 88"],
                     [f"Create event voice PSTN 9s, 3PM D+2 | Check rounded event should be 9s | Check Pricing Item ID | Check Service Filter : {serviceFilter}","Charged IDR 113"],
                     [f"Create event voice FWA 8s, 5PM D+2 | Check rounded event should be 8s | Check Pricing Item ID | Check Service Filter : {serviceFilter}","Charged IDR 100"],
                     [f"Create event voice onnet 181s, 5PM D+2 | Check rounded event should be 181s | Check Pricing Item ID | Check Service Filter : {serviceFilter}","Charged IDR 1358"],
                     [f"Create event voice offnet 190s, 5PM D+2 | Check rounded event should be 190s | Check Pricing Item ID | Check Service Filter : {serviceFilter}","Charged IDR 2375"],
                     [f"Create event voice PSTN 187s, 5PM D+2 | Check rounded event should be 87s | Check Pricing Item ID | Check Service Filter : {serviceFilter}","Charged IDR 2338"],
                     [f"Create event voice FWA 191s, 5PM D+2 | Check rounded event should be 191s | Check Pricing Item ID | Check Service Filter : {serviceFilter}","Charged IDR 2388"],
                     ["Create event voice International to Singapore (+65) 60s, 7AM D+3","Charged IDR 6364"],
                     ["Create event voice spesial number to 14045 60s, 9AM D+3","Charged IDR 1200"],
                     ["Create event 100KB GPRS RG17, 1PM D+4","Charged IDR 600"],
                     ["Create event 1 sms onnet 5AM D+4","Charged IDR 150"],
                     [f"Remove offer {offerName} ","Offer removed"],
                     [f"Create event voice onnet 43s, D+4 1AM | Check rounded | Check Pricing Item ID | Check Service Filter : {serviceFilter}","Charged IDR 516"],
                     [f"Create event voice offnet 10s, D+4 2AM | Check rounded | Check Pricing Item ID | Check Service Filter : {serviceFilter}","Charged IDR 200"],
                     [f"Create event voice PSTN 60s, D+4 3AM | Check rounded | Check Pricing Item ID | Check Service Filter : {serviceFilter}","Charged IDR 1200"],
                     [f"Create event voice FWA 38s, D+4 4AM | Check rounded | Check Pricing Item ID | Check Service Filter : {serviceFilter}","Charged IDR 760"],
                     ["Create event voice International to Singapore (+65) 60s, D+4 6AM","Charged IDR 6364"],
                     ["Check Cycle Month","Checked"],
                     ["Check INDIRA ","Checked"],
                     ["Check RB Log ","Checked"],
                     ["Check Rated event (Before tax 11%) ","Checked"],
                     "Invoicing",
              ]

              allSteps = stepsCaseAfter

              # Write Header Row
              header = [f'{offerID} - {offerName} ']
              ws.append(header)

              # Merge Header Cells
              startColumnIndex = 3  # Example of a dynamic column index
              startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
              endColumn = "E"
              startRow = 1
              endRow = 1
              cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
              ws.merge_cells(cellRange)

              headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
              ws.append(headerRow)

              no = 1
              for num, step in enumerate(allSteps, start=1):
                     
                     if isinstance(step, str):
                            row = [
                                   no,
                                   step,
                                   "Success",
                                   "No Bonus",
                                   "XYZ"
                            ]
                            no = no+1
                     else:
                            if len(step) == 5:
                                   row = [
                                          step[0],
                                          step[1],
                                          step[2],
                                          step[3],
                                          step[4]
                                   ]
                            else:
                                   row = [
                                          no,
                                          step[0],
                                          step[1],
                                          "No Bonus",
                                          "XYZ"
                                   ]
                                   no = no+1
                     ws.append(row)

       print("Testing Scenario Successfully Generated")
       
       # Save Excel File
       wb.save('Result/Scenario '+str(eventName)+'.xlsx')

def exportExcelNewPP(eventName, params=None, neededParams = None):
       # Export Test Cases to Excel File
       wb = Workbook()
       ws = wb.active

       #Defined Param
       type = ''
       name = ''
       case = ''
       rounded = 0
       rateOnnet = 0
       rateOffnet = 0
       rateLokalPTSN = 0
       PPTo = ''
       unit = ''
       rate = 0
       welcomeMessage = ''
       preloadBonus = ''
       preloadBalance = 0
       validity = 0
       preloadBonusPPTo = ''
       ratePerZone = ''
       cardType = ''
       roundedUnit = ''
       landingPage = ''
       bonusData = ''
       remainingAllowance = 0

       if isinstance(params, list):
              for params in params:
                     if "Type" in params:
                            type = params['Type'][0]

                     if "PP Name" in params:
                            name = params["PP Name"]
                     
                     if f"Case ({type})" in params:
                            case = params[f"Case ({type})"][0]
                     
                     if "Rounded (Time Unit)" in params:
                            rounded = int(params["Rounded (Time Unit)"])
                     
                     if "Rate Onnet" in params:
                            rateOnnet = int(params["Rate Onnet"])

                     if "Rate Offnet" in params:
                            rateOffnet = int(params["Rate Offnet"])

                     if "PP To" in params:
                            PPTo = params["PP To"]
                     
                     if "Rate Lokal PSTN" in params:
                            rateLokalPTSN = int(params["Rate Lokal PSTN"])

                     if "Unit" in params:
                            unit = params["Unit"]
                     
                     if "Rate" in params:
                            rate = int(params['Rate'])

                     if "Welcome Message" in params:
                            welcomeMessage = params['Welcome Message']
                     
                     if "Preload Bonus" in params:
                            preloadBonus = params["Preload Bonus"]

                     if "Preload Balance" in params:
                            preloadBalance = params["Preload Balance"]
                     
                     if "Validity" in params:
                            validity = int(params["Validity"])

                     if "Preload Bonus (For PP TO)" in params:
                            preloadBonusPPTo = params["Preload Bonus (For PP TO)"]

                     if "Rate Per Zone" in params:
                            ratePerZone = params["Rate Per Zone"]

                     if "Card Type" in params:
                            cardType = params["Card Type"]
                     
                     if "Rounded/Unit" in params:
                            roundedUnit = params["Rounded/Unit"]
                     
                     if "Landing Page" in params:
                            landingPage = params["Landing Page"]
                     
                     if "Rate Nol" in params:
                            bonusData = params["Rate Nol"]
                     
                     if "Remaining Allowance" in params:
                            remainingAllowance = params["Remaining Allowance"]

                     
                     if type == 'Postpaid':
                            if case == 'Tarif Voice':
                                   steps = tarifVoicePostpaid(name, rounded, rateOnnet, rateOffnet, rateLokalPTSN)
                            elif case == 'Offline Event':
                                   steps = offlineEventPostpaid(name)
                            elif case == 'PAYU Capped GPRSROAM':
                                   steps = GPRSRoamPostpaid(name)
                            elif case == 'Attach offer RC, OC, and Flexible Offer':
                                   steps = AttachOfferPostpaid(name)
                            elif case == 'Tarif SMS':
                                   steps = tarifSMSPostpaid(name, rateOnnet, rateOffnet)
                            elif case == 'Change PP':
                                   steps = changePPPostpaid(name, PPTo)
                            elif case == 'GPRS & Landing Page':
                                   steps = GPRSLandingPagePostpaid(name, unit, rate)
                            else:
                                   print ("Sorry the data is not ready!!")
                                   exit()
                     else:
                            if case == 'Attach offer & diameter':
                                   steps = attachOfferPrepaid(name)
                            elif case == "Preload Bonus":
                                   steps = preloadBonusPrepaid(name, welcomeMessage, preloadBonus, preloadBalance, validity)
                            elif case == "PAYU Capped GPRSROAM":
                                   steps = GPRSRoamPrepaid(name)
                            elif case == 'Split Recharge':
                                   steps = splitRechargePrepaid(name, welcomeMessage)
                            elif case == "Change PP":
                                   steps = changePPPrepaid(name, PPTo, preloadBonus, preloadBonusPPTo)
                            elif case == "Tarif SMS":
                                   steps = tarifSMSPrepaid(name, ratePerZone)
                            elif case == 'GPRS & Landing Page':
                                   steps = GPRSLandingPagePrepaid(name, cardType, roundedUnit, rate, landingPage, bonusData, remainingAllowance)
                            else:
                                   print ("Sorry the data is not ready!!")
                                   exit()

                     # Write Header Row
                     header = [f'{eventName} | {name} | {case}']
                     ws.append(header)

                     # Merge Header Cells
                     startColumnIndex = 3  # Example of a dynamic column index
                     startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
                     endColumn = "E"
                     startRow = 1
                     endRow = 1
                     cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
                     ws.merge_cells(cellRange)

                     headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
                     ws.append(headerRow)

                     no = 1
                     for num, step in enumerate(steps, start=1):   
                            if isinstance(step, str):
                                   row = [
                                          no,
                                          step,
                                          "Success",
                                          "No Bonus",
                                          "XYZ"
                                   ]
                                   no = no+1
                            else:
                                   if len(step) == 5:
                                          row = [
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 step[3],
                                                 step[4]
                                          ]
                                   elif len(step) == 4:
                                          row = [
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 step[3],
                                                 "XYZ"
                                          ] 
                                   elif len(step) == 3:
                                          row = [
                                                 no,
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 "XYZ"
                                          ]
                                          no = no+1
                                   else:
                                          row = [
                                                 no,
                                                 step[0],
                                                 step[1],
                                                 "No Bonus",
                                                 "XYZ"
                                          ]
                                          no = no+1
                            ws.append(row)

              print("Testing Scenario Successfully Generated")
              
              # Save Excel File
              wb.save(f'Result/Scenario {eventName} {type} {name} {case}.xlsx')
       else:
              if "Type" in params:
                     type = params['Type'][0]

              if "PP Name" in params:
                     name = params["PP Name"]
              
              if f"Case ({type})" in params:
                     case = params[f"Case ({type})"][0]
              
              if "Rounded (Time Unit)" in params:
                     rounded = int(params["Rounded (Time Unit)"])
              
              if "Rate Onnet" in params:
                     rateOnnet = int(params["Rate Onnet"])

              if "Rate Offnet" in params:
                     rateOffnet = int(params["Rate Offnet"])
              
              if "Rate Lokal PSTN" in params:
                     rateLokalPTSN = int(params["Rate Lokal PSTN"])
              
              if "PP To" in params:
                     PPTo = params["PP To"]
              
              if "Unit" in params:
                     unit = params["Unit"]
              
              if "Rate" in params:
                     rate = int(params['Rate'])

              if "Welcome Message" in params:
                     welcomeMessage = int(params['Welcome Message'])
              
              if "Preload Bonus" in params:
                     preloadBonus = params["Preload Bonus"]

              if "Preload Balance" in params:
                     preloadBalance = params["Preload Balance"]
              
              if "Validity" in params:
                     validity = int(params["Validity"])
              
              if "Preload Bonus (For PP TO)" in params:
                     preloadBonusPPTo = params["Preload Bonus (For PP TO)"]
              
              if "Rate Per Zone" in params:
                     ratePerZone = params["Rate Per Zone"]

              if "Card Type" in params:
                     cardType = params["Card Type"]
              
              if "Rounded/Unit" in params:
                     roundedUnit = params["Rounded/Unit"]
              
              if "Landing Page" in params:
                     landingPage = params["Landing Page"]
              
              if "Rate Nol" in params:
                     bonusData = params["Rate Nol"]
              
              if "Remaining Allowance" in params:
                     remainingAllowance = params["Remaining Allowance"]

              if type == 'Postpaid':
                     if case == 'Tarif Voice':
                            steps = tarifVoicePostpaid(name, rounded, rateOnnet, rateOffnet, rateLokalPTSN)
                     elif case == 'Offline Event':
                            steps = offlineEventPostpaid(name)
                     elif case == 'PAYU Capped GPRSROAM':
                            steps = GPRSRoamPostpaid(name)
                     elif case == 'Attach offer RC, OC, and Flexible Offer':
                            steps = AttachOfferPostpaid(name)
                     elif case == 'Tarif SMS':
                            steps = tarifSMSPostpaid(name, rateOnnet, rateOffnet)
                     elif case == 'Change PP':
                            steps = changePPPostpaid(name, PPTo)
                     elif case == 'GPRS & Landing Page':
                            steps = GPRSLandingPagePostpaid(name, unit, rate)
                     else:
                            print ("Sorry the data is not ready!!")
                            exit()
              else:
                     if case == 'Attach offer & diameter':
                            steps = attachOfferPrepaid(name)
                     elif case == "Preload Bonus":
                            steps = preloadBonusPrepaid(name, welcomeMessage, preloadBonus, preloadBalance, validity)
                     elif case == "PAYU Capped GPRSROAM":
                            steps = GPRSRoamPrepaid(name)
                     elif case == 'Split Recharge':
                            steps = splitRechargePrepaid(name, welcomeMessage)
                     elif case == "Change PP":
                            steps = changePPPrepaid(name, PPTo, preloadBonus, preloadBonusPPTo)
                     elif case == 'GPRS & Landing Page':
                            steps = GPRSLandingPagePrepaid(name, cardType, roundedUnit, rate, landingPage, bonusData, remainingAllowance)
                     else:
                            print ("Sorry the data is not ready!!")
                            exit()

              # Write Header Row
              header = [f'{eventName} | {name} | {case}']
              ws.append(header)

              # Merge Header Cells
              startColumnIndex = 3  # Example of a dynamic column index
              startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
              endColumn = "E"
              startRow = 1
              endRow = 1
              cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
              ws.merge_cells(cellRange)

              headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
              ws.append(headerRow)

              no = 1
              for num, step in enumerate(steps, start=1):   
                     if isinstance(step, str):
                            row = [
                                   no,
                                   step,
                                   "Success",
                                   "No Bonus",
                                   "XYZ"
                            ]
                            no = no+1
                     else:
                            if len(step) == 5:
                                   row = [
                                          step[0],
                                          step[1],
                                          step[2],
                                          step[3],
                                          step[4]
                                   ]
                            elif len(step) == 4:
                                   row = [
                                          step[0],
                                          step[1],
                                          step[2],
                                          step[3],
                                          "XYZ"
                                   ] 
                            elif len(step) == 3:
                                   row = [
                                          no,
                                          step[0],
                                          step[1],
                                          step[2],
                                          "XYZ"
                                   ]
                                   no = no+1
                            else:
                                   row = [
                                          no,
                                          step[0],
                                          step[1],
                                          "No Bonus",
                                          "XYZ"
                                   ]
                                   no = no+1
                     ws.append(row)

              print("Testing Scenario Successfully Generated")
              
              # Save Excel File
              wb.save(f'Result/Scenario {eventName} {type} {name} {case}.xlsx')

def tarifVoicePostpaid(name, rounded, rateOnnet, rateOffnet, rateLokalPTSN):
       zone = ["zone 1", "zone 2", "zone 4"]
       checkedStep = []
       for val in zone:
              randCond1 = random.randint(1, rounded)
              randCond2 = rounded
              randCond3 = random.randint(rounded, rounded*2)
              valOnnet  = rounded*rateOnnet
              valOnnet2 = (rounded*2)*rateOnnet
              valOffnet  = rounded*rateOffnet
              valOffnet2 = (rounded*2)*rateOffnet
              valLokalPTSN  = rounded*rateLokalPTSN
              valLokalPTSN2 = (rounded*2)*rateLokalPTSN
              steps = [
                     [f"Create Event Voice {randCond1}s Onnet on {val} | Check rounded should be {rounded}s",f"Charged {valOnnet}"],
                     [f"Create Event Voice {randCond2}s Onnet on {val} | Check rounded should be {rounded}s",f"Charged {valOnnet}"],
                     [f"Create Event Voice {randCond3}s Onnet on {val} | Check rounded should be {rounded*2}s",f"Charged {valOnnet2}"],
                     [f"Create Event Voice {randCond1}s Offnet on {val} | Check rounded should be {rounded}s",f"Charged {valOffnet}"],
                     [f"Create Event Voice {randCond2}s Offnet on {val} | Check rounded should be {rounded}s",f"Charged {valOffnet}"],
                     [f"Create Event Voice {randCond3}s Offnet on {val} | Check rounded should be {rounded*2}s",f"Charged {valOffnet2}"],
                     [f"Create Event Voice {randCond1}s PSTN on {val} | Check rounded should be {rounded}s",f"Charged {valLokalPTSN}"],
                     [f"Create Event Voice {randCond2}s PSTN on {val}| Check rounded should be {rounded}s",f"Charged {valLokalPTSN}"],
                     [f"Create Event Voice {randCond3}s PSTN on {val}| Check rounded should be {rounded*2}s",f"Charged {valLokalPTSN2}"],
                     [f"Create Event Voice {randCond1}s FWA on {val} | Check rounded should be {rounded}s",f"Charged {valOffnet}"],
                     [f"Create Event Voice {randCond2}s FWA on {val}| Check rounded should be {rounded}s",f"Charged {valOffnet}"],
                     [f"Create Event Voice {randCond3}s FWA on {val}| Check rounded should be {rounded*2}s",f"Charged {valOffnet2}"],
              ]
              checkedStep.extend(steps)

       step = [
                     [f"Create & Activate new subscriber PP {name}","Check active period"],
                     ["Check PP name (PP name based on BRD)","Checked"],
                     ["Update Parameter","Parameter Updated"],
                     ["Set New Credit Limit Service (offer id : 3669334) as 1.000.000 | 3669334","Offer Attached"],
                     ["Set New Credit Limit Service Roaming (offer id : 3669354) as 10.000.000","Offer Attached"],
                     ["Attach Offer International Roaming - 36327","Offer Attached"],
                     ["Check 888","Checked"],
                     ["Check 889","Checked"],
                     ["Check GetBonusInfo","Checked"],
                     #checked Step
                     ["Create event voice International to Singapore (+65) 15s","Charged"],
                     ["Create event voice spesial number to 14045 15s","Charged"],
                     ["Create event voice roaming MT Home in Qatar (+974) 15s","Charged"],
                     ["Create event voice roaming MT Local in Qatar (+974) 15s","Charged"],
                     ["Create event voice roaming MT Other in Qatar (+974) to Malaysia (+60) 15s","Charged"],
                     ["Create event voice roaming MO Home from Qatar (+974) 15s","Charged"],
                     ["Create event voice roaming MO Local from Qatar (+974) 15s","Charged"],
                     ["Create event voice roaming MO Other from Qatar (+974) to Singapore (+65) 15s, 5PM","Charged"],
                     ["Check cycle month","Checked"],
                     ["Check RB log","Checked"],
                     ["Check Indira","Checked"],
                     ["Check Rated Event","Checked"],

       ]

       index = step.index(['Create event voice International to Singapore (+65) 15s',"Charged"])
       step[index:index] = checkedStep

       return step

def offlineEventPostpaid(name):
       step = [
              [f"Create & Activate new subscriber PP {name}","Check active period"],
              ["Update Parameter","Parameter Updated"],
              ["Set New Credit Limit Service (offer id : 3669334) as 5.000.000 | 3669334","Offer Attached"],
              ["Set CLS Roaming (offer id : 3669354) as 10.000.000","Offer Attached"],
              ["Attach Offer International Roaming","Offer Attached"],
              ["Create offline event Voice MOC Home 60s","Charged"],
              ["Create offline event Voice MOC Local 60s","Charged"],
              ["Create offline event Voice MOC Other 60s","Charged"],
              ["Create offline event Voice MTC Home 60s","Charged"],
              ["Create offline event Voice MTC Local 60s","Charged"],
              ["Create offline event Voice MTC Other 60s","Charged"],
              ["Check Rated Event","Checked"],
              ["Check Rejected Event","Checked"],
              "Do Offline Event"	 
       ]

       return step

def GPRSRoamPostpaid(name):
       step = [
              [f"Create & Activate new subscriber {name}","Check active period"],
              ["Update Parameter (Init activation date)","Parameter Updated"],
              ["Set New Credit Limit Service (offer id : 3669334) as 10.000.000 IDR","CLS Added"],
              ["Set CLS International (offer id : 3669354) as 300.000.000 IDR","CLS Added"],
              ["Attach Offer International Roaming","Offer Attached"],
              ["Create event GPRS Roaming with operator name SAUZN 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name SAUAJ 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name SAUET 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BRATM 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BRABT 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BRASP 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BRACS 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BRARN 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name USAW3 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name USAW4 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name USAW5 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name USAW6 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name CHLMV 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name CHLSM 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name CHLTM 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name CRICL 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name CRITC 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name ECUOT 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name ARGCM 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name ARGGS 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name ARGNC 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name ARGTM 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name ARGTP 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name ARGVO 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name COLCM 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name COLCO 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name COLTI 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name COLTM 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name GHAGM 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name GHAGT 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name GHAMT 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name GHASC 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name GHAZN 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name PRYGS 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name PRYHT 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name PRYNP 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name PRYTC 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name PRYVX 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name URYAM 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name URYAN 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name URYGS 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name URYTM 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name VEN01 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name VEND2 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name VEND3 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name VENGS 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name VENMS 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name VENMV 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name MARM1 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name MARM3 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name MARMT 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name DZAA1 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name DZAOT 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name AGOUT 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name JAMDC 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name ATG03 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BHSBH 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BEN55 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BENSP 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BEN02 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BMU01 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BMUNI 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BMUBD 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BOLME 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BRACL 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BFACT 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name BDI02 1mb","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name KHMST 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name KHMSC 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name KHMVT 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name CHNCT 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name CHNTD 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name CHNCM 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name CHNCU 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name HKGTC 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name HKGHT 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name HKGH3 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name HKGMC 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name HKGNW 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name HKGPP 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name HKGSM 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name HKGM3 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name IND23 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name IND19 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name IND22 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name IND20 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name IND21 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name INDRC 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name IND09 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name INDSC 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name INDA3 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name INDA5 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name INDBL 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name INDA7 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name INDA8 1mb D+1","Charge 1802 IDR"],
              ["Create event GPRS Roaming with operator name INDA2 1mb D+1","Charge 0"],
              ["Create event GPRS Roaming with operator name INDA1 1mb D+1","Charge 0"],
              ["Create event GPRS Roaming with operator name INDA9 1mb D+2","Charge 0"],
              ["Create event GPRS Roaming with operator name INDA4 1mb D+2","Charge 0"],
              ["Create event GPRS Roaming with operator name INDA6 1mb D+2","Charge 0"],
              ["Create event GPRS Roaming with operator name INDJH 1mb D+2","Charge 0"],
              ["Create event GPRS Roaming with operator name INDJB 1mb D+2","Charge 0"],
              ["Create event GPRS Roaming with operator name INDAT 1mb D+2","Charge 0"],
              ["Create event GPRS Roaming with operator name INDMT 1mb D+2","Charge 0"],
              ["Create event GPRS Roaming with operator name INDH1 1mb D+2","Charge 0"]
       ]

       return step

def AttachOfferPostpaid(name):
       step = [
              [f"Create & Activate new subscriber PP New Price {name} (2437)","Actived"],
              ["Update Parameter (Init activation date)","Expration Update"],
              ["Set New Credit Limit Service (offer id : 3669334) as 2.000.000 IDR","Balance Updated"],
              ["Attach offer (OC Offer ) 3890329|Qlue Dashboard Yearly","Offer Attached"],
              ["Check Offer Name Disc & Offer Description Disc","Checked"],
              ["Check AMDD Code and Charge","Checked"],
              ["Create event voice onnet 300s 11am","Charged"],
              ["Create event 10 sms onnet 1pm","Charged"],
              ["Create Event GPRS 1MB RG 50 8pm","Charged"],
              ["Create event direct devit vascode google 10k 10pm","Charged"],
              ["Check 888","Checked"],
              ["Check in trb1_sub_errs","No error"],
              ["Check Indira (CHG)","Checked"],
              ["Invoicing","Checked"],
              [f"Create & Activate new subscriber PP New Price {name} (2437)","Actived"],
              ["Update Parameter (Init activation date)","Expration Update"],
              ["Set New Credit Limit Service (offer id : 3669334) as 2.000.000 IDR","Balance Updated"],
              ["Attach offer (RC  Offer ) 3888449|Halo Kick Premium 225rb","Offer Attached"],
              ["Check Offer Name Disc & Offer Description Disc","Checked"],
              ["Check AMDD Code and Charge","Checked"],
              ["Create event voice onnet 300s 11am","Charged"],
              ["Create event 10 sms onnet 1pm","Charged"],
              ["Create Event GPRS 1MB RG 50 8pm","Charged"],
              ["Create event direct devit vascode google 10k 10pm","Charged"],
              ["Check 888","Checked"],
              ["Check in trb1_sub_errs","No error"],
              ["Check Indira (CHG)","Checked"],
              ["Invoicing","Checked"],
              [f"Create & Activate new subscriber PP New Price {name} (2437)","Actived"],
              ["Update Parameter (Init activation date)","Expration Update"],
              ["Set New Credit Limit Service (offer id : 3669334) as 2.000.000 IDR","Balance Updated"],
              ["Attach offer (Flexible Offer RC Charged 3875629) with param : Commitment period|Rate|Proration|Override RC amount|Invoice description|Quotation reference|External product id|Penalty Remaining|Penalty Flat|Penalty ind - 0|200000|P|-1|Invoice description|Quotation reference|External product id|0|0|N","Offer Attached"],
              ["Check Offer Name Disc & Offer Description ","Checked"],
              ["Check AMDD Code and Charge","Checked"],
              ["Create event voice onnet 300s 11am","Consume Bonus"],
              ["Create event 10 sms onnet 1pm","Charged"],
              ["Create Event GPRS 1MB RG 50 8pm","Charged"],
              ["Create event direct devit vascode google 10k 10pm","Charged"],
              ["Check 888","Checked"],
              ["Check in trb1_sub_errs","No error"],
              ["Check Indira (CHG)","Checked"],
              ["Invoicing","Checked"]
       ]

       return step

def tarifSMSPostpaid(name, rateOnnet, rateOffnet):
       zone = ["zone 1", "zone 2", "zone 4"]
       checkedStep = []
       for val in zone:
              randCondOn = random.randint(1, 20)
              randCondOff = random.randint(1, 20)
              valOnnet  = randCondOn*rateOnnet
              valOffnet  = randCondOff*rateOffnet
              steps = [
                     [f"Create event 1 SMS onnet on zone ID {val}",f"Charged {rateOnnet}"],
                     [f"Create event {randCondOn} SMS onnet on zone ID {val}",f"Charged {valOnnet}"],
                     [f"Create event 1 SMS offnet on zone ID {val}",f"Charged {rateOffnet}"],
                     [f"Create event {randCondOff} SMS offnet on zone ID {val}",f"Charged {valOffnet}"],
              ]
              checkedStep.extend(steps)

       step = [
                     [f"Create & Activate new subscriber New Price {name}","Check active period"],
                     ["Check Message Code & Welcome Message","MessageWording : Registrasikan kartu prabayar kamu, kirimkan SMS dengan format: REG<spasi>NIK#NoKK# ke 4444 atau kunjungi tsel.me/daftar ### MessageCode : PREREG"],
                     ["Check initial balance","0 IDR"],
                     ["Check Exp Date MSISDN","Checked"],
                     ["Check bonus using L9GetBonusInfo","Success"],
                     ["Update Balance 500K","Success"],
                     ["Check bonus 889*1","Success"],
                     ["Check bonus 889*2","Success"],
                     ["Check bonus 889*3","Success"],

                     #checked Step
                     ["Create event 1 SMS international to Malaysia (+60), 1PM D+10","Charged Not Based on BRD"],
                     ["Create event 1 SMS international to Singgapore (+65), 2PM D+10","Charged Not Based on BRD"],
                     ["Create event 1 SMS international to Argentina (+54), 3PM D+10","Charged Not Based on BRD"],
                     ["Create event 1 SMS international to Brasil (+55), 4PM D+10","Charged Not Based on BRD"],
                     ["Create event 1 SMS international to Africa Selatan (+27), 5PM D+10","Charged Not Based on BRD"],
                     ["Create event 1 SMS international to Nigeria (+234), 6PM D+10","Charged Not Based on BRD"],
                     ["Create event direct debit using vascode rw_asia_3in1_7000, 1PM D+10","Charged Not Based on BRD"],
                     ["Create event direct debit using vascode google 55000 5PM D+10","Charged Not Based on BRD"],
                     ["Create event 1 SMS spesial number 69888, 5PM D+10","Charged Not Based on BRD"],
                     ["Create event 1 SMS international to Canada (+1), 5PM D+11","Charged Not Based on BRD"],
                     ["Create event 1 SMS international to Kazakhstan (+7), 8PM D+11","Charged Not Based on BRD"],
                     ["Create event 1 MMS International to Malaysia (+60), 9PM D+11","Charged Not Based on BRD"],
                     ["Create event 1 MMS International to Singgapore (+65), 10PM D+11","Charged Not Based on BRD"],
                     ["Check Indira","Checked"],
                     ["Check RB Log","Checked"],
                     ["Check Rated Event","Checked"]
       ]

       index = step.index(["Create event 1 SMS international to Malaysia (+60), 1PM D+10","Charged Not Based on BRD"])
       step[index:index] = checkedStep

       return step

def changePPPostpaid(name, PPTo):
       current_month = datetime.datetime.now().month
       step = [
              [f"Create & Actived MSISDN PP {name}","SUCCESS"],
              ["Update Parameter (Init activation date)","SUCCESS"],
              ["Set New Credit Limit Service (offer id : 3669334) as 10.000.000","SUCCESS"],
              ["Set CLS Roaming (offer id : 3669354) as 20.000.000","SUCCESS"],
              [f"Change PP to {PPTo}",f"PP {name} to PP {PPTo}"],
              ["Attach Offer Postpaid Orbit 50GB 80rb","Offer Attached"],
              ["Check 888","Checked"],
              ["Create event voice onnet 600s","Charged"],
              ["Create event voice offnet 600s","Charged"],
              ["Voice PSTN event 600s","Charged"],
              ["Voice FWA event 600s","Charged"],
              ["Create event 1 SMS onnet","Charged"],
              ["Create event 1 SMS offnet","Charged"],
              ["Create event GPRS 1MB RG 55","Charged"],
              ["Create event MMS onnet","Charged"],
              ["Create event MMS offnet","Charged"],
              ["Create event direct debit with vascode google (100k)","Charged"],
              ["Create event voice international to Malaysia (60) 60s","Charged"],
              ["Create event 1 SMS International to Malaysia (60)","Charged"],
              ["Create event GPRS 1MB RG 55","Charged"],
              ["Do Remove Offer Postpaid Orbit 50GB 80rb","Offer Removed"],
              ["Check notification","Checked"],
              ["Check cycle month",current_month],
              ["Check TRB1_SUB_Errs","Checked"],
              ["Check table RC/OC","Checked"]
       ]

       return step

def GPRSLandingPagePrepaid(name, cardType, roundedUnit, rate, landingPage, bonusData, remainingAllowance):
       cardType                    = cardType[0]
       roundedUnit                 = roundedUnit.split('/')
       unit                        = roundedUnit[1]
       bonusDataSplit              = bonusData.split(' ')
       unitBonusData               = bonusDataSplit[1]
       amountBonusData             = bonusDataSplit[0]
       rate                        = int(rate)
       
       #change rounded to kb (rate is per rounded)
       if unit.lower() == 'kb':
              rounded       = roundedUnit[0]
              # rate          = rate      
       elif unit.lower() == 'mb':
              rounded       = int(roundedUnit[0])*1024
              # rate          = rate/1024
       elif unit.lower() == 'gb':
              rounded       = int(roundedUnit[0])*1048576
              # rate          = rate/1048576
       
       #change Rate Nol to kb
       if unitBonusData.lower() == 'kb':
              amountBonusData      = amountBonusData   
       elif unitBonusData.lower() == 'mb':
              amountBonusData      = int(amountBonusData)*1024
       elif unitBonusData.lower() == 'gb':
              amountBonusData      = int(amountBonusData)*1048576 
       
       if cardType == 'Telkomsel Prabayar':
              checkedDaily  = []
              day           = 1
              remainingAllowanceSplit     = remainingAllowance.split(';')
              unitRemainingAllowance      = remainingAllowanceSplit[1]
              amountRemainingAllowance    = remainingAllowanceSplit[0]
              
              #change remaining allowance to kb
              if unitRemainingAllowance.lower() == 'kb':
                     amountRemainingAllowance      = amountRemainingAllowance   
              elif unitRemainingAllowance.lower() == 'mb':
                     amountRemainingAllowance      = int(amountRemainingAllowance)*1024
              elif unitRemainingAllowance.lower() == 'gb':
                     amountRemainingAllowance      = int(amountRemainingAllowance)*1048576

              while day <= 2:
                     landingPageSteps = []
                     landingPageSplit = landingPage.split(",")
                     landingPageAmount= rate
                     for landingPages in landingPageSplit:
                            landingPagesSplit           = landingPages.split(';')
                            landingPageBorder           = landingPagesSplit[0]
                            landingPageErrorCode        = landingPagesSplit[1]
                            intermediateLandingPage     = math.ceil((int(landingPageBorder) - int(landingPageAmount)) * (float(rounded) / rate))
                            chargedLandingPage          = intermediateLandingPage/int(rounded)*int(rate)
                            landingPageAmount           += chargedLandingPage

                            intermediateSteps           = []
                            while intermediateLandingPage > 0:
                                   if intermediateLandingPage > 512:
                                          page_count    = 512
                                          errorPage     = 'Success'
                                   else:
                                          page_count    = intermediateLandingPage
                                          errorPage     = "Error code "+str(landingPageErrorCode)

                                   intermediateStep = [
                                          ["Create event GPRS Intermediate "+str(page_count)+" "+str(unit)+" RG 55 at 11AM", errorPage, "No Bonus"],
                                   ]
                                   intermediateSteps.extend(intermediateStep)
                                   intermediateLandingPage -= page_count

                            stepLandingPage = [
                                   ["Create event GPRS Initial "+str(landingPageBorder)+" notif RG 55 at 11AM", "Success", "No Bonus"],
                            ]
                            stepLandingPageTerminate = [
                                   ["Create event GPRS Terminate 0Kb RG 55 at 11AM", "Charged "+str(chargedLandingPage)+" IDR", "No Bonus"],
                            ]
                            stepLandingPage.extend(intermediateSteps)
                            stepLandingPage.extend(stepLandingPageTerminate)

                            landingPageSteps.extend(stepLandingPage)
                     
                     #steps for reduce Rate Nol
                     percentageRemainingAllowance = int(amountRemainingAllowance)/int(amountBonusData)

                     halfPercentBonusData         = int(amountBonusData)*0.005
                     halfPercentBonusDataStr      = changeFormatData(halfPercentBonusData)

                     threePercentBonusData        = int(int(amountBonusData)*(float(percentageRemainingAllowance)-0.02))
                     threePercentBonusDataStr     = changeFormatData(threePercentBonusData)

                     nineteenPercentBonusData     = int(amountBonusData)*0.19
                     nineteenPercentBonusDataStr  = changeFormatData(nineteenPercentBonusData)

                     restPercentBonusData         = int(amountBonusData)*(1-(percentageRemainingAllowance+0.2))
                     restPercentBonusDataStr      = changeFormatData(restPercentBonusData)

                     stepsBonusData = [
                            ["Create event GPRS Intial RG 55 at D+"+str(day)+" 3PM", "Success", "No Bonus"],
                            ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 3PM", "Success", "No Bonus"],
                            ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 3PM", "Success", "No Bonus"],
                            ["Create event GPRS Terminate "+str(nineteenPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 3PM", "Charged 0 IDR", "No Bonus"],
                            ["Create Event GPRS "+str(restPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 5PM", "Charged 0 IDR", "No Bonus"],
                            ["Create event GPRS Intial RG 55 at D+"+str(day)+" 5PM", "Success", "No Bonus"],
                            ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 5PM", "Error code 4977", "No Bonus"],
                            ["Create event GPRS Terminate "+str(threePercentBonusDataStr)+" RG 55 at D+"+str(day)+" 5PM", "Charged 0 IDR", "No Bonus"],
                            ["Create event GPRS Intial RG 55 at D+"+str(day)+" 7PM", "Success", "No Bonus"],
                            ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 7PM", "Success", "No Bonus"],
                            ["Create event GPRS Terminate 0Kb RG 55 at D+"+str(day)+" 7PM", "Charged 0 IDR", "No Bonus"],
                            ["Create event GPRS Intial RG 55 at D+"+str(day)+" 9PM", "Success", "No Bonus"],
                            ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 9PM", "Success", "No Bonus"],
                            ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 9PM", "Reject with Error Code 4848", "No Bonus"],
                            ["Create event GPRS Terminate 0Kb RG 55 at D+"+str(day)+" 9PM", "Consume Bonus", "No Bonus"],
                            ["Create event GPRS Intial RG 55 at D+"+str(day)+" 11PM", "Reject with Error Code 4848", "No Bonus"],
                            ["Create event GPRS Terminate 0Kb RG 55 at D+"+str(day)+" 11PM", "Charged 0 IDR", "No Bonus"],
                     ]
                     
                     step = [
                            ["Create event GPRS Initial first notif RG 55 at D+"+str(day)+" 11AM", "Success", "No Bonus"],
                            ["Create event GPRS Intermediate "+str(rounded)+str(unit)+" RG 55 at D+"+str(day)+" 11AM", "Error code 4859", "No Bonus"],
                            ["Create event GPRS Terminate 0Kb RG 55 at D+"+str(day)+" 11AM", "Charged "+str(rate)+" IDR", "No Bonus"],
                     ]
                     step.extend(landingPageSteps)
                     step.extend(stepsBonusData)

                     checkedDaily.extend(step)
                     day += 1

              steps = [
                     ["Create & Activate new subscriber PP Prepaid "+name, "Check Active Period", "150 MB Data"],
                     ["Update Expiration Date", "ExpDate Updated", "150 MB Data"],
                     ["Update Balance", "Balance Updated", "150 MB Data"],
                     ["Check Bonus Preload", "Checked", "150 MB Data"],
                     ["Create event GPRS Intial RG 55 at 10AM", "Consume Bonus", "No Bonus"],
                     ["Create event GPRS Intermediate 153088kb RG 55 at 10AM", "Success", "No Bonus"],
                     ["Create event GPRS Intermediate 512kb RG 55 at 10AM", "Error code 4920", "No Bonus"],
                     ["Create event GPRS Terminate 0Kb RG 55 at 10AM", "Success", "No Bonus"],
              ]

              getCharge1MB         = 1024/float(rounded)*rate
              lastSteps = [
                     ["Create Event GPRS Event RG 17 apn telkomsel 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 50 apn internet 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 55 apn telkomsel 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 75 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 77 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 11 (International)  1024 kb", "Charged 1802 IDR", "No Bonus"], 
                     ["Check Indira","Success", "No Bonus"]
              ]

              steps.extend(checkedDaily)
              steps.extend(lastSteps)
       elif cardType == 'Simpati/Loop':
              checkedDaily  = []
              day           = 1

              while day <= 2:
                     landingPageSteps            = []
                     remainingAllowanceSteps     = []
                     landingPageSplit            = landingPage.split(",")
                     remainingAllowanceSplit     = remainingAllowance.split(',')
                     landingPageAmount           = rate

                     #Steps for landing page
                     for landingPages in landingPageSplit:
                            landingPagesSplit           = landingPages.split(';')
                            landingPageBorder           = landingPagesSplit[0]
                            landingPageErrorCode        = landingPagesSplit[1]
                            intermediateLandingPage     = math.ceil((int(landingPageBorder) - int(landingPageAmount)) * (float(rounded) / rate))
                            chargedLandingPage          = intermediateLandingPage/int(rounded)*int(rate)
                            landingPageAmount           += chargedLandingPage
                            
                            intermediateSteps           = []
                            while intermediateLandingPage > 0:
                                   if intermediateLandingPage > 512:
                                          page_count    = 512
                                          errorPage     = 'Success'
                                   else:
                                          page_count    = intermediateLandingPage
                                          errorPage     = "Error code "+str(landingPageErrorCode)

                                   intermediateStep = [
                                          ["Create event GPRS Intermediate "+str(page_count)+" "+str(unit)+" RG 55 at 11AM", errorPage, "No Bonus"],
                                   ]
                                   intermediateSteps.extend(intermediateStep)
                                   intermediateLandingPage -= page_count

                            stepLandingPage = [
                                   ["Create event GPRS Initial "+str(landingPageBorder)+" notif RG 55 at 11AM", "Success", "No Bonus"],
                            ]
                            stepLandingPageTerminate = [
                                   ["Create event GPRS Terminate 0Kb RG 55 at 11AM", "Charged "+str(chargedLandingPage)+" IDR", "No Bonus"],
                            ]
                            stepLandingPage.extend(intermediateSteps)
                            stepLandingPage.extend(stepLandingPageTerminate)

                            landingPageSteps.extend(stepLandingPage)
                     
                     #steps for reduce Rate Nol
                     for remainingAllowanceData in remainingAllowanceSplit:
                            remainingAllowanceSplitData = remainingAllowanceData.split(';')
                            amountRemainingAllowance    = remainingAllowanceSplitData[0]
                            unitRemainingAllowance      = remainingAllowanceSplitData[1]
                            errorCodeRemainingAllowance = remainingAllowanceSplitData[2]
                            
                            #change remaining allowance to kb
                            if unitRemainingAllowance.lower() == 'kb':
                                   amountRemainingAllowance      = amountRemainingAllowance   
                            elif unitRemainingAllowance.lower() == 'mb':
                                   amountRemainingAllowance      = float(amountRemainingAllowance)*1024
                            elif unitRemainingAllowance.lower() == 'gb':
                                   amountRemainingAllowance      = float(amountRemainingAllowance)*1048576

                            percentageRemainingAllowance = int(amountRemainingAllowance)/int(amountBonusData)
                            useRemainingAllowance        = amountBonusData-amountRemainingAllowance

                            halfPercentBonusData         = int(useRemainingAllowance)*0.005
                            halfPercentBonusDataStr      = changeFormatData(halfPercentBonusData)

                            threePercentBonusData        = int(useRemainingAllowance)*0.003
                            threePercentBonusDataStr     = changeFormatData(threePercentBonusData)

                            nineteenPercentBonusData     = int(useRemainingAllowance)*0.19
                            nineteenPercentBonusDataStr  = changeFormatData(nineteenPercentBonusData)

                            restPercentBonusData         = int(useRemainingAllowance)-(useRemainingAllowance*0.23)
                            restPercentBonusDataStr      = changeFormatData(restPercentBonusData)

                            stepsRemainingAllowanceData = [
                                   ["Create event GPRS Intial RG 55 at D+"+str(day)+" 3PM", "Success", "No Bonus"],
                                   ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 3PM", "Success", "No Bonus"],
                                   ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 3PM", "Success", "No Bonus"],
                                   ["Create event GPRS Terminate "+str(nineteenPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 3PM", "Charged 0 IDR", "No Bonus"],
                                   ["Create Event GPRS "+str(restPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 5PM", "Charged 0 IDR", "No Bonus"],
                                   ["Create event GPRS Intial RG 55 at D+"+str(day)+" 5PM", "Success", "No Bonus"],
                                   ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 5PM", "Error code "+errorCodeRemainingAllowance, "No Bonus"],
                                   ["Create event GPRS Terminate "+str(threePercentBonusDataStr)+" RG 55 at D+"+str(day)+" 5PM", "Charged 0 IDR", "No Bonus"],
                            ]
                            remainingAllowanceSteps.extend(stepsRemainingAllowanceData)
                            amountBonusData = amountRemainingAllowance
                     
                     step = [
                            ["Create event GPRS Initial first notif RG 55 at D+"+str(day)+" 11AM", "Success", "No Bonus"],
                            ["Create event GPRS Intermediate "+str(rounded)+str(unit)+" RG 55 at D+"+str(day)+" 11AM", "Error code 4859", "No Bonus"],
                            ["Create event GPRS Terminate 0Kb RG 55 at D+"+str(day)+" 11AM", "Charged "+str(rate)+" IDR", "No Bonus"],
                     ]
                     step.extend(landingPageSteps)
                     step.extend(remainingAllowanceSteps)

                     checkedDaily.extend(step)
                     day += 1

              steps = [
                     ["Create & Activate new subscriber PP "+name, "Checkl Active Period", "No Bonus"],
                     ["Update Exp Date", "Exp Date Updated", "No Bonus"],
                     ["Update Balance 1000000", "Balance Updated", "No Bonus"],
                     ["Create Event GPRS initial RG 55 3pm ", "Initial Success", "No Bonus"],
                     ["Create Event GPRS intermediate 512KB RG 55 3pm", "Initial Success", "No Bonus"],
                     ["Create Event GPRS intermediate 512KB RG 55 3pm", "Rejected 4920", "No Bonus"],
                     ["Create Event GPRS Terminate 0kb RG 55 3pm", "Charged 104 IDR", "No Bonus"],
                     #Landing Page
                     
              ]

              getCharge1MB         = 1024/float(rounded)*rate
              lastSteps = [
                     ["Create Event GPRS Event RG 17 apn telkomsel 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 50 apn internet 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 55 apn telkomsel 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 75 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 77 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 11 (International)  1024 kb", "Charged 1802 IDR", "No Bonus"], 
                     ["Check Indira","Success", "No Bonus"]
              ]

              steps.extend(checkedDaily)
              steps.extend(lastSteps)
       elif cardType == 'KartuAs':
              #3G First
              steps         = []
              checkedDaily  = []
              day           = 1

              while day <= 2:
                     landingPageSteps            = []
                     remainingAllowanceSteps     = []
                     landingPageSplit            = landingPage.split(",")
                     remainingAllowanceSplit     = remainingAllowance.split(',')
                     landingPageAmount           = rate

                     #Steps for landing page
                     for landingPages in landingPageSplit:
                            landingPagesSplit           = landingPages.split(';')
                            landingPageBorder           = landingPagesSplit[0]
                            landingPageErrorCode        = landingPagesSplit[1]
                            intermediateLandingPage     = math.ceil((int(landingPageBorder) - int(landingPageAmount)) * (float(rounded) / rate))
                            chargedLandingPage          = intermediateLandingPage/int(rounded)*int(rate)
                            landingPageAmount           += chargedLandingPage

                            intermediateSteps           = []
                            while intermediateLandingPage > 0:
                                   if intermediateLandingPage > 512:
                                          page_count    = 512
                                          errorPage     = 'Success'
                                   else:
                                          page_count    = intermediateLandingPage
                                          errorPage     = "Error code "+str(landingPageErrorCode)

                                   intermediateStep = [
                                          ["Create event GPRS Intermediate "+str(page_count)+" "+str(unit)+" RG 55 at 11AM", errorPage, "No Bonus"],
                                   ]
                                   intermediateSteps.extend(intermediateStep)
                                   intermediateLandingPage -= page_count

                            stepLandingPage = [
                                   ["Create event GPRS Initial "+str(landingPageBorder)+" notif RG 55 at 11AM", "Success", "No Bonus"],
                            ]
                            stepLandingPageTerminate = [
                                   ["Create event GPRS Terminate 0Kb RG 55 at 11AM", "Charged "+str(chargedLandingPage)+" IDR", "No Bonus"],
                            ]
                            stepLandingPage.extend(intermediateSteps)
                            stepLandingPage.extend(stepLandingPageTerminate)

                            landingPageSteps.extend(stepLandingPage)
                     
                     #steps for reduce Rate Nol
                     for remainingAllowanceData in remainingAllowanceSplit:
                            remainingAllowanceSplitData = remainingAllowanceData.split(';')
                            amountRemainingAllowance    = remainingAllowanceSplitData[0]
                            unitRemainingAllowance      = remainingAllowanceSplitData[1]
                            errorCodeRemainingAllowance = remainingAllowanceSplitData[2]
                            
                            #change remaining allowance to kb
                            if unitRemainingAllowance.lower() == 'kb':
                                   amountRemainingAllowance      = amountRemainingAllowance   
                            elif unitRemainingAllowance.lower() == 'mb':
                                   amountRemainingAllowance      = float(amountRemainingAllowance)*1024
                            elif unitRemainingAllowance.lower() == 'gb':
                                   amountRemainingAllowance      = float(amountRemainingAllowance)*1048576

                            percentageRemainingAllowance = int(amountRemainingAllowance)/int(amountBonusData)
                            useRemainingAllowance        = amountBonusData-amountRemainingAllowance

                            halfPercentBonusData         = int(useRemainingAllowance)*0.005
                            halfPercentBonusDataStr      = changeFormatData(halfPercentBonusData)

                            threePercentBonusData        = int(useRemainingAllowance)*0.003
                            threePercentBonusDataStr     = changeFormatData(threePercentBonusData)

                            nineteenPercentBonusData     = int(useRemainingAllowance)*0.19
                            nineteenPercentBonusDataStr  = changeFormatData(nineteenPercentBonusData)

                            restPercentBonusData         = int(useRemainingAllowance)-(useRemainingAllowance*0.23)
                            restPercentBonusDataStr      = changeFormatData(restPercentBonusData)

                            stepsRemainingAllowanceData = [
                                   ["Create event GPRS Intial RG 55 at D+"+str(day)+" 3PM", "Success", "No Bonus"],
                                   ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 3PM", "Success", "No Bonus"],
                                   ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 3PM", "Success", "No Bonus"],
                                   ["Create event GPRS Terminate "+str(nineteenPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 3PM", "Charged 0 IDR", "No Bonus"],
                                   ["Create Event GPRS "+str(restPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 5PM", "Charged 0 IDR", "No Bonus"],
                                   ["Create event GPRS Intial RG 55 at D+"+str(day)+" 5PM", "Success", "No Bonus"],
                                   ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day)+" 5PM", "Error code "+errorCodeRemainingAllowance, "No Bonus"],
                                   ["Create event GPRS Terminate "+str(threePercentBonusDataStr)+" RG 55 at D+"+str(day)+" 5PM", "Charged 0 IDR", "No Bonus"],
                            ]
                            remainingAllowanceSteps.extend(stepsRemainingAllowanceData)
                            amountBonusData = amountRemainingAllowance
                     
                     step = [
                            ["Create event GPRS Initial first notif RG 55 at D+"+str(day)+" 11AM", "Success", "No Bonus"],
                            ["Create event GPRS Intermediate "+str(rounded)+str(unit)+" RG 55 at D+"+str(day)+" 11AM", "Error code 4859", "No Bonus"],
                            ["Create event GPRS Terminate 0Kb RG 55 at D+"+str(day)+" 11AM", "Charged "+str(rate)+" IDR", "No Bonus"],
                     ]
                     step.extend(landingPageSteps)
                     step.extend(remainingAllowanceSteps)

                     checkedDaily.extend(step)
                     day += 1

              steps3G = [
                     ["Create & Activate new subscriber PP KartuAs Extra Ampuh Murah 3264714", "Check active period", "25MB Internet Pedana"],
                     ["Create event update balance 500000", "success", "No Bonus"],
                     ["Create Event GPRS initial RG 55 3pm ", "Initial Success", "No Bonus"],
                     ["Create Event GPRS intermediate 512KB RG 55 3pm", "Rejected 4949", "No Bonus"],
                     ["Create Event GPRS terminate 0kb RG 55 3pm", "Charged 3380 IDR", "No Bonus"],
              ]

              getCharge1MB         = 1024/float(rounded)*rate
              lastSteps = [
                     ["Create Event GPRS Event RG 17 apn telkomsel 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 50 apn internet 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 55 apn telkomsel 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 75 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 77 1024 kb", "Charged "+str(getCharge1MB)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 11 (International)  1024 kb", "Charged 1802 IDR", "No Bonus"], 
                     ["Check Indira","Success", "No Bonus"]
              ]

              steps.extend(steps3G)
              steps.extend(checkedDaily)
              steps.extend(lastSteps)
              #End 3G Step

              #4G Steps
              checkedDaily4G  = []
              day4G           = 1

              while day4G <= 2:
                     landingPageSteps            = []
                     remainingAllowanceSteps     = []
                     landingPageSplit            = landingPage.split(",")
                     remainingAllowanceSplit     = remainingAllowance.split(',')
                     landingPageAmount           = rate

                     #Steps for landing page
                     for landingPages in landingPageSplit:
                            landingPagesSplit           = landingPages.split(';')
                            landingPageBorder           = landingPagesSplit[0]
                            landingPageErrorCode        = landingPagesSplit[2]
                            intermediateLandingPage     = math.ceil((int(landingPageBorder) - int(landingPageAmount)) * (float(rounded) / rate))
                            chargedLandingPage          = intermediateLandingPage/int(rounded)*int(rate)
                            landingPageAmount           += chargedLandingPage

                            intermediateSteps           = []
                            while intermediateLandingPage > 0:
                                   if intermediateLandingPage > 512:
                                          page_count    = 512
                                          errorPage     = 'Success'
                                   else:
                                          page_count    = intermediateLandingPage
                                          errorPage     = "Error code "+str(landingPageErrorCode)

                                   intermediateStep = [
                                          ["Create event GPRS Intermediate "+str(page_count)+" "+str(unit)+" RG 55 at 11AM", errorPage, "No Bonus"],
                                   ]
                                   intermediateSteps.extend(intermediateStep)
                                   intermediateLandingPage -= page_count

                            stepLandingPage = [
                                   ["Create event GPRS Initial "+str(landingPageBorder)+" notif RG 55 at 11AM", "Success", "No Bonus"],
                            ]
                            stepLandingPageTerminate = [
                                   ["Create event GPRS Terminate 0Kb RG 55 at 11AM", "Charged "+str(chargedLandingPage)+" IDR", "No Bonus"],
                            ]
                            stepLandingPage.extend(intermediateSteps)
                            stepLandingPage.extend(stepLandingPageTerminate)

                            landingPageSteps.extend(stepLandingPage)
                     
                     #steps for reduce Rate Nol
                     for remainingAllowanceData in remainingAllowanceSplit:
                            remainingAllowanceSplitData = remainingAllowanceData.split(';')
                            amountRemainingAllowance    = remainingAllowanceSplitData[0]
                            unitRemainingAllowance      = remainingAllowanceSplitData[1]
                            errorCodeRemainingAllowance = remainingAllowanceSplitData[2]
                            
                            #change remaining allowance to kb
                            if unitRemainingAllowance.lower() == 'kb':
                                   amountRemainingAllowance      = amountRemainingAllowance   
                            elif unitRemainingAllowance.lower() == 'mb':
                                   amountRemainingAllowance      = float(amountRemainingAllowance)*1024
                            elif unitRemainingAllowance.lower() == 'gb':
                                   amountRemainingAllowance      = float(amountRemainingAllowance)*1048576

                            percentageRemainingAllowance = int(amountRemainingAllowance)/int(amountBonusData)
                            useRemainingAllowance        = amountBonusData-amountRemainingAllowance

                            halfPercentBonusData         = int(useRemainingAllowance)*0.005
                            halfPercentBonusDataStr      = changeFormatData(halfPercentBonusData)

                            threePercentBonusData        = int(useRemainingAllowance)*0.003
                            threePercentBonusDataStr     = changeFormatData(threePercentBonusData)

                            nineteenPercentBonusData     = int(useRemainingAllowance)*0.19
                            nineteenPercentBonusDataStr  = changeFormatData(nineteenPercentBonusData)

                            restPercentBonusData         = int(useRemainingAllowance)-(useRemainingAllowance*0.23)
                            restPercentBonusDataStr      = changeFormatData(restPercentBonusData)

                            stepsRemainingAllowanceData = [
                                   ["Create event GPRS Intial RG 55 at D+"+str(day4G)+" 3PM", "Success", "No Bonus"],
                                   ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day4G)+" 3PM", "Success", "No Bonus"],
                                   ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day4G)+" 3PM", "Success", "No Bonus"],
                                   ["Create event GPRS Terminate "+str(nineteenPercentBonusDataStr)+" RG 55 at D+"+str(day4G)+" 3PM", "Charged 0 IDR", "No Bonus"],
                                   ["Create Event GPRS "+str(restPercentBonusDataStr)+" RG 55 at D+"+str(day4G)+" 5PM", "Charged 0 IDR", "No Bonus"],
                                   ["Create event GPRS Intial RG 55 at D+"+str(day4G)+" 5PM", "Success", "No Bonus"],
                                   ["Create event GPRS Intermediate "+str(halfPercentBonusDataStr)+" RG 55 at D+"+str(day4G)+" 5PM", "Error code "+errorCodeRemainingAllowance, "No Bonus"],
                                   ["Create event GPRS Terminate "+str(threePercentBonusDataStr)+" RG 55 at D+"+str(day4G)+" 5PM", "Charged 0 IDR", "No Bonus"],
                            ]
                            remainingAllowanceSteps.extend(stepsRemainingAllowanceData)
                            amountBonusData = amountRemainingAllowance
                     
                     step = [
                            ["Create event GPRS Initial first notif RG 55 at D+"+str(day4G)+" 11AM", "Success", "No Bonus"],
                            ["Create event GPRS Intermediate "+str(rounded)+str(unit)+" RG 55 at D+"+str(day4G)+" 11AM", "Error code 4859", "No Bonus"],
                            ["Create event GPRS Terminate 0Kb RG 55 at D+"+str(day4G)+" 11AM", "Charged "+str(rate)+" IDR", "No Bonus"],
                     ]
                     step.extend(landingPageSteps)
                     step.extend(remainingAllowanceSteps)

                     checkedDaily4G.extend(step)
                     day4G += 1

              steps4G = [
                     ["Create & Activate new subscriber PP KartuAS Gampang Internetan", "", "No Bonus"],
                     ["Attach Offer Landing Page Kartu As 4G - 3361244", "", "No Bonus"],
                     ["Create Event GPRS initial RG 17 3pm ", "Initial Success", "No Bonus"],
                     ["Create Event GPRS intermediate 25088KB RG 17 3pm", "Intermediate Success", "No Bonus"],
                     ["Create Event GPRS intermediate 512KB RG 17 3pm", "Rejected 4960", "No Bonus"],
                     ["Create Event GPRS terminate 0kb RG 17 3pm", "Consume Bonus", "No Bonus"],
                     ["Create event update balance 500000", "success", "No Bonus"],
                     ["Create Event GPRS initial RG 17 3pm ", "Initial Success", "No Bonus"],
                     ["Create Event GPRS intermediate 512KB RG 17 3pm", "Rejected 4949", "No Bonus"],
                     ["Create Event GPRS terminate 0kb RG 17 3pm", "Charged 3380 IDR", "No Bonus"],
              ]

              getCharge1MB4G         = 1024/float(rounded)*rate
              lastSteps4G = [
                     ["Create Event GPRS Event RG 17 apn telkomsel 1024 kb", "Charged "+str(getCharge1MB4G)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 50 apn internet 1024 kb", "Charged "+str(getCharge1MB4G)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 55 apn telkomsel 1024 kb", "Charged "+str(getCharge1MB4G)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 75 1024 kb", "Charged "+str(getCharge1MB4G)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 77 1024 kb", "Charged "+str(getCharge1MB4G)+" IDR", "No Bonus"], 
                     ["Create Event GPRS Event RG 11 (International)  1024 kb", "Charged 1802 IDR", "No Bonus"], 
                     ["Check Indira","Success", "No Bonus"]
              ]

              steps.extend(steps4G)
              steps.extend(checkedDaily4G)
              steps.extend(lastSteps4G)
       
       return steps

def GPRSLandingPagePostpaid(name, unit, rate):
       staticRate           = 1024 * rate
       halfStaticRate       = 512 * rate
       marginBottom         = 50000
       restKuota            = math.ceil((marginBottom - halfStaticRate) / rate)
       allPrice             = (restKuota+512) * rate
       step = [
              [f"Create & Activate new subscriber PP {name}","Success"],
              ["Update Parameter","Success"],
              ["Attach Offer New CLS","Success"],
              ["Attach offer international Roaming","Success"],
              [f"Create Event GPRS Event RG 17 apn telkomsel 1024 {unit}",f"Charged {staticRate} IDR"],
              [f"Create Event GPRS Event RG 50 apn internet 1024 {unit}",f"Charged {staticRate} IDR"],
              [f"Create Event GPRS Event RG 55 apn telkomsel 1024 {unit}",f"Charged {staticRate} IDR"],
              [f"Create Event GPRS Event RG 75 1024 {unit}",f"Charged {staticRate} IDR"],
              [f"Create Event GPRS Event RG 77 1024 {unit}",f"Charged {staticRate} IDR"],
              ["Create Event GPRS Event RG 11 (International) D+1 11am","Charged 1802 IDR"],
              ["Create Event GPRS Initial RG 50 D+1 11am","Success"],
              [f"Create Event GPRS Intermediate {restKuota} {unit} RG 50 D+1 11am","Success"],
              [f"Create Event GPRS Terminate 512 {unit} D+1 11am",f"Charged {allPrice} IDR"],
              ["Create Event GPRS Initial RG 50 D+1 11am","Error Code 4831"],
              [f"Create Event GPRS Terminate 0 {unit} D+1 11am","Charged 0"],
              ["Create Event GPRS Initial RG 50 D+1 11am","No Error Code"],
              [f"Create Event GPRS Terminate 0 {unit} D+1 11am","Charged 0"],
              ["Create Event GPRS Initial RG 50 D+1 11am","Success"],
              [f"Create Event GPRS Intermediate {restKuota} {unit} RG 50 D+1 11am","Success"],
              [f"Create Event GPRS Terminate 512 {unit} D+1 11am",f"Charged {allPrice} IDR"],
              ["Create Event GPRS Initial RG 50 D+1 11am","Error Code 4832"],
              [f"Create Event GPRS Terminate 0 {unit} D+1 11am","Charged 0"],
              ["Create Event GPRS Initial RG 50 D+1 11am","No Error Code"],
              [f"Create Event GPRS Terminate 0 {unit} D+1 11am","Charged 0"],
              ["Create Event GPRS Initial RG 50 D+1 11am","Success"],
              [f"Create Event GPRS Intermediate {restKuota} {unit} RG 50 D+1 11am", "Success"],
              [f"Create Event GPRS Terminate 512 {unit} D+1 11am",f"Charged {allPrice} IDR"],
              ["Create Event GPRS Initial RG 50 D+1 11am","Error Code 4833"],
              [f"Create Event GPRS Terminate 0 {unit} D+1 11am","Charged 0"],
              ["Create Event GPRS Initial RG 50 D+1 11am","No Error Code"],
              [f"Create Event GPRS Terminate 0 {unit} D+1 11am","Charged 0"],
              ["Create Event GPRS Initial RG 50 D+1 11am","Success"],
              [f"Create Event GPRS Intermediate {restKuota} {unit} RG 50 D+1 11am","Success"],
              [f"Create Event GPRS Terminate 512 {unit} D+1 11am",f"Charged {allPrice} IDR"],
              ["Create Event GPRS Initial RG 50 D+1 11am","Error Code 4834"],
              [f"Create Event GPRS Terminate 0 {unit} D+1 11am","Charged 0"],
              ["Create Event GPRS Initial RG 50 D+1 11am","No Error Code"],
              [f"Create Event GPRS Terminate 0 {unit} D+1 11am","Charged 0"],
              ["Create Event GPRS Initial RG 50 D+2 11am","Success"],
              [f"Create Event GPRS Intermediate {restKuota} {unit} RG 50 D+2 11am","Success"],
              [f"Create Event GPRS Terminate 512 {unit} D+2 11am",f"Charged {allPrice} IDR"],
              ["Create Event GPRS initial D+2 11am","Success"],
              ["Create Event GPRS intermediate 7000 D+2 11am","Success"],
              ["Create Event GPRS Terminate 512 D+2 11am",f"Charged {halfStaticRate} IDR"],
              ["Create Event GPRS Initial RG 50 D+2 11am","No Error Code"],
              [f"Create Event GPRS Terminate 0 {unit} D+2 11am","Charged 0"],
              ["Check Indira","Checked"],
              ["Check error in table trb_subs_errs","Checked"] 
       ]

       return step

def attachOfferPrepaid(name):
       now = datetime.now()
       current_year = now.year
       current_month = now.month
       current_day = now.day
       current_date = now.date()
       next_date = current_date + timedelta(days=10)
       next_date = next_date.strftime("%Y%m%d")
       step = [
              [f"Create & Activate new subscriber PP {name}", "Check active period", "150MB Internet Perdana"],
              ["Update Balance", "Balance Updated", "150MB Internet Perdana"],
              ["Update Exp Date", "Expired Date Updated", "150MB Internet Perdana"],
              ["Check bonus 889*2", "Success", "No Bonus"],
              ["Check bonus 889*1", "Success", "150MB Internet Perdana"],
              ["Attach offer BI-WEEKLY Voice (901161)", "Offer Attached", "Tsel 1000 Minutes, AllOpr 100 Minutes; 150MB Internet Perdana"],
              ["Create event voice onnet 1800s, 11AM", "Consume Bonus", "Tsel 970 Minutes, AllOpr 100 Minutes; 150MB Internet Perdana"],
              ["Create event voice offnet 600s, 1PM", "Consume Bonus", "Tsel 970 Minutes, AllOpr 90 Minutes; 150MB Internet Perdana"],
              ["Create event voice PSTN 600s, 5PM", "Consume Bonus", "Tsel 970 Minutes, AllOpr 80 Minutes; 150MB Internet Perdana"],
              ["Create event voice FWA 600s, 8PM", "Consume Bonus", "Tsel 970 Minutes, AllOpr 70 Minutes; 150MB Internet Perdana"],
              ["Create event 1 SMS onnet, 11PM", "Charged", "Tsel 970 Minutes, AllOpr 70 Minutes; 150MB Internet Perdana"],
              ["Create event voice onnet 6000s, 1AM D+1", "Consume Bonus", "Tsel 870 Minutes, AllOpr 70 Minutes; 150MB Internet Perdana"],
              ["Create event voice onnet 6000s, 5AM D+1", "Consume Bonus", "Tsel 770 Minutes, AllOpr 70 Minutes; 150MB Internet Perdana"],
              ["Create event GPRS 150 MB RG17, 11AM D+1", "Consume Bonus", "Tsel 770 Minutes, AllOpr 70 Minutes"],
              ["Attach offer BI-WEEKLY Voice (901161)", "Offer Attached", "Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 770 Minutes, AllOpr 70 Minutes"],
              ["Attach offer BI-WEEKLY Voice (901161)", "Offer Attached", "Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 770 Minutes, AllOpr 70 Minutes"],
              ["Attach offer BI-WEEKLY Voice (901161)", "Offer Attached", "Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 770 Minutes, AllOpr 70 Minutes"],
              ["Attach offer BI-WEEKLY Voice (901161)", "Offer Attached", "Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 770 Minutes, AllOpr 70 Minutes"],
              ["Attach offer BI-WEEKLY Voice (901161)", "Offer Attached", "Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 1000 Minutes, AllOpr 100 Minutes; Tsel 770 Minutes, AllOpr 70 Minutes"],
              [f"Create & Activate new subscriber PP {name}", "Check active period", "150MB Internet Perdana"],
              ["Update Balance", "Success", "150MB Internet Perdana"],
              ["Attach Talkmania | 99L15", "Diameter Attached", "Tsel 200 minutes, Tsel 20 sms;150MB Internet Perdana"],
              ["Create event voice onnet 3000s, 11AM", "Consume Bonus", "Tsel 150 minutes, Tsel 20 sms;150MB Internet Perdana"],
              ["Create event voice offnet 60s, 1PM", "Charged", "Tsel 150 minutes, Tsel 20 sms;150MB Internet Perdana"],
              ["Create event 20 SMS onnet, 5PM", "Consume Bonus", "Tsel 150 minutes;150MB Internet Perdana"],
              ["Create event voice onnet 3000s, 8PM", "Consume Bonus", "Tsel 100 minutes;150MB Internet Perdana"],
              ["Create event voice onnet 6000s, 11PM", "Consume Bonus", "150MB Internet Perdana"],
              ["Create event GPRS 50 MB RG17, 1AM D+1", "Consume Bonus", "100MB Internet Perdana"],
              [f"Create & Activate new subscriber PP {name}", "Check active period", "150MB Internet Perdana"],
              ["Update Balance", "Success", "150MB Internet Perdana"],
              ["Create event voice onnet 600s ", "Charged", "150MB Internet Perdana"],
              ["Cek Bonus Info", "Success", "150MB Internet Perdana"],
              ["Create event GPRS 50 MB RG17", "Consume Bonus", "100MB Internet Perdana"],
              ["Cek Bonus Info", "Success", "100MB Internet Perdana"],
              [f"Attach offer BI-WEEKLY Voice (99901161) with AILC 3701034$V$0.3$6000${current_year}{current_month:02d}{current_day:02d}235959${next_date}235959$", "Offer Attached", "100 Min Tsel;100MB Internet Perdana"],
              [f"Attach offer Talkmania | 99A52 with AILC : 3713016$V$0.3$3000${current_year}{current_month:02d}{current_day:02d}235959${next_date}235959$;", "Diameter Attached", "50 Min Tsel; 100 Min Tsel;100MB Internet Perdana"],
              ["Create event voice onnet 2400s, 11AM", "Consume Bonus", "50 Min Tsel; 60 Min Tsel;100MB Internet Perdana"],
              ["Create event voice offnet 60s, 1PM", "Charged", "50 Min Tsel; 60 Min Tsel;100MB Internet Perdana"],
              ["Check bonus 889 D+0 11PM", "Success", "50 Min Tsel; 60 Min Tsel;100MB Internet Perdana"],
              ["Update Accumulator from Method Invoker 6000;Override;V;3700784;3701034|0", "Success", "50 Min Tsel; 100 Min Tsel;100MB Internet Perdana"],
       ]

       return step

def preloadBonusPrepaid(name, welcomeMessage, preloadBonus, preloadBalance, validity):

       preloadBonusSplit           = preloadBonus.split(";")
       preloadBonusData            = preloadBonusSplit[0]
       preloadBonusDataString      = preloadBonusData+" Internet Perdana"
       RGUsed                      = ["17","55","75","77"]
       day                         = 1
       preloadBonusDataMin         = int(preloadBonusData) - 30
       countPreloadBonusData       = preloadBonusDataMin
       checkedStep = []
       while day < validity:
              random_numbers = generate_ordered_multiple_random_numbers(base=15, count=5)
              random_value = random.choice(RGUsed)
              random_number = random.choice(random_numbers)
              restPreloadBonusData = countPreloadBonusData-random_number
              if restPreloadBonusData > 0:
                     steps = [
                            [f"Create event GPRS {random_number} MB with RG {random_value} at 5PM D+{day}","Consume Bonus",f"{restPreloadBonusData}MB Internet Perdana"],
                     ]
                     countPreloadBonusData = countPreloadBonusData-random_number
                     checkedStep.extend(steps)
              else:
                     steps = [
                            [f"Create event GPRS {countPreloadBonusData} MB with RG {random_value} at 5PM D+{day}","Consume Bonus","0MB Internet Perdana"],
                     ]
                     checkedStep.extend(steps)
                     break
              day += 1     
       
       checkedStep.extend([[f"Create event GPRS 1 MB with RG 50 at 11PM D+{validity}","Charged","No Bonus"]])                
       
       steps = [
              [f"Create & Activate new subscriber PP {name}","Check active period",preloadBonusDataString],
              ["Check Welcome Message",welcomeMessage,preloadBonusDataString],
              ["Check Expired date PP","Checked",preloadBonusDataString],
              ["Check L9 Get Bonus Info and Validity","Checked",preloadBonusDataString],
              ["Check PRIT_NAME","BNS_BonusNameFollow",preloadBonusDataString],
              ["Check Initial Balance",preloadBalance,preloadBonusDataString],
              ["Check Bonus 889","Checked",preloadBonusDataString],
              ["Check 889*1","Checked",preloadBonusDataString],
              ["Check 889*2","Checked","No Bonus"],
              ["Check 889*3","Checked","No Bonus"],
              ["Check 889*4","Checked","No Bonus"],
              ["Create event voice onnet 10s","Charged",preloadBonusDataString],
              ["Create event 1 SMS offnet","Charged",preloadBonusDataString],
              ["Create event GPRS 30 MB with RG 50 at 11AM","Consume Bonus",f"{preloadBonusDataMin}MB Internet Perdana"],
              ["Check Balance 888","Balance exist after charge",f"{preloadBonusDataMin}MB Internet Perdana"],
              ["Update Balance 50000","Balance Update",f"{preloadBonusDataMin}MB Internet Perdana"],
              ["Create event 1 MMS","Charged",f"{preloadBonusDataMin}MB Internet Perdana"],
              ["Create event direct debit with vascode cm_digi_1500 at 1AM D+1","Charged",f"{preloadBonusDataMin}MB Internet Perdana"],
              ["Create event voice international to Malaysia (60) at 9PM D+1","Charged",f"{preloadBonusDataMin}MB Internet Perdana"],
              ["Create event sms international to Malaysia (60) at 11PM D+1","Charged",f"{preloadBonusDataMin}MB Internet Perdana"],
              ["Create event transfer balance 5000 from A number with transferparty SO","Success",f"{preloadBonusDataMin}MB Internet Perdana"],
              ["Create event transfer balance 500000 to A number with transferparty TA","Success",f"{preloadBonusDataMin}MB Internet Perdana"],
              ["Check Balance 888","should be added 500000 IDR",f"{preloadBonusDataMin}MB Internet Perdana"],
              ["Create event GPRS Roaming 1 MB using tapcode BGDBL (MCCMNC : 47003) with APN telkomsel.r RG11, 11am D+0","Charged",f"{preloadBonusDataMin}MB Internet Perdana"],
       ]

       steps.extend(checkedStep)
       return steps

def generate_ordered_multiple_random_numbers(base, count):
    random_numbers = []
    for _ in range(count):
        random_number = random.randint(1, 100)  # Adjust the range as needed
        multiple = (random_number // base) * base
        random_numbers.append(multiple)
    random_numbers.sort()  # Sort the list in ascending order
    return random_numbers

def GPRSRoamPrepaid(name):
       # Get the current date
       current_date = datetime.now()
       next_day = current_date + timedelta(days=1)
       next_day_plus_2 = current_date + timedelta(days=2)
       next_day_plus_3 = current_date + timedelta(days=3)

       # Format the date as "day Month Year"
       formatted_date = current_date.strftime("%d %B %Y")
       formatted_next_day = next_day.strftime("%d %B %Y")
       formatted_next_day_plus_2 = next_day_plus_2.strftime("%d %B %Y")
       formatted_next_day_plus_3 = next_day_plus_3.strftime("%d %B %Y")
       steps = [
              [f"Create & Activate new subscriber PP {name}", "Check active period", "No Bonus"],
              ["Update Balance 10M", "Balance Updated", "No Bonus"],
              ["Update exp date", "Success", "No Bonus"],
              [f"Create Event GPRS Roaming 10kb in AUSHU (MCCMNC : 50506) using APN internet.r RG 11 and Group International Roaming - Capped Data Roaming Operator ID-150k on {formatted_date} 00.00.00", "Charged 2000 IDR", "No Bonus"],
              ["Check Rounded GPRS  Above Events", "Should be 1024kb / 1048576 bytes", "No Bonus"],
              [f"Create Event GPRS Roaming 100kb in BRATM (MCCMNC : 72431) using APN telkomsel.r RG 11 and Group International Roaming - Capped Data Roaming Operator ID-200k on {formatted_date} 00.00.00", "Charged 2000 IDR", "No Bonus"],
              ["Check Rounded GPRS  Above Events", "Should be 1024kb / 1048576 bytes", "No Bonus"],
              [f"Create Event GPRS Roaming 1536kb in ALBVF (MCCMNC : 27602) using APN internet.r RG 11 and Group International Roaming - Capped Data Roaming Operator ID-250k on {formatted_date} 00.00.00", "Charged 4000 IDR", "No Bonus"],
              ["Check Rounded GPRS  Above Events", "Should be 2048kb / 2097152 bytes", "No Bonus"],
              [f"Create Event GPRS Roaming 1024kb in SAUET (MCCMNC : 42003) using APN telkomsel.r RG 11 and Group SAU Operator - Capped GPRS Roaming on {formatted_date} 00.00.00", "Charged 2000 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 20480kb in AZEAF (MCCMNC : 40004) using APN internet.r RG 11 and Group non listed on BRD on {formatted_date} 00.00.00", "Charged not 40000 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 102400kb in CHNCT (MCCMNC : 46000) using APN telkomsel.r RG 11 and Group International Roaming - Capped Data Roaming Operator ID-150k on {formatted_date} 00.00.00", "Charged 190000 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 1024kb in IND23 (MCCMNC : 405801) using APN internet.r RG 11 and Group International Roaming - Capped Data Roaming Operator ID-150k on {formatted_date} 00.00.00", "Charged 0 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 102400kb in CHLMV (MCCMNC : 73001) using APN telkomsel.r RG 11 and Group International Roaming - Capped Data Roaming Operator ID-200k on {formatted_date} 00.00.00", "Charged 0 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 1048576kb in BELKO (MCCMNC : 20620) using APN internet.r RG 11 and Group International Roaming - Capped Data Roaming Operator ID-250k on {formatted_date} 00.00.00", "Charged 0 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 1048576kb in SAUAJ (MCCMNC : 42001) using APN telkomsel.r RG 11 and Group SAU Operator - Capped GPRS Roaming on {formatted_date} 00.00.00", "Charged 0 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 10kb in GUMDP (MCCMNC : 310370) using APN internet.r RG 11 and Group non listed on BRD on {formatted_date} 00.00.00", "Charged not 2000 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 10kb in JPNEM (MCCMNC : 44000) using APN internet.r RG 11 and Group International Roaming - Capped Data Roaming Operator ID-150k on on {formatted_next_day} 00.00.00", "Charged 2000 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 100kb in ARGTP (MCCMNC : 72234) using APN telkomsel.r RG 11 and Group International Roaming - Capped Data Roaming Operator ID-200k on on {formatted_next_day} 00.00.00", "Charged 2000 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 1536kb in CZEET (MCCMNC : 23002) using APN internet.r RG 11 and Group International Roaming - Capped Data Roaming Operator ID-250k on on {formatted_next_day} 00.00.00", "Charged 4000 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 1024kb in SAUZN (MCCMNC : 42004) using APN telkomsel.r RG 11 and Group SAU Operator - Capped GPRS Roaming on on {formatted_next_day} 00.00.00", "Charged 2000 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 20480kb in FJIVF (MCCMNC : 54201) using APN internet.r RG 11 and Group non listed on BRD on on {formatted_next_day} 00.00.00", "Charged not 40000 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 97280kb in KORKT (MCCMNC : 45002) using APN telkomsel.r RG 11 and Group International Roaming - Capped Data Roaming Operator ID-150k on on {formatted_next_day} 00.00.00", "Charged 190000 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 5242880kb in SAUET (MCCMNC : 42003) using APN internet.r RG 11 and Group SAU Operator - Capped GPRS Roaming on on {formatted_next_day} 00.00.00", "Charged 0 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 51200kb in CHNCT (MCCMNC : 46000) using APN telkomsel.r RG 11 and Group International Roaming - Capped Data Roaming Operator ID-150k on on {formatted_next_day_plus_2} 00.00.00", "Charged 100000 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 51200kb in ARGTP (MCCMNC : 72234) using APN internet.r RG 11 and Group International Roaming - Capped Data Roaming Operator ID-200k on on {formatted_next_day_plus_3} 00.00.00", "Charged 100000 IDR", "No Bonus"],
              [f"Create Event GPRS Roaming 1536kb in CZEET (MCCMNC : 23002) using APN telkomsel.r RG 11 and Group International Roaming - Capped Data Roaming Operator ID-250k on on {formatted_next_day_plus_3} 00.00.00", "Charged 4000 IDR", "No Bonus"],
              ["Check rated vent", "Success", "No Bonus"],
              ["Check RB Log", "Success", "No Bonus"],
              ["Check INDIRA", "Success", "No Bonus"],
              ["Set New Credit Limit Service (offer id : 3669334) as 4.000.000", "Offer Attached", "No Bonus"],
              ["Attach Offer Flexible Voice Offnet Darurat with Parameter : Quota|UOM|Validity end date|RC indicator|Invoice description|Quotation reference|External product id|TransactionID Value : 3600|V|2022-02-25 22:00:00|0|String1|String2|String3|String4", "Offer Attached", "60 Min Opr lain"],
              ["Create event Voice offnet 1s 1pm D+1", "Consume Bonus", "59 Min Opr lain"],
              ["Create event Voice offnet 60s 2pm D+2", "Consume Bonus", "58 Min Opr lain"],
              ["Create event voice offnet 120s 5pm D+2", "Consume Bonus", "56 Min Opr lain"],
              ["Create event voice onnet 10s 11pm D+4", "Charged", "56 Min Opr lain"],
              ["Create event 1 SMS onnet 2pm D+5", "Charged", "56 Min Opr lain"],
              ["Create event voice PSTN 10s 6pm D+5", "Charged", "56 Min Opr lain"],
              ["Create event voice FWA 10s 8pm D+5", "Charged", "56 Min Opr lain"],
              [f"Create event voice onnet 180s 5pm at {formatted_date}", "Consume Bonus", "53 Min Opr lain"],
       ]

       return steps

def splitRechargePrepaid(name, welcomeMessage):
       kartuAs = name.lower().find('kartu as')
       if kartuAs == -1:
              steps = [
                     [f"Create & Activate new subscriber {name}", "Check active period", "No Bonus"],
                     ["Check Welcome Message in XML", welcomeMessage, "No Bonus"],
                     ["Update Exp Date ", "Exp Date Updated", "No Bonus"],
                     ["Check Bonus Info", "No Bonus", "No Bonus"],
                     ["Check Bonus 889 11am", "No Bonus", "No Bonus"],
                     ["Check Balance 888 11am", "Preload Balance 5000 IDR", "No Bonus"],
                     ["Do Recharge Rp 10000 using Split Code 002", "Recharge Success", "No Bonus"],
                     ["Check Balance 888 11am", "Balance 10000 IDR", "No Bonus"],
                     ["Check Bonus Info", "No Bonus", "No Bonus"],
                     ["Do Recharge Rp 10000 using Split Code 075", "Recharge Success", "No Bonus"],
                     ["Check Balance 888 11am", "Balance 10001 IDR", "No Bonus"],
                     ["Check Bonus Info", "No Bonus", "No Bonus"],
                     ["Do Recharge Rp 50000 using Split Code 006", "Recharge Success", "30 min Tsel"],
                     ["Check Balance 888 11am", "Balance 60001 IDR", "30 min Tsel"],
                     ["Check Bonus Info", "30 min Tsel", "30 min Tsel"],
                     ["Do Recharge Rp 10000 using Split Code 014", "Recharge Success", "30 min Tsel, 30MB 2G/3G"],
                     ["Check Balance 888 11am", "Balance 60002 IDR", "30 min Tsel, 30MB 2G/3G"],
                     ["Check Bonus Info", "30 min Tsel, 30MB 2G/3G", "30 min Tsel, 30MB 2G/3G"],
                     ["Do Recharge Rp 20000 using Split Code 068", "Recharge Success", "30 min Tsel, 30MB 2G/3G, 4000 IDR Monetary"],
                     ["Check Balance 888 11am", "Balance 80002 IDR", "30 min Tsel, 30MB 2G/3G, 4000 IDR Monetary"],
                     ["Check Bonus Info", "30 min Tsel, 30MB 2G/3G, 4000 IDR Monetary", "30 min Tsel, 30MB 2G/3G, 4000 IDR Monetary"],
                     ["Do Recharge Rp 1000 using Split Code 022", "Recharge Success", "30 min Tsel, 30MB 2G/3G, 4000 IDR Monetary, 1000 SMS Tsel "],
                     ["Check Balance 888 11am", "Balance 80003 IDR", "30 min Tsel, 30MB 2G/3G, 4000 IDR Monetary, 1000 SMS Tsel "],
                     ["Check Bonus Info", "30 min Tsel, 30MB 2G/3G, 4000 IDR Monetary, 1000 SMS Tsel", "30 min Tsel, 30MB 2G/3G, 4000 IDR Monetary, 1000 SMS Tsel "]
              ]
       else:
              steps = [
                     [f"Create & Activate new subscriber {name}","Check active period","No Bonus"],
                     ["Check Welcome Message in XML",welcomeMessage,"No Bonus"],
                     ["Update Exp Date ","Exp Date Updated","No Bonus"],
                     ["Check Bonus Info","No Bonus","No Bonus"],
                     ["Check Bonus 889 11am","No Bonus","No Bonus"],
                     ["Check Balance 888 11am","Preload Balance 5000 IDR","No Bonus"],
                     ["Do Recharge Rp 10000 using Split Code 002","Recharge Success","Unlimited Voice Tsel, Unlimited SMS Tsel"],
                     ["Check Balance 888 11am","Balance 150 IDR","Unlimited Voice Tsel, Unlimited SMS Tsel"],
                     ["Check Bonus Info","Checked","Unlimited Voice Tsel, Unlimited SMS Tsel"],
                     ["Do Recharge Rp 10000 using Split Code 075","Recharge Success","Unlimited Voice Tsel, Unlimited SMS Tsel"],
                     ["Check Balance 888 11am","Balance 151 IDR","Unlimited Voice Tsel, Unlimited SMS Tsel"],
                     ["Check Bonus Info","Checked","Unlimited Voice Tsel, Unlimited SMS Tsel"],
                     ["Do Recharge Rp 50000 using Split Code 006","Recharge Success","Unlimited Voice Tsel, Unlimited SMS Tsel, 30 min Tsel"],
                     ["Check Balance 888 11am","Balance 50151 IDR","Unlimited Voice Tsel, Unlimited SMS Tsel, 30 min Tsel"],
                     ["Check Bonus Info","Checked","Unlimited Voice Tsel, Unlimited SMS Tsel, 30 min Tsel"],
                     ["Do Recharge Rp 10000 using Split Code 014","Recharge Success","Unlimited Voice Tsel, Unlimited SMS Tsel, 30 min Tsel, 30MB 2G/3G"],
                     ["Check Balance 888 11am","Balance 50152 IDR","Unlimited Voice Tsel, Unlimited SMS Tsel, 30 min Tsel, 30MB 2G/3G"],
                     ["Check Bonus Info","Checked","Unlimited Voice Tsel, Unlimited SMS Tsel, 30 min Tsel, 30MB 2G/3G"],
                     ["Do Recharge Rp 20000 using Split Code 068","Recharge Success","Unlimited Voice Tsel, Unlimited SMS Tsel, 30 min Tsel, 30MB 2G/3G, 4000 IDR Monetary"],
                     ["Check Balance 888 11am","Balance 70152 IDR","Unlimited Voice Tsel, Unlimited SMS Tsel, 30 min Tsel, 30MB 2G/3G, 4000 IDR Monetary"],
                     ["Check Bonus Info","Checked","Unlimited Voice Tsel, Unlimited SMS Tsel, 30 min Tsel, 30MB 2G/3G, 4000 IDR Monetary"],
                     ["Do Recharge Rp 1000 using Split Code 019","Recharge Success","Unlimited Voice Tsel, Unlimited SMS Tsel, 30 min Tsel, 30MB 2G/3G, 4000 IDR Monetary, 1000 SMS Tsel"],
                     ["Check Balance 888 11am","Balance 70153 IDR","Unlimited Voice Tsel, Unlimited SMS Tsel, 30 min Tsel, 30MB 2G/3G, 4000 IDR Monetary, 1000 SMS Tsel"],
                     ["Check Bonus Info","Checked","Unlimited Voice Tsel, Unlimited SMS Tsel, 30 min Tsel, 30MB 2G/3G, 4000 IDR Monetary, 1000 SMS Tsel"],
              ]

       return steps

def changePPPrepaid(name, PPTo, PreloadBonus, PreloadBonusPPTo):
       preloadBonusSplit           = PreloadBonusPPTo.split(";") #For PP Before (Not PP Name)
       preloadBonusPPBefore        = PreloadBonus.split(";") #For PP Name are used

       preloadBonusData            = preloadBonusSplit[0] #For PP Before (Not PP Name)
       preloadBonusDataPPBefore    = preloadBonusPPBefore[0] #For PP Name are used

       preloadBonusDataString             = preloadBonusData+"MB Internet Perdana" if preloadBonusData != '' else 'No Bonus' #For PP Before (Not PP Name)
       preloadBonusPPBeforeDataString     = preloadBonusDataPPBefore+"MB Internet Perdana" if preloadBonusDataPPBefore != '' else 'No Bonus' #For PP Name are used

       RGUsed                      = ["17","55","75","77"]

       restPreloadBonusDataFirst = int(preloadBonusData) if preloadBonusData != '' else 0 #for PP Before (not PP Name)
       restPreloadBonusDataSecond = int(preloadBonusDataPPBefore) if preloadBonusDataPPBefore != '' else 0 #for PP name (PP Used)

       checkedStepPPTo = []
       checkedStepPPName = []
       while int(restPreloadBonusDataFirst) > 0:
              random_numbers = generate_ordered_multiple_random_numbers(base=15, count=5)
              random_value = random.choice(RGUsed)
              random_number = random.choice(random_numbers)
              restPreloadBonusData = restPreloadBonusDataFirst-random_number
              if restPreloadBonusData > 0:
                     steps = [
                            [f"Create event GPRS {random_number} MB with RG {random_value}","Consume Bonus",f"{restPreloadBonusData}MB Internet Perdana"],
                     ]
                     checkedStepPPTo.extend(steps)
              else:
                     steps = [
                            [f"Create event GPRS {restPreloadBonusDataFirst} MB with RG {random_value}","Consume Bonus","No Bonus"],
                     ]
                     checkedStepPPTo.extend(steps)
                     break
              restPreloadBonusDataFirst -= random_number 
       
       while int(restPreloadBonusDataSecond) > 0:
              random_numbers = generate_ordered_multiple_random_numbers(base=15, count=5)
              random_value = random.choice(RGUsed)
              random_number = random.choice(random_numbers)
              restPreloadBonusDataForSecondPP = restPreloadBonusDataSecond-random_number
              if restPreloadBonusDataForSecondPP > 0:
                     steps = [
                            [f"Create event GPRS {random_number} MB with RG {random_value}","Consume Bonus",f"{restPreloadBonusDataForSecondPP}MB Internet Perdana"],
                     ]
                     checkedStepPPName.extend(steps)
              else:
                     steps = [
                            [f"Create event GPRS {restPreloadBonusDataSecond} MB with RG {random_value}","Consume Bonus","No Bonus"],
                     ]
                     checkedStepPPName.extend(steps)
                     break
              restPreloadBonusDataSecond -= random_number 

       steps = [
              [f"Create & Activate new subscriber PP {PPTo}", "Check active period", preloadBonusDataString],
              ["Update exp date 2020-12-31", "Success", preloadBonusDataString],
              #Add Checked Step PP To
              [f"Change PP to PP {name}", "Success", preloadBonusPPBeforeDataString],
              ["Check Notification Message", "CHANSP Notif Code", preloadBonusPPBeforeDataString],
              ["Update balance 1.000.000", "Success", preloadBonusPPBeforeDataString],
              ["Check Bonus 889", "Success", preloadBonusPPBeforeDataString],
              #Add Checked Step PP Name
              [f"Create and Activate new PP {name}", "Success", preloadBonusPPBeforeDataString],
              [f"Change Price Plan to PP {PPTo}", "Success", preloadBonusDataString],
              ["Check Notification Message", "No Notification", preloadBonusDataString],
              ["Check Bonus 889", "Success", preloadBonusDataString],
              ["Create event voice 1s", "Charged", preloadBonusDataString],
              ["Create event 1 sms onnet", "Charged", preloadBonusDataString],
              ["Update Balance 1.000.000", "Success", preloadBonusDataString]
              #Add Checked Step PP To
       ]

       index = steps.index([f"Change PP to PP {name}", "Success", preloadBonusPPBeforeDataString])
       steps[index:index] = checkedStepPPTo

       index = steps.index([f"Create and Activate new PP {name}", "Success", preloadBonusPPBeforeDataString])
       steps[index:index] = checkedStepPPName

       steps.extend(checkedStepPPTo)
       
       return steps

def getZoneMapping(name, zoneNumber):
    # Load the Excel file
    workbook = openpyxl.load_workbook("Database/Zone Mapping.xlsx")
    sheet = workbook.active

    # Initialize data_dict
    data_dict = {}

    # Get the column headers (zone names) from the first row
    column_headers = [sheet.cell(row=1, column=col).value for col in range(2, sheet.max_column + 1)]

    # Read data from the second row onwards and populate the dictionary
    for row_num in range(2, sheet.max_row + 1):
        key = str(sheet.cell(row=row_num, column=1).value)
        data_dict[key] = {}
        for col_num in range(2, sheet.max_column + 1):
            zone_name = column_headers[col_num - 2]
            value = sheet.cell(row=row_num, column=col_num).value
            data_dict[key][zone_name] = value

    return data_dict[str(zoneNumber)][name]

def tarifSMSPrepaid(name, ratePerZone):
       #A;B;C;D;E;F;G;H;Special <-- Format for rate per zone
       ratePerZoneSplit = ratePerZone.split(";")
       if len(ratePerZoneSplit) > 1:
              rateZone = {
                     'a'            : ratePerZoneSplit[0],
                     'b'            : ratePerZoneSplit[1],
                     'c'            : ratePerZoneSplit[2],
                     'd'            : ratePerZoneSplit[3],
                     'e'            : ratePerZoneSplit[4],
                     'f'            : ratePerZoneSplit[5],
                     'g'            : ratePerZoneSplit[6],
                     'h'            : ratePerZoneSplit[7],
                     'special zone'      : ratePerZoneSplit[8],
              }
       else:
              rateZone = {
                     'a'            : ratePerZoneSplit[0],
                     'b'            : ratePerZoneSplit[0],
                     'c'            : ratePerZoneSplit[0],
                     'd'            : ratePerZoneSplit[0],
                     'e'            : ratePerZoneSplit[0],
                     'f'            : ratePerZoneSplit[0],
                     'g'            : ratePerZoneSplit[0],
                     'h'            : ratePerZoneSplit[0],
                     'special zone' : ratePerZoneSplit[0],
              }

       zone     = 1
       checkedStep   = []
       while zone <= 54:
              zoneMapping   = allZoneMapping[str(zone)][name]
              chargeRate    = rateZone[zoneMapping.lower()]
              smsVal        = random.randint(1, 10)
              chargeSMS     = smsVal*int(chargeRate)
              steps = [
                     [f"Create event {smsVal} SMS onnet on zone id {zone}",f"Charged {chargeSMS} IDR", "No Bonus"],
              ]
              checkedStep.extend(steps)
              zone += 1

       steps = [
              [f"Create & Activate new subscriber PP {name}", "Check active period", "150 MB Internet Perdana"],
              ["Update Balance 500K", "Success", "Balance added"],
              ["Check bonus 889*2", "Success", "No Bonus"],
              ["Check bonus 889*1", "Success", "150 MB Internet Perdana"],
              ["Create event GPRS 150MB RG17, 1PM", "Consume Bonus", "No Bonus"],
              #add checked step
              ["Create event 1 SMS international to Malaysia (+60) , 1PM D+10", "Charged not 350", "No Bonus"],
              ["Create event direct debit using vascode rw_asia_3in1_7000, 1PM D+10", "Charged not 350", "No Bonus"],
              ["Create event direct debit using vascode charged google 55000 5PM D+10", "Charged not 350", "No Bonus"],
              ["Create event 1 SMS spesial number 97080, 5PM D+10", "Charged not 350", "No Bonus"],
              ["Create event 1 SMS international to +1 , 5PM D+11", "Charged 1500 IDR", "No Bonus"],
              ["Create event 1 SMS international to +5 , 8PM D+11", "Charged 1500 IDR", "No Bonus"],
              ["Create event 1 SMS international to +3 , 8PM D+11", "Charged 1500 IDR", "No Bonus"],
              ["Create event 1 SMS international to +4 , 8PM D+11", "Charged 1500 IDR", "No Bonus"],
              ["Create event 1 SMS international to +6 , 8PM D+11", "Charged 1500 IDR", "No Bonus"],
              ["Create event 1 SMS international to +6 , 8PM D+11", "Charged 1500 IDR", "No Bonus"],              
       ]

       index = steps.index(["Create event 1 SMS international to Malaysia (+60) , 1PM D+10", "Charged not 350", "No Bonus"])
       steps[index:index] = checkedStep

       return steps

def exportExcelPrepaidOffer(eventName, params=None, neededParams = None):
       wb = Workbook()
       ws = wb.active

       offerType            = ''
       offerName            = ''
       PPName               = ''
       quota                = ''
       bonusDesc            = ''
       preloadBonus         = ''
       validity             = ''
       timeband             = ''
       eligible             = ''
       startDateValidity    = ''
       endDateValidity      = ''
       endDateValidity60    = ''
       endDateValidityBack  = ''
       itemId               = ''
       allowance            = ''
       accessCodePositif    = ''
       accessCodeNegatif    = ''
       bonDesc              = ''
       countryPositif       = ''
       countryNegatif       = ''

       for params in params:
              if "Offer Type" in params:
                     offerType = params['Offer Type'][0]
              else:
                     offerType = ''
              
              if "Offer Name" in params:
                     offerName = params['Offer Name']
              else:
                     offerName = '' 
              
              if offerName == '':
                     continue
              
              if "Price Plan Name" in params:
                     PPName = params['Price Plan Name']
              else:
                     PPName = ''
              
              if "Quota" in params:
                     quota = params["Quota"]
              
              if "Bonus Description" in params:
                     bonusDesc = params["Bonus Description"][0]
              
              if "Eligible" in params:
                     eligible = params["Eligible"][0]
              
              if "Start Date Validity" in params:
                     startDateValidity = params["Start Date Validity"]
              
              if "End Date Validity" in params:
                     endDateValidity = params["End Date Validity"]
              
              if "End Date Validity For More Than 60 Days" in params:
                     endDateValidity60 = params["End Date Validity For More Than 60 Days"]
              
              if "End Date Validity For Back Date" in params:
                     endDateValidityBack = params["End Date Validity For Back Date"]
              
              if "Item ID" in params:
                     itemId = params["Item ID"]
              
              if "Allowance" in params:
                     allowance = params["Allowance"]
              
              if "Preload Bonus" in params:
                     preloadBonus = params["Preload Bonus"]
              
              if "Validity" in params:
                     validity = params["Validity"]

              if "Timeband" in params:
                     timeband = params["Timeband"]
              
              if "Access Code (for positif test case)" in params:
                     accessCodePositif = params["Access Code (for positif test case)"]
              
              if "Access Code (for negatif test case)" in params:
                     accessCodeNegatif = params["Access Code (for negatif test case)"]
              
              if "Bonus Desc" in params:
                     bonDesc = params["Bonus Desc"]
              
              if "Country (for positif test case)" in params:
                     countryPositif = params["Country (for positif test case)"]
              
              if "Country (for negatif test case)" in params:
                     countryNegatif = params["Country (for negatif test case)"]
              
              if offerType == 'Offer Fix':
                     steps = stepOfferFix(offerName, PPName, quota, bonusDesc, preloadBonus, validity, timeband)
              elif offerType == "Offer TM":
                     steps = stepOfferTM(offerName, PPName, quota, bonusDesc, preloadBonus, validity, timeband)
              elif offerType == 'Offer Flexible':
                     steps = stepOfferFlexible(offerName, PPName, preloadBonus, eligible, bonusDesc, startDateValidity, endDateValidity, endDateValidity60, endDateValidityBack, itemId, allowance, timeband)
              elif offerType == 'Offer Voice IDD':
                     steps = stepOfferVoiceIDD(offerName, accessCodePositif, accessCodeNegatif, allowance, validity, timeband, bonDesc, countryPositif, countryNegatif)
              elif offerType == 'Offer Voice IDD (Offer Flexible)':
                     steps = stepOfferVoiceIDDFlexible(offerName, PPName, preloadBonus, startDateValidity, endDateValidity, endDateValidity60, endDateValidityBack, itemId, allowance, timeband, accessCodePositif, accessCodeNegatif, bonDesc, countryPositif, countryNegatif)
              else:
                     print("Sorry, Scenario isn't ready yet")
                     exit('')

              # Write Header Row
              header = [f'{eventName} | {offerName}']
              ws.append(header)

              # Merge Header Cells
              startColumnIndex = 3  # Example of a dynamic column index
              startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
              endColumn = "E"
              startRow = 1
              endRow = 1
              cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
              ws.merge_cells(cellRange)

              headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
              ws.append(headerRow)

              no = 1
              for num, step in enumerate(steps, start=1):
                     if isinstance(step, str):
                            row = [
                                   no,
                                   step,
                                   "Success",
                                   "No Bonus",
                                   "XYZ"
                            ]
                            no = no+1
                     else:
                            if step is None:
                                   continue
                            else:
                                   if len(step) == 5:
                                          row = [
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 step[3],
                                                 step[4]
                                          ]
                                   elif len(step) == 4:
                                          row = [
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 step[3],
                                                 "XYZ"
                                          ]
                                   elif len(step) == 3:
                                          row = [
                                                 no,
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 "XYZ"
                                          ]
                                          no = no+1
                                   else:
                                          row = [
                                                 no,
                                                 step[0],
                                                 step[1],
                                                 "No Bonus",
                                                 "XYZ"
                                          ]
                                          no = no+1
                     ws.append(row)

       print("Testing Scenario Successfully Generated")
       
       # Save Excel File
       wb.save('Result/Scenario '+str(eventName)+' '+str(offerType)+'.xlsx')

def stepOfferFix(offerName, PPName, Quota, bonusDesc, preloadBonus, validity, timeband):
       QuotaSplitString     = Quota.split(';')
       QuotaSMS             = 0
       firstQuotaSMS        = 0
       QuotaVoice           = int(QuotaSplitString[0])
       firstQuotaVoice      = int(QuotaSplitString[0])
       stepConsumePreload   = None
       bonusPreload         = 'No Bonus'
       bonusVoice           = 'No Bonus'
       bonusSMS             = 'No Bonus'
       start_hour, end_hour = map(int, timeband.split('-'))
       validity             = int(validity)
       

       if preloadBonus != '' and preloadBonus != 0:
              stepConsumePreload   = ["Consume Bonus Preload","Consume Bonus","No Bonus"]
              bonusPreload         = preloadBonus

       if bonusDesc == 'All Opr':
              stringBonus = "All Opr"
       elif bonusDesc == 'Tsel (Onnet, Onbrand for Loop)':
              stringBonus = "Tsel"
       else:
              stringBonus = "Opr Lain"

       if len(QuotaSplitString) > 1:
              QuotaSMS             = int(QuotaSplitString[1])
              firstQuotaSMS        = int(QuotaSplitString[1])
       
       # if QuotaVoice > 0:
       #       stepsConsumeVoice, QuotaVoice, detailQuotaVoice = getStepRecudeVoice(QuotaVoice, stringBonus, QuotaSMS, dataEvent, bonusDesc, start_hour, end_hour, validity)

       # if QuotaSMS > 0:
       #        stepsConsumeSMS = getStepReduceSMS(QuotaVoice, stringBonus, QuotaSMS, dataEvent, bonusDesc, start_hour, end_hour, validity, detailQuotaVoice)

       if QuotaVoice > 0 or QuotaSMS > 0:
              stepsConsumeBonus, QuotaVoice, QuotaSMS = getStepRecudeQuota(QuotaVoice, QuotaSMS, stringBonus, bonusDesc, start_hour, end_hour, validity)

       stringBonusAll = ''
       if firstQuotaVoice > 0:
              stringBonusAll = str(firstQuotaVoice)+" Min "+stringBonus
              bonusVoice     = str(firstQuotaVoice)+" Min "+stringBonus
       
       if firstQuotaSMS > 0:
              stringBonusAll = stringBonusAll+" "+str(firstQuotaSMS)+" SMS "+stringBonus
              bonusSMS       = str(firstQuotaSMS)+" SMS "+stringBonus
       
       steps = [
              ["Create & Activate new subscriber PP "+PPName,"Check active period",bonusPreload],
              stepConsumePreload,
              ["Update Balance 10000000","Balance Updated","No Bonus"],
              ["Update Exp Date","Update Expired Date","No Bonus"],
              ["Attach Offer "+str(offerName)+" Voice "+str(stringBonus)+"","Offer Attached",stringBonusAll],
              ["Check Bonus 889*1 11am","Checked","No Bonus"],
              ["Check Bonus 889*2 11am","Checked",bonusVoice],
              ["Check Bonus 889*3 11am","Checked",bonusSMS]
       ]

       if firstQuotaVoice > 0 or firstQuotaSMS > 0:
              steps.extend(stepsConsumeBonus)

       # if firstQuotaVoice > 0:
       #        steps.extend(stepsConsumeVoice)
       # if firstQuotaSMS > 0:
       #        steps.extend(stepsConsumeSMS)

       return steps

def stepOfferTM(offerName, PPName, Quota, bonusDesc, preloadBonus, validity, timeband):
       QuotaSplitString     = Quota.split(';')
       QuotaSMS             = 0
       firstQuotaSMS        = 0
       QuotaVoice           = int(QuotaSplitString[0])
       firstQuotaVoice      = int(QuotaSplitString[0])
       stepConsumePreload   = None
       bonusPreload         = 'No Bonus'
       bonusVoice           = 'No Bonus'
       bonusSMS             = 'No Bonus'
       start_hour, end_hour = map(int, timeband.split('-'))
       validity             = int(validity)
       

       if preloadBonus != '' and preloadBonus != 0:
              stepConsumePreload   = ["Consume Bonus Preload","Consume Bonus","No Bonus"]
              bonusPreload         = preloadBonus

       if bonusDesc == 'All Opr':
              stringBonus = "All Opr"
       elif bonusDesc == 'Tsel (Onnet, Onbrand for Loop)':
              stringBonus = "Tsel"
       else:
              stringBonus = "Opr Lain"

       if len(QuotaSplitString) > 1:
              QuotaSMS             = int(QuotaSplitString[1])
              firstQuotaSMS        = int(QuotaSplitString[1])
       
       # if QuotaVoice > 0:
       #       stepsConsumeVoice, QuotaVoice, detailQuotaVoice = getStepRecudeVoice(QuotaVoice, stringBonus, QuotaSMS, dataEvent, bonusDesc, start_hour, end_hour, validity)

       # if QuotaSMS > 0:
       #        stepsConsumeSMS = getStepReduceSMS(QuotaVoice, stringBonus, QuotaSMS, dataEvent, bonusDesc, start_hour, end_hour, validity, detailQuotaVoice)

       if QuotaVoice > 0 or QuotaSMS > 0:
              stepsConsumeBonus, QuotaVoice, QuotaSMS = getStepRecudeQuota(QuotaVoice, QuotaSMS, stringBonus, bonusDesc, start_hour, end_hour, validity)

       stringBonusAll = ''
       if firstQuotaVoice > 0:
              stringBonusAll = str(firstQuotaVoice)+" Min "+stringBonus
              bonusVoice     = str(firstQuotaVoice)+" Min "+stringBonus
       
       if firstQuotaSMS > 0:
              stringBonusAll = stringBonusAll+" "+str(firstQuotaSMS)+" SMS "+stringBonus
              bonusSMS       = str(firstQuotaSMS)+" SMS "+stringBonus
       
       steps = [
              ["Create & Activate new subscriber PP "+PPName,"Check active period",bonusPreload],
              stepConsumePreload,
              ["Update Balance 10000000","Balance Updated","No Bonus"],
              ["Update Exp Date","Update Expired Date","No Bonus"],
              ["Attach Offer "+str(offerName)+" Voice "+str(stringBonus)+"","Offer Attached",stringBonusAll],
              ["Check Bonus 889*1 11am","Checked","No Bonus"],
              ["Check Bonus 889*2 11am","Checked",bonusVoice],
              ["Check Bonus 889*3 11am","Checked",bonusSMS]
       ]

       if firstQuotaVoice > 0 or firstQuotaSMS > 0:
              steps.extend(stepsConsumeBonus)

       # if firstQuotaVoice > 0:
       #        steps.extend(stepsConsumeVoice)
       # if firstQuotaSMS > 0:
       #        steps.extend(stepsConsumeSMS)

       return steps

def stepOfferFlexible(offerName, PPName, preloadBonus, eligible, bonusDesc, startDateValidity, endDateValidity, endDateValidity60, endDateValidityBack, itemId, allowance, timeband):
       steps                = []
       stepConsumePreload   = None
       start_hour, end_hour = map(int, timeband.split('-'))

       itemIDSplit   = itemId.split(';')
       itemIDVoice   = itemIDSplit[0]
       itemIDSMS     = 0
       if len(itemIDSplit) > 1:
              itemIDSMS = itemIDSplit[1]
       
       allowanceSplit       = allowance.split(';')
       allowanceVoice       = int(allowanceSplit[0])
       QuotaVoice           = int(allowanceVoice/60) if allowanceVoice != 0 else 0
       QuotaVoiceCase2      = 0
       QuotaVoiceCase3      = 0
       QuotaVoiceCase4      = int(allowanceVoice/60) if allowanceVoice != 0 else 0
       firstQuotaVoice      = QuotaVoice
       allowanceSMS         = 0
       if len(allowanceSplit) > 1:
              allowanceSMS  = int(allowanceSplit[1])
       QuotaSMS             = allowanceSMS if allowanceSMS != 0 else 0
       QuotaSMSCase2        = 0
       QuotaSMSCase3        = 0
       QuotaSMSCase4        = allowanceSMS if allowanceSMS != 0 else 0
       firstQuotaSMS        = QuotaSMS

       if preloadBonus != '' and preloadBonus != 0:
              stepConsumePreload   = ["Consume Bonus Preload","Consume Bonus","No Bonus"]
       
       if bonusDesc == 'All Opr':
              stringBonus = "All Opr"
       elif bonusDesc == 'Tsel (Onnet, Onbrand for Loop)':
              stringBonus = "Tsel"
       else:
              stringBonus = "Opr Lain"

       if eligible == 'Voice':
              UOM = 'V'
              attachOfferString = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOM}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity}235959$;'
              attachOfferStringCase2 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}$S$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity}235959$;'
              attachOfferStringCase3 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOM}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidityBack}235959$;'
              attachOfferStringCase4 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOM}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity60}235959$;'
       elif eligible == 'SMS':
              UOM = 'S'
              attachOfferString = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDSMS}${UOM}$0.3${allowanceSMS}${startDateValidity}000000${endDateValidity}235959$;'
              attachOfferStringCase2 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDSMS}$V$0.3${allowanceSMS}${startDateValidity}000000${endDateValidity}235959$;'
              attachOfferStringCase3 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOM}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidityBack}235959$;'
              attachOfferStringCase4 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDSMS}${UOM}$0.3${allowanceSMS}${startDateValidity}000000${endDateValidity60}235959$;'
       elif eligible == 'Voice & SMS':
              UOMV = 'V'
              UOMS = 'S'
              attachOfferString = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOMV}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity}235959$;{itemIDSMS}${UOMS}$0.3${allowanceSMS}${startDateValidity}000000${endDateValidity}235959$'
              attachOfferStringCase2 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}$O$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity}235959$;{itemIDSMS}$O$0.3${allowanceSMS}${startDateValidity}000000${endDateValidity}235959$'
              attachOfferStringCase3 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOMV}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidityBack}235959$;{itemIDSMS}${UOMS}$0.3${allowanceSMS}${startDateValidity}000000${endDateValidityBack}235959$'
              attachOfferStringCase4 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOMV}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity60}235959$;{itemIDSMS}${UOMS}$0.3${allowanceSMS}${startDateValidity}000000${endDateValidity60}235959$'


       stringBonusAll       = ''
       bonusVoice           = 'No Bonus'
       bonusSMS             = 'No Bonus'
       if firstQuotaVoice > 0:
              stringBonusAll = str(firstQuotaVoice)+" Min "+stringBonus
              bonusVoice     = str(firstQuotaVoice)+" Min "+stringBonus
       if firstQuotaSMS > 0:
              stringBonusAll = stringBonusAll+" "+str(firstQuotaSMS)+" SMS "+stringBonus
              bonusSMS       = str(firstQuotaSMS)+" SMS "+stringBonus
       
       start_datetime       = datetime.strptime(str(startDateValidity), '%Y%m%d')
       end_datetime         = datetime.strptime(str(endDateValidity), '%Y%m%d')
       end_datetimecase4    = datetime.strptime(str(endDateValidity60), '%Y%m%d')
       
       validity      = (end_datetime - start_datetime).days
       validityCase4 = (end_datetimecase4 - start_datetime).days

       if QuotaVoice > 0 or QuotaSMS > 0:
              stepsConsumeBonus, QuotaVoice, QuotaSMS = getStepRecudeQuota(QuotaVoice, QuotaSMS, stringBonus, bonusDesc, start_hour, end_hour, validity)
              stepsConsumeBonusCase4, QuotaVoiceCase4, QuotaSMSCase4 = getStepRecudeQuota(QuotaVoiceCase4, QuotaSMSCase4, stringBonus, bonusDesc, start_hour, end_hour, validityCase4)
              

       #Case 1 = Positif Case
       stepCase1 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonus],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferString,"Offer attached",stringBonusAll],
              ["Check Bonus 889*1","Bonus Checked","No Bonus"],
              ["Check Bonus 889*2","Bonus Checked",bonusVoice],
              ["Check Bonus 889*3","Bonus Checked",bonusSMS],
              ["Check Bonus 889*4","Bonus Checked","No Bonus"],
              #Reduce Allowance
       ]

       stepCase1.extend(stepsConsumeBonus)
       stepCase1.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 2 = Negatif (Berdasarkan UOM)
       stepsConsumeBonusCase2, QuotaVoiceCase2, QuotaSMSCase2 = getStepRecudeQuota(QuotaVoiceCase2, QuotaSMSCase2, stringBonus, bonusDesc, start_hour, end_hour, 5)
       stepCase2 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonus],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferStringCase2,"Offer attached","No Bonus"],
              ["Check Bonus 889*1","Bonus Checked","No Bonus"],
              ["Check Bonus 889*2","Bonus Checked","No Bonus"],
              ["Check Bonus 889*3","Bonus Checked","No Bonus"],
              ["Check Bonus 889*4","Bonus Checked","No Bonus"],
              #Reduce Allowance
       ]
       stepCase2.extend(stepsConsumeBonusCase2)
       stepCase2.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 3 = Negatif (Backdate)
       stepsConsumeBonusCase3, QuotaVoiceCase3, QuotaSMSCase3 = getStepRecudeQuota(QuotaVoiceCase3, QuotaSMSCase3, stringBonus, bonusDesc, start_hour, end_hour, 5)
       stepCase3 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonus],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferStringCase3,"Offer attached","No Bonus"],
              ["Check Bonus 889*1","Bonus Checked","No Bonus"],
              ["Check Bonus 889*2","Bonus Checked","No Bonus"],
              ["Check Bonus 889*3","Bonus Checked","No Bonus"],
              ["Check Bonus 889*4","Bonus Checked","No Bonus"],
              #Reduce Allowance
       ]
       stepCase3.extend(stepsConsumeBonusCase3)
       stepCase3.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 4 = Positif (Lebih dari 60 hari)
       stepCase4 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonus],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferStringCase4,"Offer attached",stringBonusAll],
              ["Check Bonus 889*1","Bonus Checked","No Bonus"],
              ["Check Bonus 889*2","Bonus Checked",bonusVoice],
              ["Check Bonus 889*3","Bonus Checked",bonusSMS],
              ["Check Bonus 889*4","Bonus Checked","No Bonus"],
              #Reduce Allowance
       ]

       stepCase4.extend(stepsConsumeBonusCase4)
       stepCase4.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 5 = Multiple Attach (6x)
       bonus6x = ""
       if firstQuotaVoice > 0:
              totalVoice    = firstQuotaVoice*6
              bonus6x       = str(totalVoice)+" Min "+stringBonus
       
       if firstQuotaSMS > 0:
              totalSMS      = firstQuotaSMS*6
              if bonus6x != '':
                     bonus6x = str(bonus6x)+", "+str(totalSMS)+" SMS "+stringBonus
              else:
                     bonus6x = str(totalSMS)+" SMS "+stringBonus

       stepCase5 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonus],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferString,"Offer attached",stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              ["Check 889","Checked",bonus6x],
              ["Check on database","Success",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              ["Check PI on Indira","Success","No Bonus"]
       ]

       #Case 6
       stepCase6 = [
              [f"{offerName} SCN For check BSZ Extract | D-2", "", "", "", ""],
              [f"Create & Activate new subscriber PP {PPName}", "Success", preloadBonus],
              stepConsumePreload,
              ["Update Exp Date", "Success", "No Bonus"],
              ["Update Balance 10000000", "Balance Update", "No Bonus"],
              [attachOfferString, "Offer attached", stringBonusAll],
              ["Check notification after add offer", "Success", "Checked"],
              ["Check Offer Name & Description", "Success", offerName],
              ["Check GetBonusInfo and validity", "Success", "Checked"],
              ["Check Bonus 889 and bonus description", "Success", "Checked"],
              ["Check Bonus 889*1", "Success", "No Bonus"],
              ["Check Bonus 889*2", "Success", bonusVoice],
              ["Check Bonus 889*3", "Success", bonusSMS],
              ["Check Bonus 889*4", "Success", "Success"],
              ["Check PRIT Name", "Success", "Success"],
              ["Create event vas with eligible vascode param_vascode", "Success", "Success"],
              ["Check notification after first event consume", "Success", "Success"],
              ["Create event voice Onnet 60s", "Success", "Success"],
              ["run adjustment so it will expired by today -2", "Success", "No Bonus"],
              ["check bonus info (bonus should be gone)", "Success", "No Bonus"],
              ["check 888", "Success", "No Bonus"],
              ["run BSZ eod, and check bsz seizure", "Success", "No Bonus"],
       ]

       steps.extend(stepCase1)
       steps.extend(stepCase2)
       steps.extend(stepCase3)
       steps.extend(stepCase4)
       steps.extend(stepCase5)
       steps.extend(stepCase6)

       return steps

def stepOfferVoiceIDD(offerName, accessCodePositif, accessCodeNegatif, allowance, validity, timeband, bonDesc, countryPositif, countryNegatif):
       steps = []

       allowanceSplit       = allowance.split(';')
       voiceIDD             = int(allowanceSplit[0])/60
       allowanceString      = str(int(voiceIDD))+" Min "+bonDesc
       countryPositifSplit  = countryPositif.split(';')
       firstCountryPos      = countryPositifSplit[0]
       countryPositifData   = [{"name": name, "status": "Positif"} for name in countryPositifSplit]
       countryNegatifSplit  = countryNegatif.split(';') 
       countryNegatifData   = [{"name": name} for name in countryNegatifSplit] 
       mergedCountryData    = countryPositifData + countryNegatifData
       accessCodePosSplit   = accessCodePositif.split(";")
       firstAccessCodePos   = accessCodePosSplit[0]
       accessCodePosData    = [{"name": name, "status": "Positif"} for name in accessCodePosSplit]
       accessCodeNegSplit   = accessCodeNegatif.split(";")
       accessCodeNegData    = [{"name": name, "status": "Negatif"} for name in accessCodeNegSplit]
       mergedAccessCode     = accessCodePosData + accessCodeNegData
       start_hour, end_hour = map(int, timeband.split('-'))

       step = [
              ["Create & Activate new subscriber PP KartuAS Regular","Check active period","No Bonus"],
              ["Update Balance 1000000","Success","No Bonus"],
              [f"Attach offer Prepaid {offerName}","Offer attached",allowanceString],
              ["Check bonus 889*2","Checked",allowanceString],
              ["Check bonus 889*3","Checked","No Bonus"]
       ]

       stepConsumeVoiceIDD = getStepReduceQuotaVoiceIDD(voiceIDD, mergedCountryData, validity, start_hour, end_hour, bonDesc, mergedAccessCode, firstCountryPos, firstAccessCodePos, countryPositifData, accessCodePosData)
       
       steps.extend(step)
       steps.extend(stepConsumeVoiceIDD)

       return steps

def stepOfferVoiceIDDFlexible(offerName, PPName, preloadBonus, startDateValidity, endDateValidity, endDateValidity60, endDateValidityBack, itemId, allowance, timeband, accessCodePositif, accessCodeNegatif, bonDesc, countryPositif, countryNegatif):
       steps                = []
       stepConsumePreload   = None
       start_hour, end_hour = map(int, timeband.split('-'))
       eligible             = 'Voice'
       itemIDSplit          = itemId.split(';')
       itemIDVoice          = itemIDSplit[0]
       allowanceSplit       = allowance.split(';')
       allowanceVoice       = int(allowanceSplit[0])
       QuotaVoice           = int(allowanceVoice/60) if allowanceVoice != 0 else 0
       QuotaVoiceCase2      = 0
       QuotaVoiceCase3      = 0
       QuotaVoiceCase4      = int(allowanceVoice/60) if allowanceVoice != 0 else 0
       firstQuotaVoice      = QuotaVoice
       countryPositifSplit  = countryPositif.split(';')
       firstCountryPos      = countryPositifSplit[0]
       countryPositifData   = [{"name": name, "status": "Positif"} for name in countryPositifSplit]
       countryNegatifSplit  = countryNegatif.split(';') 
       countryNegatifData   = [{"name": name} for name in countryNegatifSplit] 
       mergedCountryData    = countryPositifData + countryNegatifData
       accessCodePosSplit   = accessCodePositif.split(";")
       firstAccessCodePos   = accessCodePosSplit[0]
       accessCodePosData    = [{"name": name, "status": "Positif"} for name in accessCodePosSplit]
       accessCodeNegSplit   = accessCodeNegatif.split(";")
       accessCodeNegData    = [{"name": name, "status": "Negatif"} for name in accessCodeNegSplit]
       mergedAccessCode     = accessCodePosData + accessCodeNegData

       if preloadBonus != '' and preloadBonus != 0:
              stepConsumePreload   = ["Consume Bonus Preload","Consume Bonus","No Bonus"]

       if eligible == 'Voice':
              UOM = 'V'
              attachOfferString = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOM}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity}235959$;'
              attachOfferStringCase2 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}$S$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity}235959$;'
              attachOfferStringCase3 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOM}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidityBack}235959$;'
              attachOfferStringCase4 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOM}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity60}235959$;'


       stringBonusAll       = ''
       bonusVoice           = 'No Bonus'
       if firstQuotaVoice > 0:
              stringBonusAll = str(firstQuotaVoice)+" Min "+bonDesc
              bonusVoice     = str(firstQuotaVoice)+" Min "+bonDesc
       
       start_datetime       = datetime.strptime(str(startDateValidity), '%Y%m%d')
       end_datetime         = datetime.strptime(str(endDateValidity), '%Y%m%d')
       end_datetimecase4    = datetime.strptime(str(endDateValidity60), '%Y%m%d')
       
       validity      = (end_datetime - start_datetime).days
       validityCase4 = (end_datetimecase4 - start_datetime).days

       if QuotaVoice > 0:
              stepsConsumeBonus = getStepReduceQuotaVoiceIDD(QuotaVoice, mergedCountryData, validity, start_hour, end_hour, bonDesc, mergedAccessCode, firstCountryPos, firstAccessCodePos, countryPositifData, accessCodePosData)
              stepsConsumeBonusCase4 = getStepReduceQuotaVoiceIDD(QuotaVoiceCase4, mergedCountryData, validityCase4, start_hour, end_hour, bonDesc, mergedAccessCode, firstCountryPos, firstAccessCodePos, countryPositifData, accessCodePosData)
              

       #Case 1 = Positif Case
       stepCase1 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonus],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferString,"Offer attached",stringBonusAll],
              ["Check Bonus 889*1","Bonus Checked","No Bonus"],
              ["Check Bonus 889*2","Bonus Checked",bonusVoice],
              ["Check Bonus 889*3","Bonus Checked","No Bonus"],
              ["Check Bonus 889*4","Bonus Checked","No Bonus"],
              #Reduce Allowance
       ]

       stepCase1.extend(stepsConsumeBonus)
       stepCase1.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 2 = Negatif (Berdasarkan UOM)
       stepsConsumeBonusCase2 = getStepReduceQuotaVoiceIDD(QuotaVoiceCase2, mergedCountryData, 1, start_hour, end_hour, bonDesc, mergedAccessCode, firstCountryPos, firstAccessCodePos, countryPositifData, accessCodePosData)
       stepCase2 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonus],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferStringCase2,"Offer attached","No Bonus"],
              #Reduce Allowance
       ]
       stepCase2.extend(stepsConsumeBonusCase2)
       stepCase2.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 3 = Negatif (Backdate)
       stepsConsumeBonusCase3 = getStepReduceQuotaVoiceIDD(QuotaVoiceCase3, mergedCountryData, 1, start_hour, end_hour, bonDesc, mergedAccessCode, firstCountryPos, firstAccessCodePos, countryPositifData, accessCodePosData)
       stepCase3 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonus],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferStringCase3,"Offer attached","No Bonus"],
              #Reduce Allowance
       ]
       stepCase3.extend(stepsConsumeBonusCase3)
       stepCase3.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 4 = Positif (Lebih dari 60 hari)
       stepCase4 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonus],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferStringCase4,"Offer attached",stringBonusAll],
              #Reduce Allowance
       ]

       stepCase4.extend(stepsConsumeBonusCase4)
       stepCase4.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 5 = Multiple Attach (6x)
       bonus6x = ""
       if firstQuotaVoice > 0:
              totalVoice    = firstQuotaVoice*6
              bonus6x       = str(totalVoice)+" Min "+bonDesc

       stepCase5 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonus],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferString,"Offer attached",stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              ["Check 889","Checked",bonus6x],
              ["Check on database","Success",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              ["Check PI on Indira","Success","No Bonus"]
       ]

       #Case 6
       stepCase6 = [
              [f"{offerName} SCN For check BSZ Extract | D-2", "", "", "", ""],
              [f"Create & Activate new subscriber PP {PPName}", "Success", preloadBonus],
              stepConsumePreload,
              ["Update Exp Date", "Success", "No Bonus"],
              ["Update Balance 10000000", "Balance Update", "No Bonus"],
              [attachOfferString, "Offer attached", stringBonusAll],
              ["Check notification after add offer", "Success", "Checked"],
              ["Check Offer Name & Description", "Success", offerName],
              ["Check GetBonusInfo and validity", "Success", "Checked"],
              ["Check Bonus 889 and bonus description", "Success", "Checked"],
              ["Check PRIT Name", "Success", "Success"],
              ["Create event vas with eligible vascode param_vascode", "Success", "Success"],
              ["Check notification after first event consume", "Success", "Success"],
              ["Create event voice Onnet 60s", "Success", "Success"],
              ["run adjustment so it will expired by today -2", "Success", "No Bonus"],
              ["check bonus info (bonus should be gone)", "Success", "No Bonus"],
              ["check 888", "Success", "No Bonus"],
              ["run BSZ eod, and check bsz seizure", "Success", "No Bonus"],
       ]

       steps.extend(stepCase1)
       steps.extend(stepCase2)
       steps.extend(stepCase3)
       steps.extend(stepCase4)
       steps.extend(stepCase5)
       steps.extend(stepCase6)

       return steps

def getStepReduceQuotaVoiceIDD(QuotaVoice, countryData, validity, start_hour, end_hour, bonDesc, mergedAccessCode, firstCountryPos, firstAccessCodePos, countryPosData, accessCodePosData):
       stepsConsume         = []
       dayString            = 0
       validity             = int(validity)
       maxValidity          = validity+1
       
       # Generate a shuffled list of numbers from dayString to validity - 1
       days          = list(range(dayString, maxValidity))
       firstDate     = days[0]
       lastDate      = days[len(days) - 1]

       if validity > 1:
              # Choose a random number of elements to select from data
              num_elements = random.randint(1, len(days) - 2)

              # Randomly select elements from data
              selected_data = random.sample(days[1:-1], num_elements)

              # Sort selected_data based on the index in the original data list
              selected_data = sorted(selected_data, key=lambda x: days.index(x))

              # Merge variables into a single list
              merged_data = [firstDate] + selected_data + [lastDate]
       else:
              merged_data = [firstDate] + [lastDate]
       
       random.shuffle(mergedAccessCode)

       if len(countryData) < len(merged_data):
              random.shuffle(countryData)
              stepsConsume = validateStepNormal(QuotaVoice, merged_data, countryData, mergedAccessCode, start_hour, end_hour, validity, bonDesc, firstCountryPos, firstAccessCodePos, countryPosData, accessCodePosData)
       else:
              stepsConsume = validateStepShortValidity(QuotaVoice, merged_data, countryData, mergedAccessCode, start_hour, end_hour, validity, bonDesc, firstCountryPos, firstAccessCodePos)

       return stepsConsume

def getStepReduceSMS(QuotaVoice, stringBonus, QuotaSMS, dataEvent, bonusDesc, start_hour, end_hour, validity, detailQuotaVoice):
       stepsConsumeSMS      = []
       dayString            = 0
       timeString           = ''
       consumeOrCharged     = ''
       restBonus            = 'No Bonus'
       validity             = int(validity)
       maxValidity          = validity+5
       reduceOrNot          = False
       
       # Generate a shuffled list of numbers from dayString to validity - 1
       days = list(range(dayString, maxValidity))

       for strValidity in days:
              random_event  = random.choice(dataEvent)
              event_name    = random_event["Name"]
              event_param   = random_event["Param"]
              timeNumber    = random.randint(0, 23)
              timeString    = timeNumber
              if len(detailQuotaVoice) > 0:
                     filtered_dicts = [item for item in detailQuotaVoice if item["Day"] == strValidity]

                     # Check if there are any matches
                     if filtered_dicts:
                            # Extract the "Quota" value from the first matching dictionary (assuming there's only one)
                            quota_value = filtered_dicts[0]["Quota"]
                            stringQuotaVoice = stringBonus + ' ' + str(quota_value) + " minutes,"
                     else:
                            stringQuotaVoice = ''
              else:
                     if QuotaVoice > 0:
                            stringQuotaVoice = stringBonus + ' ' + str(QuotaVoice) + " minutes,"
                     else:
                            stringQuotaVoice = ''

              # Check if random event is included in bonus desc
              if bonusDesc in event_param:
                     if strValidity < validity:
                            if start_hour <= end_hour:
                                   # Time range does not span midnight
                                   if start_hour <= timeNumber <= end_hour:
                                          # Number is within the time range Timeband
                                          consumeOrCharged = 'Consume Bonus'
                                          reduceOrNot      = True
                                   else:
                                          # Number is not within the time range Timeband
                                          consumeOrCharged = 'Charged'
                                          reduceOrNot      = True
                            else:
                                   # Time range spans midnight
                                   if timeNumber >= start_hour or timeNumber <= end_hour:
                                          # Number is within the time range Timeband
                                          consumeOrCharged = 'Consume Bonus'
                                          reduceOrNot      = True
                                   else:
                                          # Number is not within the time range Timeband
                                          consumeOrCharged = 'Charged'
                                          reduceOrNot      = True
                     else:
                            consumeOrCharged = 'Charged'
                            reduceOrNot      = True

              else:
                     consumeOrCharged = 'Charged'

              if QuotaSMS > 0 and reduceOrNot:
                     decreasingQuotaSMS = round((QuotaSMS * 0.5) / 4)
                     QuotaSMS -= decreasingQuotaSMS
                     eventString = decreasingQuotaSMS
              else:
                     eventString = '1'

              if int(timeString) > 12:
                     timeString = str(int(timeString) - 12) + 'PM'
              else:
                     timeString = str(timeString) + "AM"

              if int(strValidity) >= validity:
                     restBonus = "No Bonus"
              else:
                     restBonus = f"{stringQuotaVoice} {stringBonus} {QuotaSMS} sms"

              step = [
                     f"Create event {eventString} SMS {event_name} {timeString} D+{strValidity}",
                     consumeOrCharged,
                     restBonus
              ]
              stepsConsumeSMS.append(step)

       return stepsConsumeSMS

def getStepRecudeVoice(QuotaVoice, stringBonus, QuotaSMS, dataEvent, bonusDesc, start_hour, end_hour, validity):
       if QuotaSMS > 0:
              stringQuotaSMS = ", " + stringBonus + ' ' + str(QuotaSMS) + " sms"
       else:
              stringQuotaSMS = ''

       stepsConsumeVoice    = []
       dayString            = 0
       timeString           = ''
       consumeOrCharged     = ''
       restBonus            = 'No Bonus'
       validity             = int(validity)
       maxValidity          = validity+5
       reduceOrNot          = False
       detailQuotaVoice     = []
       
       # Generate a shuffled list of numbers from dayString to validity - 1
       days = list(range(dayString, maxValidity))

       for strValidity in days:
              random_event  = random.choice(dataEvent)
              event_name    = random_event["Name"]
              event_param   = random_event["Param"]
              timeNumber    = random.randint(0, 23)
              timeString    = timeNumber

              # Check if random event is included in bonus desc
              if bonusDesc in event_param:
                     if strValidity < validity:
                            if start_hour <= end_hour:
                                   # Time range does not span midnight
                                   if start_hour <= timeNumber <= end_hour:
                                          # Number is within the time range Timeband
                                          consumeOrCharged = 'Consume Bonus'
                                          reduceOrNot      = True
                                   else:
                                          # Number is not within the time range Timeband
                                          consumeOrCharged = 'Charged'
                                          reduceOrNot      = False
                            else:
                                   # Time range spans midnight
                                   if timeNumber >= start_hour or timeNumber <= end_hour:
                                          # Number is within the time range Timeband
                                          consumeOrCharged = 'Consume Bonus'
                                          reduceOrNot      = True
                                   else:
                                          # Number is not within the time range Timeband
                                          consumeOrCharged = 'Charged'
                                          reduceOrNot      = False
                     else:
                            consumeOrCharged     = 'Charged'
                            reduceOrNot          = False

              else:
                     consumeOrCharged = 'Charged'
                     reduceOrNot      = False

              if QuotaVoice > 0 and reduceOrNot:
                     decreasingQuotaVoice = round((QuotaVoice * 0.5) / 4)
                     QuotaVoice -= decreasingQuotaVoice
                     eventString = decreasingQuotaVoice
              else:
                     eventString = '1'

              if int(timeString) > 12:
                     timeString = str(int(timeString) - 12) + 'PM'
              else:
                     timeString = str(timeString) + "AM"

              if int(strValidity) >= validity:
                     restBonus = "No Bonus"
                     detailVoiceQuota = {
                            "Day" : strValidity,
                            "Quota" : "Not Found"
                     }
              else:
                     restBonus = f"{stringBonus} {QuotaVoice} minutes {stringQuotaSMS}"
                     detailVoiceQuota = {
                            "Day" : strValidity,
                            "Quota" : QuotaVoice
                     }

              step = [
                     f"Create event {eventString} minutes voice {event_name} {timeString} D+{strValidity}",
                     consumeOrCharged,
                     restBonus
              ]
              stepsConsumeVoice.append(step)
              detailQuotaVoice.append(detailVoiceQuota)

       return stepsConsumeVoice, QuotaVoice, detailQuotaVoice

def getStepRecudeQuota(QuotaVoice, QuotaSMS, stringBonus, bonusDesc, start_hour, end_hour, validity):
       stepsConsume         = []
       dayString            = 0
       validity             = int(validity)
       maxValidity          = validity+1
       first                = True
       dataEvent            = [
              {
                     "Name" : 'Onnet',
                     "Param" : ["All Opr", "Tsel (Onnet, Onbrand for Loop)"],
                     "ShowEvent" : True
              },
              {
                     "Name" : 'Offnet',
                     "Param" : ["All Opr", "Opr Lain (Include fwa,pstn)", "Opr Lain (Exclude fwa,pstn)"],
                     "ShowEvent" : True
              },
              {
                     "Name" : 'FWA',
                     "Param" : ["All Opr", "Opr Lain (Include fwa,pstn)"],
                     "ShowEvent" : True
              },
              {
                     "Name" : 'International',
                     "Param" : [],
                     "ShowEvent" : True
              },
              {
                     "Name" : 'GPRS 1MB RG 50 ',
                     "Param" : [],
                     "ShowEvent" : False
              },
              {
                     "Name" : 'Direct Debit bank_digi_250 ',
                     "Param" : [],
                     "ShowEvent" : False
              },
       ]

       # Iterate through the list and add the "Priority" key based on the mapping
       for item in dataEvent:
              if bonusDesc in item["Param"]:
                     item["Priority"] = 1
              else:
                     item["Priority"] = 0
       
       # Generate a shuffled list of numbers from dayString to validity - 1
       days          = list(range(dayString, maxValidity))
       firstDate     = days[0]
       lastDate      = days[len(days) - 1]

       # Choose a random number of elements to select from data
       num_elements = random.randint(1, len(days) - 2)

       # Randomly select elements from data
       selected_data = random.sample(days[1:-1], num_elements)

       # Sort selected_data based on the index in the original data list
       selected_data = sorted(selected_data, key=lambda x: days.index(x))

       # Merge variables into a single list
       merged_data = [firstDate] + selected_data + [lastDate]

       for strValidity in merged_data:
              stepConsumeVoice, QuotaVoice       = getStepConsumeVoice(QuotaVoice, QuotaSMS, stringBonus, dataEvent, bonusDesc, start_hour, end_hour, strValidity, validity, first)
              stepConsumeSMS, QuotaSMS           = getStepConsumeSMS(QuotaVoice, stringBonus, QuotaSMS, dataEvent, bonusDesc, start_hour, end_hour, strValidity, validity, first)
              stepsConsume.append(stepConsumeVoice)
              stepsConsume.append(stepConsumeSMS)
              first = False

       return stepsConsume, QuotaVoice, QuotaSMS

def getStepConsumeVoice(QuotaVoice, QuotaSMS, stringBonus, dataEvent, bonusDesc, start_hour, end_hour, days, validity, first):
       
       if QuotaSMS > 0:
              stringQuotaSMS = stringBonus + ' ' + str(QuotaSMS) + " sms"
       else:
              stringQuotaSMS = ''
       
       timeString           = ''
       consumeOrCharged     = ''
       restBonus            = 'No Bonus'
       reduceOrNot          = False
 
       if first:
              random_event = next((item for item in dataEvent if item["Priority"] == 1), None)
              event_name    = random_event["Name"]
              event_param   = random_event["Param"]
              event_show    = random_event["ShowEvent"]
              timeNumber    = random.randint(start_hour, end_hour)
       else:
              random_event  = random.choice(dataEvent)
              event_name    = random_event["Name"]
              event_param   = random_event["Param"]
              event_show    = random_event["ShowEvent"]
              timeNumber    = random.randint(0, 23)
       
       timeString    = timeNumber

       # Check if random event is included in bonus desc
       if bonusDesc in event_param:
              if days < validity:
                     if start_hour <= end_hour:
                            # Time range does not span midnight
                            if start_hour <= timeNumber <= end_hour:
                                   # Number is within the time range Timeband
                                   consumeOrCharged = 'Consume Bonus'
                                   reduceOrNot      = True
                            else:
                                   # Number is not within the time range Timeband
                                   consumeOrCharged = 'Charged'
                                   reduceOrNot      = False
                     else:
                            # Time range spans midnight
                            if timeNumber >= start_hour or timeNumber <= end_hour:
                                   # Number is within the time range Timeband
                                   consumeOrCharged = 'Consume Bonus'
                                   reduceOrNot      = True
                            else:
                                   # Number is not within the time range Timeband
                                   consumeOrCharged = 'Charged'
                                   reduceOrNot      = False
              else:
                     consumeOrCharged     = 'Charged'
                     reduceOrNot          = False

       else:
              consumeOrCharged = 'Charged'
              reduceOrNot      = False

       if QuotaVoice > 0 and reduceOrNot:
              decreasingQuotaVoice = round((QuotaVoice * 0.5) / 4)
              QuotaVoice -= decreasingQuotaVoice
              eventString = decreasingQuotaVoice
       else:
              eventString = '1'
              consumeOrCharged = 'Charged'

       if int(timeString) > 12:
              timeString = str(int(timeString) - 12) + 'PM'
       else:
              timeString = str(timeString) + "AM"
       
       if QuotaSMS <= 0 and QuotaVoice > 0:
              restBonus = f"{stringBonus} {QuotaVoice} minutes"
       elif QuotaSMS > 0 and QuotaVoice <= 0:
              restBonus = f"{stringQuotaSMS}"
       elif QuotaSMS <= 0 and QuotaVoice <= 0:
              restBonus = "No Bonus" 
       else:
              restBonus = f"{stringBonus} {QuotaVoice} minutes, {stringQuotaSMS}"

       if int(days) >= validity:
              restBonus = "No Bonus"

       if event_show:
              eventLabel = f"Create event {eventString} minutes voice {event_name} {timeString} D+{days}"
       else:
              eventLabel = f"Create event {event_name} {timeString} D+{days}"

       step = [
              eventLabel,
              consumeOrCharged,
              restBonus
       ]

       return step, QuotaVoice

def getStepConsumeSMS(QuotaVoice, stringBonus, QuotaSMS, dataEvent, bonusDesc, start_hour, end_hour, days, validity, first):
       
       timeString           = ''
       consumeOrCharged     = ''
       restBonus            = 'No Bonus'
       reduceOrNot          = False

       if first:
              random_event = next((item for item in dataEvent if item["Priority"] == 1), None)
              event_name    = random_event["Name"]
              event_param   = random_event["Param"]
              event_show    = random_event["ShowEvent"]
              timeNumber    = random.randint(start_hour, end_hour)
       else:
              random_event  = random.choice(dataEvent)
              event_name    = random_event["Name"]
              event_param   = random_event["Param"]
              event_show    = random_event["ShowEvent"]
              timeNumber    = random.randint(0, 23)
       
       timeString           = timeNumber

       if QuotaVoice > 0:
              stringQuotaVoice = stringBonus + ' ' + str(QuotaVoice) + " minutes"
       else:
              stringQuotaVoice = ''

       # Check if random event is included in bonus desc
       if bonusDesc in event_param:
              if days < validity:
                     if start_hour <= end_hour:
                            # Time range does not span midnight
                            if start_hour <= timeNumber <= end_hour:
                                   # Number is within the time range Timeband
                                   consumeOrCharged = 'Consume Bonus'
                                   reduceOrNot      = True
                            else:
                                   # Number is not within the time range Timeband
                                   consumeOrCharged = 'Charged'
                                   reduceOrNot      = False
                     else:
                            # Time range spans midnight
                            if timeNumber >= start_hour or timeNumber <= end_hour:
                                   # Number is within the time range Timeband
                                   consumeOrCharged = 'Consume Bonus'
                                   reduceOrNot      = True
                            else:
                                   # Number is not within the time range Timeband
                                   consumeOrCharged = 'Charged'
                                   reduceOrNot      = False
              else:
                     consumeOrCharged = 'Charged'
                     reduceOrNot      = False

       else:
              consumeOrCharged = 'Charged'

       if QuotaSMS > 0 and reduceOrNot:
              decreasingQuotaSMS = round((QuotaSMS * 0.5) / 4)
              QuotaSMS -= decreasingQuotaSMS
              eventString = decreasingQuotaSMS
       else:
              eventString = '1'
              consumeOrCharged = 'Charged'

       if int(timeString) > 12:
              timeString = str(int(timeString) - 12) + 'PM'
       else:
              timeString = str(timeString) + "AM"

       if QuotaSMS <= 0 and QuotaVoice > 0:
              restBonus = f"{stringQuotaVoice}"
       elif QuotaSMS <= 0 and QuotaVoice <= 0:
              restBonus = "No Bonus"
       elif QuotaSMS > 0 and QuotaVoice <= 0:
              restBonus = f"{stringBonus} {QuotaSMS} sms"
       else:
              restBonus = f"{stringQuotaVoice}, {stringBonus} {QuotaSMS} sms"

       if int(days) >= validity:
              restBonus = "No Bonus"

       if event_show:
              eventLabel = f"Create event {eventString} SMS {event_name} {timeString} D+{days}"
       else:
              eventLabel = f"Create event {event_name} {timeString} D+{days}"

       step = [
              eventLabel,
              consumeOrCharged,
              restBonus
       ]

       return step, QuotaSMS

def generatingScenario(eventName, offerName, offerDesc, offerType, steps):
       wb = Workbook()
       ws = wb.active

       # Write Header Row
       header = [f'{eventName} | {offerName} | {offerDesc}']
       ws.append(header)

       # Merge Header Cells
       startColumnIndex = 3  # Example of a dynamic column index
       startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
       endColumn = "E"
       startRow = 1
       endRow = 1
       cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
       ws.merge_cells(cellRange)

       headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
       ws.append(headerRow)

       for no, step in enumerate(steps):
              if isinstance(step, str):
                     row = [
                            no,
                            step,
                            "Success",
                            "No Bonus",
                            "XYZ"
                     ]
                     no = no+1
              else:
                     if step is None:
                            continue
                     else:
                            if len(step) == 5:
                                   row = [
                                          step[0],
                                          step[1],
                                          step[2],
                                          step[3],
                                          step[4]
                                   ]
                            elif len(step) == 4:
                                   row = [
                                          step[0],
                                          step[1],
                                          step[2],
                                          step[3],
                                          "XYZ"
                                   ]
                            elif len(step) == 3:
                                   row = [
                                          no,
                                          step[0],
                                          step[1],
                                          step[2],
                                          "XYZ"
                                   ]
                                   no = no+1
                            else:
                                   row = [
                                          no,
                                          step[0],
                                          step[1],
                                          "No Bonus",
                                          "XYZ"
                                   ]
                                   no = no+1
              ws.append(row)

       print("Testing Scenario Successfully Generated")
       
       # Save Excel File
       wb.save('Result/Scenario '+str(eventName)+' '+str(offerType)+'.xlsx')

def validateStepNormal(QuotaVoice, merged_data, countryData, mergedAccessCode, start_hour, end_hour, validity, bonDesc, firstCountryPos, firstAccessCodePos, countryPosData, accessCodePosData):
       stepsConsume         = []
       additionalNegatifCase = [
              "Create Voice Onnet 1 Min",
              "Create event 1 SMS onnet",
              "Create event GPRS 1MB RG 50",
              "Create Event Voice Offnet 5s",
              "Create Event Voice PSTN 5s",
              "Create Event Voice FWA 180s",
              f"Create Event Voice Roaming MOC Home in {firstCountryPos}"
       ]

       getData       = 0
       getAccessCode = 0
       count         = 1

       #Variable for count how much country positive and access code positive is out
       priorityOut   = 0

       for strValidity in merged_data:
              if priorityOut >= len(countryPosData):
                     if getData == len(countryData):
                            getData = 0
                     if getAccessCode == len(mergedAccessCode):
                            getAccessCode = 0
                     country       = countryData[getData]
                     countryName   = country["name"]

                     accessCode           = mergedAccessCode[getAccessCode]
                     accessCodeName       = accessCode["name"]
                     accessCodeStatus     = accessCode["status"]
                     accessCodeUsed       = accessCodeName
              else:
                     country       = countryPosData[priorityOut]
                     countryName   = country["name"]

                     accessCode           = random.choice(accessCodePosData)
                     accessCodeName       = accessCode["name"]
                     accessCodeStatus     = accessCode["status"]
                     accessCodeUsed       = accessCodeName

                     priorityOut += 1
              
              timeNumber    = random.randint(start_hour, end_hour)
              timeString    = timeNumber
              
              if "status" in country:
                     if accessCodeStatus == 'Positif':
                            if strValidity < validity:
                                   if start_hour <= end_hour:
                                          # Time range does not span midnight
                                          if start_hour <= timeNumber <= end_hour:
                                                 # Number is within the time range Timeband
                                                 consumeOrCharged     = 'Consume Bonus'
                                                 reduceOrNot          = True
                                          else:
                                                 # Number is not within the time range Timeband
                                                 consumeOrCharged     = 'Charged'
                                                 reduceOrNot          = False
                                   else:
                                          # Time range spans midnight
                                          if timeNumber >= start_hour or timeNumber <= end_hour:
                                                 # Number is within the time range Timeband
                                                 consumeOrCharged     = 'Consume Bonus'
                                                 reduceOrNot          = True
                                          else:
                                                 # Number is not within the time range Timeband
                                                 consumeOrCharged     = 'Charged'
                                                 reduceOrNot          = False
                            else:
                                   consumeOrCharged     = 'Charged'
                                   reduceOrNot          = False
                     else:
                            consumeOrCharged     = 'Charged'
                            reduceOrNot          = False
              else:
                     consumeOrCharged     = 'Charged'
                     reduceOrNot          = False
                     accessCodeUsed       = firstAccessCodePos
              
              if QuotaVoice > 0 and reduceOrNot:
                     decreasingQuotaVoice = round((QuotaVoice * 0.5) / 4)
                     QuotaVoice -= decreasingQuotaVoice
                     eventString = decreasingQuotaVoice
              else:
                     eventString = '1'
                     consumeOrCharged = 'Charged'

              if int(timeString) > 12:
                     timeString = str(int(timeString) - 12) + 'PM'
              else:
                     timeString = str(timeString) + "AM"
              
              QuotaVoice    = int(QuotaVoice)
              if QuotaVoice > 0:
                     restBonus = f"{QuotaVoice} Min {bonDesc}"
              elif QuotaVoice <= 0:
                     restBonus = "No Bonus" 

              if int(strValidity) >= validity:
                     restBonus = "No Bonus"

              eventLabel = f"Create event voice IDD to {countryName} using access code {accessCodeUsed} {timeString} {eventString} Min D+{strValidity}"

              step = [
                     eventLabel,
                     consumeOrCharged,
                     restBonus
              ]

              stepsConsume.append(step)

              if len(additionalNegatifCase) > 0:
                     showOrNot = random.randint(0,1)
                     if showOrNot == 1:
                            selected_value = random.choice(additionalNegatifCase)
                            stepAdd = [
                                   str(selected_value)+" "+str(timeString)+" D+"+str(strValidity),
                                   "Charged",
                                   restBonus
                            ]
                            stepsConsume.append(stepAdd)
                            additionalNegatifCase.remove(selected_value)
              
              if count == len(merged_data):
                     stepLast = [
                            f"Create event voice IDD to {firstCountryPos} using access code {firstAccessCodePos} {timeString} {eventString} Min D+{strValidity}",
                            "Charged",
                            restBonus
                     ]
                     stepsConsume.append(stepLast)
                            
              getData += 1
              count += 1
              getAccessCode += 1

       return stepsConsume

def validateStepShortValidity(QuotaVoice, merged_data, countryDatas, mergedAccessCode, start_hour, end_hour, validity, bonDesc, firstCountryPos, firstAccessCodePos):
       stepsConsume         = []
       additionalNegatifCase = [
              "Create Voice Onnet 1 Min",
              "Create event 1 SMS onnet",
              "Create event GPRS 1MB RG 50",
              "Create Event Voice Offnet 5s",
              "Create Event Voice PSTN 5s",
              "Create Event Voice FWA 180s",
              f"Create Event Voice Roaming MOC Home in {firstCountryPos}"
       ]
       getData       = 0
       getAccessCode = 0
       count         = 1

       for strValidity in merged_data:
              for countryData in countryDatas:
                     if getAccessCode == len(mergedAccessCode):
                            getAccessCode = 0
                     country       = countryData
                     countryName   = country["name"]

                     accessCode           = mergedAccessCode[getAccessCode]
                     accessCodeName       = accessCode["name"]
                     accessCodeStatus     = accessCode["status"]
                     accessCodeUsed       = accessCodeName
                     
                     timeNumber    = random.randint(start_hour, end_hour)
                     timeString    = timeNumber
              
                     if accessCodeStatus == 'Positif':
                            if "status" in country:
                                   if strValidity < validity:
                                          if start_hour <= end_hour:
                                                 # Time range does not span midnight
                                                 if start_hour <= timeNumber <= end_hour:
                                                        # Number is within the time range Timeband
                                                        consumeOrCharged     = 'Consume Bonus'
                                                        reduceOrNot          = True
                                                 else:
                                                        # Number is not within the time range Timeband
                                                        consumeOrCharged     = 'Charged'
                                                        reduceOrNot          = False
                                          else:
                                                 # Time range spans midnight
                                                 if timeNumber >= start_hour or timeNumber <= end_hour:
                                                        # Number is within the time range Timeband
                                                        consumeOrCharged     = 'Consume Bonus'
                                                        reduceOrNot          = True
                                                 else:
                                                        # Number is not within the time range Timeband
                                                        consumeOrCharged     = 'Charged'
                                                        reduceOrNot          = False
                                   else:
                                          consumeOrCharged     = 'Charged'
                                          reduceOrNot          = False
                            else:
                                   consumeOrCharged     = 'Charged'
                                   reduceOrNot          = False
                     else:
                            consumeOrCharged     = 'Charged'
                            reduceOrNot          = False
                     
                     if QuotaVoice > 0 and reduceOrNot:
                            decreasingQuotaVoice = round((QuotaVoice * 0.5) / 4)
                            QuotaVoice -= decreasingQuotaVoice
                            eventString = decreasingQuotaVoice
                     else:
                            eventString = '1'
                            consumeOrCharged = 'Charged'

                     if int(timeString) > 12:
                            timeString = str(int(timeString) - 12) + 'PM'
                     else:
                            timeString = str(timeString) + "AM"
                     
                     QuotaVoice    = int(QuotaVoice)
                     if QuotaVoice > 0:
                            restBonus = f"{QuotaVoice} Min {bonDesc}"
                     elif QuotaVoice <= 0:
                            restBonus = "No Bonus" 

                     if int(strValidity) >= validity:
                            restBonus = "No Bonus"

                     eventLabel = f"Create event voice IDD to {countryName} using access code {accessCodeUsed} {timeString} {eventString} Min D+{strValidity}"

                     step = [
                            eventLabel,
                            consumeOrCharged,
                            restBonus
                     ]

                     stepsConsume.append(step)

                     if len(additionalNegatifCase) > 0:
                            showOrNot = random.randint(0,1)
                            if showOrNot == 1:
                                   selected_value = random.choice(additionalNegatifCase)
                                   stepAdd = [
                                          str(selected_value)+" "+str(timeString)+" D+"+str(strValidity),
                                          "Charged",
                                          restBonus
                                   ]
                                   stepsConsume.append(stepAdd)
                                   additionalNegatifCase.remove(selected_value)
                     
                     if count == len(merged_data):
                            stepLast = [
                                   f"Create event voice IDD to {firstCountryPos} using access code {firstAccessCodePos} {timeString} {eventString} Min D+{strValidity}",
                                   "Charged",
                                   restBonus
                            ]
                            stepsConsume.append(stepLast)
                                   
                     getData += 1
                     count += 1
                     getAccessCode += 1

       return stepsConsume

def exportExcelOfferRoaming(eventName, params=None, neededParams = None):
       wb = Workbook()
       ws = wb.active

       cardType             = ''
       offerType            = ''
       offerTypePostpaid    = ''
       offerName            = ''
       PPName               = ''
       preloadBonus         = ''
       eligible             = ''
       bonusDesc            = ''
       MOEligible           = ''
       MTEligible           = ''
       vascodePositif       = ''
       vascodeNegatif       = ''
       countryPositif       = ''
       countryNegatif       = ''
       validity             = ''
       startDateValidity    = ''
       endDateValidity      = ''
       endDateValidity60    = ''
       endDateValidityBack  = ''
       itemId               = ''
       allowance            = ''
       timeband             = ''
       quota                = ''
       cls                  = ''
       scenarioChoosen      = ''
       RCIndicator          = ''
       validityEndDate      = ''
       for params in params:
              if "Card Type" in params:
                     cardType = params['Card Type'][0]
              else:
                     cardType = ''

              if "Offer Type Prepaid" in params:
                     offerType = params['Offer Type Prepaid'][0]
              else:
                     offerType = ''
              
              if "Offer Type Postpaid" in params:
                     offerTypePostpaid = params['Offer Type Postpaid'][0]
              else:
                     offerTypePostpaid = ''
              
              if "Offer Name" in params:
                     offerName = params['Offer Name']
              else:
                     offerName = '' 
              
              if offerName == '':
                     continue
              
              if "Price Plan Name" in params:
                     PPName = params['Price Plan Name']
              else:
                     PPName = ''
              
              if "Bonus Description" in params:
                     bonusDesc = params["Bonus Description"]
              
              if "Credit Limit Service" in params:
                     cls = params["Credit Limit Service"]
              
              if "Eligible" in params:
                     dataEligible = {
                            "1" : "Voice",
                            "2" : "SMS",
                            "3" : "Voice & SMS"
                     }
                     if isinstance(params["Eligible"], list):
                            eligible = params["Eligible"][0]
                     else:
                            if params["Eligible"].isnumeric():
                                   eligible = dataEligible[params["Eligible"]]
                            else:
                                   eligible = params["Eligible"]
              
              if "Start Date Validity" in params:
                     startDateValidity = params["Start Date Validity"]
              
              if "End Date Validity" in params:
                     endDateValidity = params["End Date Validity"]
              
              if "End Date Validity For More Than 60 Days" in params:
                     endDateValidity60 = params["End Date Validity For More Than 60 Days"]
              
              if "End Date Validity For Back Date" in params:
                     endDateValidityBack = params["End Date Validity For Back Date"]
              
              if "Item ID" in params:
                     itemId = params["Item ID"]
              
              if "Allowance" in params:
                     allowance = params["Allowance"]
              
              if "Preload Bonus" in params:
                     preloadBonus = params["Preload Bonus"]
              
              if "Validity" in params:
                     validity = params["Validity"]

              if "Timeband" in params:
                     timeband = params["Timeband"]
              
              if "Country (for positif test case)" in params:
                     countryPositif = params["Country (for positif test case)"]
              
              if "Country (for negatif test case)" in params:
                     countryNegatif = params["Country (for negatif test case)"]
              
              if "MO Eligible" in params:
                     MOEligible = params["MO Eligible"]
              
              if "MT Eligible" in params:
                     MTEligible = params["MT Eligible"]
              
              if "Vascode (for positif test case)" in params:
                     vascodePositif = params["Vascode (for positif test case)"]
              
              if "Vascode (for negatif test case)" in params:
                     vascodeNegatif = params["Vascode (for negatif test case)"]
              
              if "Quota" in params:
                     quota = params["Quota"]
              
              if "Scenario" in params:
                     scenarioChoosen = params["Scenario"]
              
              if "RC Indicator" in params:
                     RCIndicator = params["RC Indicator"]
              
              if "Validity End Date" in params:
                     validityEndDate = params["Validity End Date"]
              
              if cardType == 'Prepaid':
                     if offerType == 'Offer Flexible':
                            steps = getStepOfferRoamingPrepaidFlexibleOffer(offerName, PPName, preloadBonus, eligible, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, validity, startDateValidity, endDateValidity, endDateValidity60, endDateValidityBack, itemId, allowance, timeband)
                     elif offerType == 'Offer Fix':
                            steps = getStepOfferRoamingPrepaidFixOffer(offerName, PPName, preloadBonus, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, validity, timeband, quota)
                     else:
                            print("Sorry, Scenario isn't ready yet")
                            exit('') 
              else:
                     if offerTypePostpaid == 'Offer Fix':
                            steps = getStepOfferRoamingPostpaidFixOffer(offerName, PPName, preloadBonus, cls, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, validity, timeband, quota)
                     elif offerTypePostpaid == 'Offer Flexible':
                            steps = getStepOfferRoamingPostpaidFlexibleOffer(offerName, PPName, cls, eligible, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, startDateValidity, allowance, scenarioChoosen, RCIndicator, validityEndDate, endDateValidityBack)
                     else:
                            print("Sorry, Scenario isn't ready yet")
                            exit('')

              # Write Header Row
              header = [f'{eventName} | {offerName}']
              ws.append(header)

              # Merge Header Cells
              startColumnIndex = 3  # Example of a dynamic column index
              startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
              endColumn = "E"
              startRow = 1
              endRow = 1
              cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
              ws.merge_cells(cellRange)

              headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
              ws.append(headerRow)

              no = 1
              
              for num, step in enumerate(steps, start=1):
                     if isinstance(step, str):
                            row = [
                                   no,
                                   step,
                                   "Success",
                                   "No Bonus",
                                   "XYZ"
                            ]
                            no = no+1
                     else:
                            if step is None:
                                   continue
                            else:
                                   if len(step) == 5:
                                          row = [
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 step[3],
                                                 step[4]
                                          ]
                                   elif len(step) == 4:
                                          row = [
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 step[3],
                                                 "XYZ"
                                          ]
                                   elif len(step) == 3:
                                          row = [
                                                 no,
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 "XYZ"
                                          ]
                                          no = no+1
                                   else:
                                          row = [
                                                 no,
                                                 step[0],
                                                 step[1],
                                                 "No Bonus",
                                                 "XYZ"
                                          ]
                                          no = no+1
                     ws.append(row)

       print("Testing Scenario Successfully Generated")
       
       # Save Excel File
       wb.save('Result/Scenario '+str(eventName)+' '+str(cardType)+' '+str(offerTypePostpaid)+' '+str(offerType)+'.xlsx')

def getStepOfferRoamingPrepaidFlexibleOffer(offerName, PPName, preloadBonus, eligible, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, validity, startDateValidity, endDateValidity, endDateValidity60, endDateValidityBack, itemId, allowance, timeband):
       steps                = []
       stepConsumePreload   = None
       start_hour, end_hour = map(int, timeband.split('-'))

       itemIDSplit   = itemId.split(';')
       itemIDVoice   = itemIDSplit[0]
       itemIDSMS     = 0
       if len(itemIDSplit) > 1:
              itemIDSMS = itemIDSplit[1]
       
       allowanceSplit       = allowance.split(';')
       allowanceVoice       = int(allowanceSplit[0])
       QuotaVoice           = int(allowanceVoice/60) if allowanceVoice != 0 else 0
       QuotaVoiceCase2      = 0
       QuotaVoiceCase3      = 0
       QuotaVoiceCase4      = int(allowanceVoice/60) if allowanceVoice != 0 else 0
       firstQuotaVoice      = QuotaVoice
       allowanceSMS         = 0
       if len(allowanceSplit) > 1:
              allowanceSMS  = int(allowanceSplit[1])
       QuotaSMS             = allowanceSMS if allowanceSMS != 0 else 0
       QuotaSMSCase2        = 0
       QuotaSMSCase3        = 0
       QuotaSMSCase4        = allowanceSMS if allowanceSMS != 0 else 0
       firstQuotaSMS        = QuotaSMS

       if preloadBonus != '' and preloadBonus != 0 and preloadBonus != "0":
              stepConsumePreload   = ["Consume Bonus Preload","Consume Bonus","No Bonus"]
              preloadBonusString = preloadBonus
       else:
              preloadBonusString = "No Bonus"
       

       if eligible == 'Voice':
              UOM = 'V'
              attachOfferString = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOM}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity}235959$;'
              attachOfferStringCase2 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}$S$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity}235959$;'
              attachOfferStringCase3 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOM}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidityBack}235959$;'
              attachOfferStringCase4 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOM}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity60}235959$;'
       elif eligible == 'SMS':
              UOM = 'S'
              attachOfferString = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDSMS}${UOM}$0.3${allowanceSMS}${startDateValidity}000000${endDateValidity}235959$;'
              attachOfferStringCase2 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDSMS}$V$0.3${allowanceSMS}${startDateValidity}000000${endDateValidity}235959$;'
              attachOfferStringCase3 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOM}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidityBack}235959$;'
              attachOfferStringCase4 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDSMS}${UOM}$0.3${allowanceSMS}${startDateValidity}000000${endDateValidity60}235959$;'
       elif eligible == 'Voice & SMS':
              UOMV = 'V'
              UOMS = 'S'
              attachOfferString = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOMV}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity}235959$;{itemIDSMS}${UOMS}$0.3${allowanceSMS}${startDateValidity}000000${endDateValidity}235959$'
              attachOfferStringCase2 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}$O$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity}235959$;{itemIDSMS}$O$0.3${allowanceSMS}${startDateValidity}000000${endDateValidity}235959$'
              attachOfferStringCase3 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOMV}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidityBack}235959$;{itemIDSMS}${UOMS}$0.3${allowanceSMS}${startDateValidity}000000${endDateValidityBack}235959$'
              attachOfferStringCase4 = f'Attach Offer {offerName} with param TransactionID|Product ID|Allow Item Level Cost and value Prod1|Trx1|{itemIDVoice}${UOMV}$0.3${allowanceVoice}${startDateValidity}000000${endDateValidity60}235959$;{itemIDSMS}${UOMS}$0.3${allowanceSMS}${startDateValidity}000000${endDateValidity60}235959$'


       stringBonusAll       = ''
       bonusVoice           = 'No Bonus'
       bonusSMS             = 'No Bonus'
       if firstQuotaVoice > 0:
              stringBonusAll = str(firstQuotaVoice)+" Min "+bonusDesc
              bonusVoice     = str(firstQuotaVoice)+" Min "+bonusDesc
       if firstQuotaSMS > 0:
              stringBonusAll = stringBonusAll+" "+str(firstQuotaSMS)+" SMS "+bonusDesc
              bonusSMS       = str(firstQuotaSMS)+" SMS "+bonusDesc
       
       start_datetime       = datetime.strptime(str(startDateValidity), '%Y%m%d')
       end_datetime         = datetime.strptime(str(endDateValidity), '%Y%m%d')
       end_datetimecase4    = datetime.strptime(str(endDateValidity60), '%Y%m%d')
       
       validity      = (end_datetime - start_datetime).days
       validityCase4 = (end_datetimecase4 - start_datetime).days

       if QuotaVoice > 0 or QuotaSMS > 0:
              stepsConsumeBonus, QuotaVoice, QuotaSMS = getStepReduceQuotaInternational(QuotaVoice, QuotaSMS, bonusDesc, start_hour, end_hour, validity, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif)
              stepsConsumeBonusCase4, QuotaVoiceCase4, QuotaSMSCase4 = getStepReduceQuotaInternational(QuotaVoiceCase4, QuotaSMSCase4, bonusDesc, start_hour, end_hour, validityCase4, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif)
              

       #Case 1 = Positif Case
       stepCase1 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferString,"Offer attached",stringBonusAll],
              ["Check Bonus 889*1","Bonus Checked","No Bonus"],
              ["Check Bonus 889*2","Bonus Checked",bonusVoice],
              ["Check Bonus 889*3","Bonus Checked",bonusSMS],
              ["Check Bonus 889*4","Bonus Checked","No Bonus"],
              #Reduce Allowance
       ]

       stepCase1.extend(stepsConsumeBonus)
       stepCase1.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 2 = Negatif (Berdasarkan UOM)
       stepsConsumeBonusCase2, QuotaVoiceCase2, QuotaSMSCase2 = getStepReduceQuotaInternational(QuotaVoiceCase2, QuotaSMSCase2, bonusDesc, start_hour, end_hour, 1, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif)
       stepCase2 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferStringCase2,"Offer attached","No Bonus"],
              #Reduce Allowance
       ]
       stepCase2.extend(stepsConsumeBonusCase2)
       stepCase2.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 3 = Negatif (Backdate)
       stepsConsumeBonusCase3, QuotaVoiceCase3, QuotaSMSCase3 = getStepReduceQuotaInternational(QuotaVoiceCase3, QuotaSMSCase3, bonusDesc, start_hour, end_hour, 1, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif)
       stepCase3 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferStringCase3,"Offer attached","No Bonus"],
              #Reduce Allowance
       ]
       stepCase3.extend(stepsConsumeBonusCase3)
       stepCase3.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 4 = Positif (Lebih dari 60 hari)
       stepCase4 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferStringCase4,"Offer attached",stringBonusAll],
              #Reduce Allowance
       ]

       stepCase4.extend(stepsConsumeBonusCase4)
       stepCase4.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 5 = Multiple Attach (6x)
       bonus6x = ""
       if firstQuotaVoice > 0:
              totalVoice    = firstQuotaVoice*6
              bonus6x       = str(totalVoice)+" Min "+bonusDesc
       
       if firstQuotaSMS > 0:
              totalSMS      = firstQuotaSMS*6
              if bonus6x != '':
                     bonus6x = str(bonus6x)+", "+str(totalSMS)+" SMS "+bonusDesc
              else:
                     bonus6x = str(totalSMS)+" SMS "+bonusDesc

       stepCase5 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [attachOfferString,"Offer attached",stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [attachOfferString,"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              ["Check 889","Checked",bonus6x],
              ["Check on database","Success",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              ["Check PI on Indira","Success","No Bonus"]
       ]

       #Case 6
       stepCase6 = [
              [f"{offerName} SCN For check BSZ Extract | D-2", "", "", "", ""],
              [f"Create & Activate new subscriber PP {PPName}", "Success", preloadBonusString],
              stepConsumePreload,
              ["Update Exp Date", "Success", "No Bonus"],
              ["Update Balance 10000000", "Balance Update", "No Bonus"],
              [attachOfferString, "Offer attached", stringBonusAll],
              ["Check notification after add offer", "Success", "Checked"],
              ["Check Offer Name & Description", "Success", offerName],
              ["Check GetBonusInfo and validity", "Success", "Checked"],
              ["Check Bonus 889 and bonus description", "Success", "Checked"],
              ["Check PRIT Name", "Success", "Success"],
              ["Create event vas with eligible vascode param_vascode", "Success", "Success"],
              ["Check notification after first event consume", "Success", "Success"],
              ["Create event voice Onnet 60s", "Success", "Success"],
              ["run adjustment so it will expired by today -2", "Success", "No Bonus"],
              ["check bonus info (bonus should be gone)", "Success", "No Bonus"],
              ["check 888", "Success", "No Bonus"],
              ["run BSZ eod, and check bsz seizure", "Success", "No Bonus"],
       ]

       steps.extend(stepCase1)
       steps.extend(stepCase2)
       steps.extend(stepCase3)
       steps.extend(stepCase4)
       steps.extend(stepCase5)
       steps.extend(stepCase6)

       return steps

def getStepOfferRoamingPrepaidFixOffer(offerName, PPName, preloadBonus, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, validity, timeband, quota):
       steps                = []
       stepConsumePreload   = None
       start_hour, end_hour = map(int, timeband.split('-'))
       
       allowanceSplit       = quota.split(';')
       allowanceVoice       = int(allowanceSplit[0])
       QuotaVoice           = int(allowanceVoice) if allowanceVoice != 0 else 0
       firstQuotaVoice      = QuotaVoice
       allowanceSMS         = 0
       if len(allowanceSplit) > 1:
              allowanceSMS  = int(allowanceSplit[1])
       QuotaSMS             = allowanceSMS if allowanceSMS != 0 else 0
       firstQuotaSMS        = QuotaSMS

       if preloadBonus != '' and preloadBonus != 0 and preloadBonus != "0":
              stepConsumePreload   = ["Consume Bonus Preload","Consume Bonus","No Bonus"]
              preloadBonusString = preloadBonus
       else:
              preloadBonusString = "No Bonus"

       stringBonusAll       = ''
       bonusVoice           = 'No Bonus'
       bonusSMS             = 'No Bonus'
       if firstQuotaVoice > 0:
              stringBonusAll = str(firstQuotaVoice)+" Min "+bonusDesc
              bonusVoice     = str(firstQuotaVoice)+" Min "+bonusDesc
       if firstQuotaSMS > 0:
              stringBonusAll = stringBonusAll+" "+str(firstQuotaSMS)+" SMS "+bonusDesc
              bonusSMS       = str(firstQuotaSMS)+" SMS "+bonusDesc

       if QuotaVoice > 0 or QuotaSMS > 0:
              stepsConsumeBonus, QuotaVoice, QuotaSMS = getStepReduceQuotaInternational(QuotaVoice, QuotaSMS, bonusDesc, start_hour, end_hour, validity, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif)
              

       #Case 1 = Positif Case
       stepCase1 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [f"Attach offer {offerName}","Offer attached",stringBonusAll],
              ["Check Bonus 889*1","Bonus Checked","No Bonus"],
              ["Check Bonus 889*2","Bonus Checked",bonusVoice],
              ["Check Bonus 889*3","Bonus Checked",bonusSMS],
              ["Check Bonus 889*4","Bonus Checked","No Bonus"],
              #Reduce Allowance
       ]

       stepCase1.extend(stepsConsumeBonus)
       stepCase1.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 5 = Multiple Attach (6x)
       bonus6x = ""
       if firstQuotaVoice > 0:
              totalVoice    = firstQuotaVoice*6
              bonus6x       = str(totalVoice)+" Min "+bonusDesc
       
       if firstQuotaSMS > 0:
              totalSMS      = firstQuotaSMS*6
              if bonus6x != '':
                     bonus6x = str(bonus6x)+", "+str(totalSMS)+" SMS "+bonusDesc
              else:
                     bonus6x = str(totalSMS)+" SMS "+bonusDesc

       stepCase5 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
              stepConsumePreload,
              ["Update Exp Date","Updated","No Bonus"],
              ["Update Balance 1000000","Balance Updated","No Bonus"],
              [f"Attach offer {offerName}","Offer attached",stringBonusAll],
              [f"Attach offer {offerName}","Offer attached",stringBonusAll+" , "+stringBonusAll],
              [f"Attach offer {offerName}","Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [f"Attach offer {offerName}","Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [f"Attach offer {offerName}","Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [f"Attach offer {offerName}","Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              ["Check 889","Checked",bonus6x],
              ["Check on database","Success",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              ["Check PI on Indira","Success","No Bonus"]
       ]

       steps.extend(stepCase1)
       steps.extend(stepCase5)

       return steps

def getStepOfferRoamingPostpaidFixOffer(offerName, PPName, preloadBonus, cls, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, validity, timeband, quota):
       steps                = []
       stepConsumePreload   = None
       start_hour, end_hour = map(int, timeband.split('-'))
       
       allowanceSplit       = quota.split(';')
       allowanceVoice       = int(allowanceSplit[0])
       QuotaVoice           = int(allowanceVoice) if allowanceVoice != 0 else 0
       firstQuotaVoice      = QuotaVoice
       allowanceSMS         = 0
       if len(allowanceSplit) > 1:
              allowanceSMS  = int(allowanceSplit[1])
       QuotaSMS             = allowanceSMS if allowanceSMS != 0 else 0
       firstQuotaSMS        = QuotaSMS

       if preloadBonus != '' and preloadBonus != 0 and preloadBonus != "0":
              stepConsumePreload   = ["Consume Bonus Preload","Consume Bonus","No Bonus"]
              preloadBonusString = preloadBonus
       else:
              preloadBonusString = "No Bonus"

       stringBonusAll       = ''
       bonusVoice           = 'No Bonus'
       bonusSMS             = 'No Bonus'
       if firstQuotaVoice > 0:
              stringBonusAll = str(firstQuotaVoice)+" Min "+bonusDesc
              bonusVoice     = str(firstQuotaVoice)+" Min "+bonusDesc
       if firstQuotaSMS > 0:
              stringBonusAll = stringBonusAll+" "+str(firstQuotaSMS)+" SMS "+bonusDesc
              bonusSMS       = str(firstQuotaSMS)+" SMS "+bonusDesc

       if QuotaVoice > 0 or QuotaSMS > 0:
              stepsConsumeBonus, QuotaVoice, QuotaSMS = getStepReduceQuotaInternational(QuotaVoice, QuotaSMS, bonusDesc, start_hour, end_hour, validity, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif)
              

       #Case 1 = Positif Case
       stepCase1 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
              stepConsumePreload,
              ["Update parameter Init activation date", "Updated", "No Bonus"],
              [f"Attach Offer New Credit Limit Service {cls} | 3669334", "Offer Attached", "No Bonus"],
              [f"Attach Offer New CLS Roaming {cls} | 3669354", "Offer Attached", "No Bonus"],
              ["Attach Offer International Roaming | 36327", "Offer Attached", "No Bonus"],
              [f"Attach offer {offerName}","Offer attached",stringBonusAll],
              ["Check Bonus 889*1","Bonus Checked","No Bonus"],
              ["Check Bonus 889*2","Bonus Checked",bonusVoice],
              ["Check Bonus 889*3","Bonus Checked",bonusSMS],
              ["Check Bonus 889*4","Bonus Checked","No Bonus"],
              #Reduce Allowance
       ]

       stepCase1.extend(stepsConsumeBonus)
       stepCase1.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 5 = Multiple Attach (6x)
       bonus6x = ""
       if firstQuotaVoice > 0:
              totalVoice    = firstQuotaVoice*6
              bonus6x       = str(totalVoice)+" Min "+bonusDesc
       
       if firstQuotaSMS > 0:
              totalSMS      = firstQuotaSMS*6
              if bonus6x != '':
                     bonus6x = str(bonus6x)+", "+str(totalSMS)+" SMS "+bonusDesc
              else:
                     bonus6x = str(totalSMS)+" SMS "+bonusDesc

       stepCase5 = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
              stepConsumePreload,
              [f"Attach offer {offerName}","Offer attached",stringBonusAll],
              [f"Attach offer {offerName}","Offer attached",stringBonusAll+" , "+stringBonusAll],
              [f"Attach offer {offerName}","Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [f"Attach offer {offerName}","Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [f"Attach offer {offerName}","Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              [f"Attach offer {offerName}","Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              ["Check 889","Checked",bonus6x],
              ["Check on database","Success",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
              ["Check PI on Indira","Success","No Bonus"],
              ["Do Offline event", "", ""]
       ]

       steps.extend(stepCase1)
       steps.extend(stepCase5)

       return steps

def getStepOfferRoamingPostpaidFlexibleOffer(offerName, PPName, cls, eligible, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, startDateValidity, allowance, scenarioChoosen, RCIndicator, validityEndDate, endDateValidityBack):
       steps                = []
       stepConsumePreload   = None
       preloadBonusString   = "No Bonus"
       scenarioChoosed      = scenarioChoosen.split(',')
       validityEndDate      = validityEndDate.split(';')
       date_object          = datetime.strptime(endDateValidityBack, "%Y%m%d")
       date_with_time       = date_object.replace(hour=17, minute=0, second=0)
       endDateValidityBack  = date_with_time.strftime("%Y-%m-%d %H:%M:%S")
       
       allowanceSplit       = allowance.split(';')
       allowanceVoice       = int(allowanceSplit[0])
       QuotaVoice           = int(allowanceVoice/60) if allowanceVoice != 0 else 0
       QuotaVoiceNegatif    = 0
       QuotaVoiceCase2      = 0
       QuotaVoiceCase3      = 0
       firstQuotaVoice      = QuotaVoice
       allowanceSMS         = 0
       if len(allowanceSplit) > 1:
              allowanceSMS  = int(allowanceSplit[1])
       QuotaSMS             = allowanceSMS if allowanceSMS != 0 else 0
       QuotaSMSCase2        = 0
       QuotaSMSNegatif      = 0
       QuotaSMSCase3        = 0
       firstQuotaSMS        = QuotaSMS

       if eligible == 'Voice':
              UOM = 'V'
              attachOfferString = f'Attach Offer {offerName} Quota|UOM|Validity end date|RC indicator|Invoice description|Quotation reference|External product id|TransactionID {allowanceVoice}|{UOM}|variable_for_historycal|variable_for_RC_Indicator|String|String|String|String'
              attachOfferStringCase2 = f'Attach Offer {offerName} Quota|UOM|Validity end date|RC indicator|Invoice description|Quotation reference|External product id|TransactionID {allowanceVoice}|S|1960-01-01 19:50:00|0|String|String|String|String'
              attachOfferStringCase3 = f'Attach Offer {offerName} Quota|UOM|Validity end date|RC indicator|Invoice description|Quotation reference|External product id|TransactionID {allowanceVoice}|{UOM}|{endDateValidityBack}|variable_for_RC_Indicator|String|String|String|String'
       elif eligible == 'SMS':
              UOM = 'S'
              attachOfferString = f'Attach Offer {offerName} Quota|UOM|Validity end date|RC indicator|Invoice description|Quotation reference|External product id|TransactionID {allowanceSMS}|{UOM}|variable_for_historycal|variable_for_RC_Indicator|String|String|String|String'
              attachOfferStringCase2 = f'Attach Offer {offerName} Quota|UOM|Validity end date|RC indicator|Invoice description|Quotation reference|External product id|TransactionID {allowanceSMS}|V|1960-01-01 19:50:00|0|String|String|String|String'
              attachOfferStringCase3 = f'Attach Offer {offerName} Quota|UOM|Validity end date|RC indicator|Invoice description|Quotation reference|External product id|TransactionID {allowanceVoice}|{UOM}|{endDateValidityBack}|0|String|String|String|String'
       elif eligible == 'Voice & SMS':
              UOMV = 'V'
              UOMS = 'S'
              attachOfferString = f'Attach Offer {offerName} Quota|UOM|Validity end date|RC indicator|Invoice description|Quotation reference|External product id|TransactionID {allowanceVoice}|{UOMV}|variable_for_historycal|variable_for_RC_Indicator|String|String|String|String;{allowanceSMS}|{UOMS}|1960-01-01 19:50:00|variable_for_RC_Indicator|String|String|String|String'
              attachOfferStringCase2 = f'Attach Offer {offerName} Quota|UOM|Validity end date|RC indicator|Invoice description|Quotation reference|External product id|TransactionID {allowanceVoice}|O|1960-01-01 19:50:00|0|String|String|String|String;{allowanceSMS}|O|1960-01-01 19:50:00|variable_for_RC_Indicator|String|String|String|String'
              attachOfferStringCase3 = f'Attach Offer {offerName} Quota|UOM|Validity end date|RC indicator|Invoice description|Quotation reference|External product id|TransactionID {allowanceVoice}|{UOMV}|{endDateValidityBack}|0|String|String|String|String;{allowanceSMS}|{UOMS}|{endDateValidityBack}|variable_for_RC_Indicator|String|String|String|String'


       stringBonusAll       = ''
       bonusVoice           = 'No Bonus'
       bonusSMS             = 'No Bonus'
       if firstQuotaVoice > 0:
              stringBonusAll = str(firstQuotaVoice)+" Min "+bonusDesc
              bonusVoice     = str(firstQuotaVoice)+" Min "+bonusDesc
       if firstQuotaSMS > 0:
              stringBonusAll = stringBonusAll+" "+str(firstQuotaSMS)+" SMS "+bonusDesc
              bonusSMS       = str(firstQuotaSMS)+" SMS "+bonusDesc
       
       #Case 1 = Negatif Attach offer without parameter
       if "1" in scenarioChoosed or "11" in scenarioChoosed:
              stepsConsumeBonusCase1 = getStepReduceQuotaInternationalFlexiblePostpaid(QuotaVoiceNegatif, QuotaSMSNegatif, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, 0, startDateValidity, "1960-01-01 19:50:00")
              stepCase1 = [
                     ["Case Attach offer without parameter", "", "", "", ""],
                     [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
                     stepConsumePreload,
                     [f"Attach Offer New Credit Limit Service {cls} | 3669334", "Offer Attached", "No Bonus"],
                     [f"Attach Offer New CLS Roaming {cls} | 3669354", "Offer Attached", "No Bonus"],
                     ["Attach Offer International Roaming | 36327", "Offer Attached", "No Bonus"],
                     [f"Attach Offer {offerName}","Rejected","No Bonus"],
                     #Reduce Allowance
              ]
              stepCase1.extend(stepsConsumeBonusCase1)
              stepCase1.extend([["Check PI on Indira","Success","No Bonus"]])
       
       #Case 2 = Negatif (Berdasarkan UOM)
       if "2" in scenarioChoosed or "11" in scenarioChoosed:
              stepsConsumeBonusCase2 = getStepReduceQuotaInternationalFlexiblePostpaid(QuotaVoiceCase2, QuotaSMSCase2, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, 0, startDateValidity, "1960-01-01 19:50:00")
              stepCase2 = [
                     ["Case Negatif (Berdasarkan UOM)", "", "", "", ""],
                     [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
                     stepConsumePreload,
                     [f"Attach Offer New Credit Limit Service {cls} | 3669334", "Offer Attached", "No Bonus"],
                     [f"Attach Offer New CLS Roaming {cls} | 3669354", "Offer Attached", "No Bonus"],
                     ["Attach Offer International Roaming | 36327", "Offer Attached", "No Bonus"],
                     [attachOfferStringCase2,"Offer attached","No Bonus"],
                     #Reduce Allowance
              ]
              stepCase2.extend(stepsConsumeBonusCase2)
              stepCase2.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 3 = Negatif (Backdate)
       if "3" in scenarioChoosed or "11" in scenarioChoosed:
              stepsConsumeBonusCase3 = getStepReduceQuotaInternationalFlexiblePostpaid(QuotaVoiceCase3, QuotaSMSCase3, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, 0, startDateValidity, "1960-01-01 19:50:00")
              stepCase3 = [
                     ["Case Backdate", "", "", "", ""],
                     [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
                     stepConsumePreload,
                     [f"Attach Offer New Credit Limit Service {cls} | 3669334", "Offer Attached", "No Bonus"],
                     [f"Attach Offer New CLS Roaming {cls} | 3669354", "Offer Attached", "No Bonus"],
                     ["Attach Offer International Roaming | 36327", "Offer Attached", "No Bonus"],
                     [attachOfferStringCase3,"Offer attached","No Bonus"],
                     #Reduce Allowance
              ]
              stepCase3.extend(stepsConsumeBonusCase3)
              stepCase3.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 4 = Positif Case RC Indicator -1 (Historical)
       if "4" in scenarioChoosed or "11" in scenarioChoosed:
              if firstQuotaVoice > 0 or firstQuotaSMS > 0:
                     stepsConsumeBonusCase4 = getStepReduceQuotaInternationalFlexiblePostpaid(QuotaVoice, QuotaSMS, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, -1, startDateValidity, "1960-01-01 19:50:00")
              stepCase4 = [
                     ["Case RC Indicator -1 (Historical)", "", "", "", ""],
                     [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
                     stepConsumePreload,
                     [f"Attach Offer New Credit Limit Service {cls} | 3669334", "Offer Attached", "No Bonus"],
                     [f"Attach Offer New CLS Roaming {cls} | 3669354", "Offer Attached", "No Bonus"],
                     ["Attach Offer International Roaming | 36327", "Offer Attached", "No Bonus"],
                     [attachOfferString.replace("variable_for_RC_Indicator", "-1").replace("variable_for_historycal", "1960-01-01 19:50:00"),"Offer attached",stringBonusAll],
                     ["Check Bonus 889*1","Bonus Checked","No Bonus"],
                     ["Check Bonus 889*2","Bonus Checked",bonusVoice],
                     ["Check Bonus 889*3","Bonus Checked",bonusSMS],
                     ["Check Bonus 889*4","Bonus Checked","No Bonus"],
                     #Reduce Allowance
              ]

              stepCase4.extend(stepsConsumeBonusCase4)
              stepCase4.extend([["Check PI on Indira","Success","No Bonus"]])
       
       #Case 5 = Positif Case RC Indicator -1 (Non Historical)
       if "5" in scenarioChoosed or "11" in scenarioChoosed:
              if firstQuotaVoice > 0 or firstQuotaSMS > 0:
                     stepsConsumeBonusCase5 = getStepReduceQuotaInternationalFlexiblePostpaid(QuotaVoice, QuotaSMS, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, -1, startDateValidity, validityEndDate[0])
              stepCase5 = [
                     ["Case RC Indicator -1 (Non Historical)", "", "", "", ""],
                     [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
                     stepConsumePreload,
                     [f"Attach Offer New Credit Limit Service {cls} | 3669334", "Offer Attached", "No Bonus"],
                     [f"Attach Offer New CLS Roaming {cls} | 3669354", "Offer Attached", "No Bonus"],
                     ["Attach Offer International Roaming | 36327", "Offer Attached", "No Bonus"],
                     [attachOfferString.replace("variable_for_RC_Indicator", "-1").replace("variable_for_historycal", validityEndDate[0]),"Offer attached",stringBonusAll],
                     #Reduce Allowance
              ]

              stepCase5.extend(stepsConsumeBonusCase5)
              stepCase5.extend([["Check PI on Indira","Success","No Bonus"]])
       
       #Case 6 = Positif Case RC Indicator 0 (Historical)
       if "6" in scenarioChoosed or "11" in scenarioChoosed:
              if firstQuotaVoice > 0 or firstQuotaSMS > 0:
                     stepsConsumeBonusCase6 = getStepReduceQuotaInternationalFlexiblePostpaid(QuotaVoice, QuotaSMS, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, 0, startDateValidity, "1960-01-01 19:50:00")
              stepCase6 = [
                     ["Case RC Indicator 0 (Historical)", "", "", "", ""],
                     [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
                     stepConsumePreload,
                     [f"Attach Offer New Credit Limit Service {cls} | 3669334", "Offer Attached", "No Bonus"],
                     [f"Attach Offer New CLS Roaming {cls} | 3669354", "Offer Attached", "No Bonus"],
                     ["Attach Offer International Roaming | 36327", "Offer Attached", "No Bonus"],
                     [attachOfferString.replace("variable_for_RC_Indicator", "0").replace("variable_for_historycal", "1960-01-01 19:50:00"),"Offer attached",stringBonusAll],
                     #Reduce Allowance
              ]

              stepCase6.extend(stepsConsumeBonusCase6)
              stepCase6.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 7 = Positif Case RC Indicator 0 (Non Historycal)
       if "7" in scenarioChoosed or "11" in scenarioChoosed:
              if firstQuotaVoice > 0 or firstQuotaSMS > 0:
                     stepsConsumeBonusCase7 = getStepReduceQuotaInternationalFlexiblePostpaid(QuotaVoice, QuotaSMS, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, 0, startDateValidity, validityEndDate[1])
              stepCase7 = [
                     ["Case RC Indicator 0 (Non Historical)", "", "", "", ""],
                     [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
                     stepConsumePreload,
                     [f"Attach Offer New Credit Limit Service {cls} | 3669334", "Offer Attached", "No Bonus"],
                     [f"Attach Offer New CLS Roaming {cls} | 3669354", "Offer Attached", "No Bonus"],
                     ["Attach Offer International Roaming | 36327", "Offer Attached", "No Bonus"],
                     [attachOfferString.replace("variable_for_RC_Indicator", "0").replace("variable_for_historycal", validityEndDate[1]),"Offer attached",stringBonusAll],
                     #Reduce Allowance
              ]

              stepCase7.extend(stepsConsumeBonusCase7)
              stepCase7.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 8 = Positif Case RC Indicator more than 0 (Historical)
       if "8" in scenarioChoosed or "11" in scenarioChoosed:
              if firstQuotaVoice > 0 or firstQuotaSMS > 0:
                     stepsConsumeBonusCase8 = getStepReduceQuotaInternationalFlexiblePostpaid(QuotaVoice, QuotaSMS, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, RCIndicator, startDateValidity, "1960-01-01 19:50:00")
              stepCase8 = [
                     ["Case RC Indicator More Than 0 (Historical)", "", "", "", ""],
                     [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
                     stepConsumePreload,
                     [f"Attach Offer New Credit Limit Service {cls} | 3669334", "Offer Attached", "No Bonus"],
                     [f"Attach Offer New CLS Roaming {cls} | 3669354", "Offer Attached", "No Bonus"],
                     ["Attach Offer International Roaming | 36327", "Offer Attached", "No Bonus"],
                     [attachOfferString.replace("variable_for_RC_Indicator", RCIndicator).replace("variable_for_historycal", "1960-01-01 19:50:00"),"Offer attached",stringBonusAll],
                     #Reduce Allowance
              ]

              stepCase8.extend(stepsConsumeBonusCase8)
              stepCase8.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 9 = Positif Case RC Indicator more than 0 (Non Historycal)
       if "9" in scenarioChoosed or "11" in scenarioChoosed:
              if firstQuotaVoice > 0 or firstQuotaSMS > 0:
                     stepsConsumeBonusCase9 = getStepReduceQuotaInternationalFlexiblePostpaid(QuotaVoice, QuotaSMS, bonusDesc, MOEligible, MTEligible, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, RCIndicator, startDateValidity, validityEndDate[2])
              stepCase9 = [
                     ["Case RC Indicator More Than 0 (Non Historical)", "", "", "", ""],
                     [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
                     stepConsumePreload,
                     [f"Attach Offer New Credit Limit Service {cls} | 3669334", "Offer Attached", "No Bonus"],
                     [f"Attach Offer New CLS Roaming {cls} | 3669354", "Offer Attached", "No Bonus"],
                     ["Attach Offer International Roaming | 36327", "Offer Attached", "No Bonus"],
                     [attachOfferString.replace("variable_for_RC_Indicator", RCIndicator).replace("variable_for_historycal", validityEndDate[2]),"Offer attached",stringBonusAll],
                     #Reduce Allowance
              ]

              stepCase9.extend(stepsConsumeBonusCase9)
              stepCase9.extend([["Check PI on Indira","Success","No Bonus"]])

       #Case 10 = Multiple Attach (6x)
       if "10" in scenarioChoosed or "11" in scenarioChoosed:
              bonus6x = ""
              if firstQuotaVoice > 0:
                     totalVoice    = firstQuotaVoice*6
                     bonus6x       = str(totalVoice)+" Min "+bonusDesc
              
              if firstQuotaSMS > 0:
                     totalSMS      = firstQuotaSMS*6
                     if bonus6x != '':
                            bonus6x = str(bonus6x)+", "+str(totalSMS)+" SMS "+bonusDesc
                     else:
                            bonus6x = str(totalSMS)+" SMS "+bonusDesc

              stepCase10 = [
                     ["Case Multiple Attach (6x)", "", "", "", ""],
                     [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
                     stepConsumePreload,
                     [f"Attach Offer New Credit Limit Service {cls} | 3669334", "Offer Attached", "No Bonus"],
                     [f"Attach Offer New CLS Roaming {cls} | 3669354", "Offer Attached", "No Bonus"],
                     ["Attach Offer International Roaming | 36327", "Offer Attached", "No Bonus"],
                     [attachOfferString.replace("variable_for_RC_Indicator", "0").replace("variable_for_historycal", "1960-01-01 19:50:00"),"Offer attached",stringBonusAll],
                     [attachOfferString.replace("variable_for_RC_Indicator", "0").replace("variable_for_historycal", "1960-01-01 19:50:00"),"Offer attached",stringBonusAll+" , "+stringBonusAll],
                     [attachOfferString.replace("variable_for_RC_Indicator", "0").replace("variable_for_historycal", "1960-01-01 19:50:00"),"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
                     [attachOfferString.replace("variable_for_RC_Indicator", "0").replace("variable_for_historycal", "1960-01-01 19:50:00"),"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
                     [attachOfferString.replace("variable_for_RC_Indicator", "0").replace("variable_for_historycal", "1960-01-01 19:50:00"),"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
                     [attachOfferString.replace("variable_for_RC_Indicator", "0").replace("variable_for_historycal", "1960-01-01 19:50:00"),"Offer attached",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
                     ["Check 889","Checked",bonus6x],
                     ["Check on database","Success",stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll+" , "+stringBonusAll],
                     ["Check PI on Indira","Success","No Bonus"],
                     ["Do Offline event", "", ""]
              ]

       #Case BSZ
       stepCaseBSZ = [
              [f"{offerName} SCN For check BSZ Extract | D-2", "", "", "", ""],
              [f"Create & Activate new subscriber PP {PPName}", "Success", preloadBonusString],
              stepConsumePreload,
              [f"Attach Offer New Credit Limit Service {cls} | 3669334", "Offer Attached", "No Bonus"],
              [f"Attach Offer New CLS Roaming {cls} | 3669354", "Offer Attached", "No Bonus"],
              ["Attach Offer International Roaming | 36327", "Offer Attached", "No Bonus"],
              [attachOfferString, "Offer attached", stringBonusAll],
              ["Check notification after add offer", "Success", "Checked"],
              ["Check Offer Name & Description", "Success", offerName],
              ["Check GetBonusInfo and validity", "Success", "Checked"],
              ["Check Bonus 889 and bonus description", "Success", "Checked"],
              ["Check PRIT Name", "Success", "Success"],
              ["Create event vas with eligible vascode param_vascode", "Success", "Success"],
              ["Check notification after first event consume", "Success", "Success"],
              ["Create event voice Onnet 60s", "Success", "Success"],
              ["run adjustment so it will expired by today -2", "Success", "No Bonus"],
              ["check bonus info (bonus should be gone)", "Success", "No Bonus"],
              ["check 888", "Success", "No Bonus"],
              ["run BSZ eod, and check bsz seizure", "Success", "No Bonus"],
       ]

       if "1" in scenarioChoosed or "11" in scenarioChoosed:
              steps.extend(stepCase1)

       if "2" in scenarioChoosed or "11" in scenarioChoosed:
              steps.extend(stepCase2)
       
       if "3" in scenarioChoosed or "11" in scenarioChoosed:
              steps.extend(stepCase3)
       
       if "4" in scenarioChoosed or "11" in scenarioChoosed:
              steps.extend(stepCase4)
       
       if "5" in scenarioChoosed or "11" in scenarioChoosed:
              steps.extend(stepCase5)
       
       if "6" in scenarioChoosed or "11" in scenarioChoosed:
              steps.extend(stepCase6)
       
       if "7" in scenarioChoosed or "11" in scenarioChoosed:
              steps.extend(stepCase7)
       
       if "8" in scenarioChoosed or "11" in scenarioChoosed:
              steps.extend(stepCase8)
       
       if "9" in scenarioChoosed or "11" in scenarioChoosed:
              steps.extend(stepCase9)
       
       if "10" in scenarioChoosed or "11" in scenarioChoosed:
              steps.extend(stepCase10)

       return steps

def getStepReduceQuotaInternational(QuotaVoice, QuotaSMS, bonDesc, start_hour, end_hour, validity, MO, MT, vascodePositif, vascodeNegatif, countryPositif, countryNegatif):
       stepsConsume         = []
       dayString            = 0
       validity             = int(validity)
       maxValidity          = validity+1
       countryPositifSplit  = countryPositif.split(';')
       firstCountryPos      = countryPositifSplit[0]
       countryPositifData   = [{"name": name, "status": "Positif"} for name in countryPositifSplit]
       countryNegatifSplit  = countryNegatif.split(';') 
       countryNegatifData   = [{"name": name} for name in countryNegatifSplit] 
       mergedCountryData    = countryPositifData + countryNegatifData
       vascodePosSplit      = vascodePositif.split(";")
       firstVascodePos      = vascodePosSplit[0]
       vascodePosData       = [{"name": name, "status": "Positif"} for name in vascodePosSplit]
       vascodeNegSplit      = vascodeNegatif.split(";")
       vascodeNegData       = [{"name": name, "status": "Negatif"} for name in vascodeNegSplit]
       mergedVascode        = vascodePosData + vascodeNegData
       MOEligible           = MO
       MTEligible           = MT
       
       # Generate a shuffled list of numbers from dayString to validity - 1
       days          = list(range(dayString, maxValidity))
       firstDate     = days[0]
       lastDate      = days[len(days) - 1]

       if validity > 1:
              # Choose a random number of elements to select from data
              num_elements = random.randint(1, len(days) - 2)

              # Randomly select elements from data
              selected_data = random.sample(days[1:-1], num_elements)

              # Sort selected_data based on the index in the original data list
              selected_data = sorted(selected_data, key=lambda x: days.index(x))

              # Merge variables into a single list
              merged_data = [firstDate] + selected_data + [lastDate]
       else:
              merged_data = [firstDate] + [lastDate]
       
       random.shuffle(mergedVascode)
       random.shuffle(mergedCountryData)
       additionalNegatifCase = [
              "Create Voice Onnet 1 Min",
              "Create event 1 SMS onnet",
              "Create event GPRS 1MB RG 50",
              "Create Event Voice Offnet 5s",
              "Create Event Voice PSTN 5s",
              "Create Event Voice FWA 180s",
       ]
       #Steps for reduce quota voice
       getDataVoice                = 0
       countVoice                  = 1
       priorityOutVoice            = 0
       negatifCaseOutTimeband      = True
       MO_MT_Data                  = [
              {"Type" : "MO", "Data" : "Home"},
              {"Type" : "MO", "Data" : "Local"},
              {"Type" : "MO", "Data" : "Other"},
              {"Type" : "MT", "Data" : "Home"},
              {"Type" : "MT", "Data" : "Local"},
              {"Type" : "MT", "Data" : "Other"}
       ]
       data_MO_MT           = 0
       for strValidity in merged_data:
              stepsConsumeVoice, QuotaVoice, getDataVoice, countVoice, priorityOutVoice, data_MO_MT, additionalNegatifCase, negatifCaseOutTimeband = validateStepNormalVoiceInternational(QuotaVoice, QuotaSMS, strValidity, merged_data, mergedCountryData, start_hour, end_hour, validity, bonDesc, firstCountryPos, countryPositifData, MOEligible, MTEligible, getDataVoice, countVoice, priorityOutVoice, countryNegatifData, MO_MT_Data, data_MO_MT, additionalNegatifCase, negatifCaseOutTimeband)
              stepsConsume.extend(stepsConsumeVoice)
              stepsConsumeSMS, QuotaSMS, additionalNegatifCase, negatifCaseOutTimeband = validateStepNormalSMSInternational(QuotaVoice, QuotaSMS, strValidity, mergedVascode, start_hour, end_hour, validity, bonDesc, additionalNegatifCase, negatifCaseOutTimeband)
              stepsConsume.extend(stepsConsumeSMS)

       return stepsConsume, QuotaVoice, QuotaSMS

def validateStepNormalVoiceInternational(QuotaVoice, QuotaSMS, day, merged_data, countryData, start_hour, end_hour, validity, bonDesc, firstCountryPos, countryPosData, MOEligible, MTEligible, getDataVoice, countVoice, priorityOutVoice, countryNegatifData, MO_MT_Data, MO_MT_Out, additionalNegatifCase, negatifCaseOutTimeband):
       stepsConsume         = []

       getData       = getDataVoice
       count         = countVoice

       #Variable for count how much country positive and access code positive is out
       priorityOut   = priorityOutVoice

       if MO_MT_Out >= len(MO_MT_Data):
              MO_MT_Out = 0
       
       MO_or_MT = MO_MT_Data[MO_MT_Out]["Type"]

       if priorityOut >= len(countryPosData):
              if getData >= len(countryData):
                     getData = 0
              country       = countryData[getData]
              countryName   = country["name"]
       else:
              country       = countryPosData[priorityOut]
              countryName   = country["name"]

              priorityOut += 1
       
       showOrNotOutTimeband = random.randint(0,1)
       if showOrNotOutTimeband == 1 and negatifCaseOutTimeband and end_hour <= 22:
              start_hour_out_timeband     = int(end_hour)+1
              timeNumber                  = random.randint(start_hour_out_timeband, 23)
              timeString                  = timeNumber
              negatifCaseOutTimeband      = False
       else:
              timeNumber    = random.randint(start_hour, end_hour)
              timeString    = timeNumber

       if MO_or_MT == 'MO':
              MO_MT_Choice = MO_MT_Data[MO_MT_Out]["Data"]
              MO_ELigible = MOEligible.lower().split(";")
              if MO_MT_Choice.lower() in MO_ELigible:
                     checked_MO_MT = True
              else:
                     checked_MO_MT = False
       else:
              MO_MT_Choice = MO_MT_Data[MO_MT_Out]["Data"]
              MT_ELigible = MTEligible.lower().split(";")
              if MO_MT_Choice.lower() in MT_ELigible:
                     checked_MO_MT = True
              else:
                     checked_MO_MT = False

       if MO_MT_Choice == 'Home':
              String_MO_MT = f'From {countryName}'
       elif MO_MT_Choice == 'Local':
              String_MO_MT = f'In {countryName}'
       else:
              countryNeg    = random.choice(countryNegatifData)
              countryTo     = countryNeg["name"]
              if MO_or_MT == 'MO':
                     String_MO_MT  = f'In {countryName} To {countryTo}'
              else:
                     String_MO_MT  = f'In {countryName} From {countryTo}'
       
       if checked_MO_MT:
              if "status" in country:
                     if day < validity:
                            if start_hour <= end_hour:
                                   # Time range does not span midnight
                                   if start_hour <= timeNumber <= end_hour:
                                          # Number is within the time range Timeband
                                          consumeOrCharged     = 'Consume Bonus'
                                          reduceOrNot          = True
                                   else:
                                          # Number is not within the time range Timeband
                                          consumeOrCharged     = 'Charged'
                                          reduceOrNot          = False
                            else:
                                   # Time range spans midnight
                                   if timeNumber >= start_hour or timeNumber <= end_hour:
                                          # Number is within the time range Timeband
                                          consumeOrCharged     = 'Consume Bonus'
                                          reduceOrNot          = True
                                   else:
                                          # Number is not within the time range Timeband
                                          consumeOrCharged     = 'Charged'
                                          reduceOrNot          = False
                     else:
                            consumeOrCharged     = 'Charged'
                            reduceOrNot          = False
              else:
                     consumeOrCharged     = 'Charged'
                     reduceOrNot          = False
       else:
              consumeOrCharged     = 'Charged'
              reduceOrNot          = False
       
       if QuotaVoice > 0 and reduceOrNot:
              decreasingQuotaVoice = round((QuotaVoice * 0.5) / 4)
              QuotaVoice -= decreasingQuotaVoice
              eventString = decreasingQuotaVoice
       else:
              eventString          = '1'
              consumeOrCharged     = 'Charged'

       if int(timeString) > 12:
              timeString = str(int(timeString) - 12) + 'PM'
       else:
              timeString = str(timeString) + "AM"
       
       QuotaVoice    = int(QuotaVoice)
       if QuotaVoice > 0 and QuotaSMS <= 0:
              restBonus = f"{QuotaVoice} Min {bonDesc}"
       elif QuotaVoice > 0 and QuotaSMS > 0:
              restBonus = f"{QuotaVoice} Min {bonDesc}, {QuotaSMS} SMS {bonDesc}"
       elif QuotaVoice <= 0 and QuotaSMS > 0:
              restBonus = f"{QuotaSMS} SMS {bonDesc}"
       elif QuotaVoice <= 0 and QuotaSMS <= 0:
              restBonus = "No Bonus" 

       if int(day) >= validity:
              restBonus = "No Bonus"

       #{Create} {event} {voice} {MO/MT Home/Local/Other} {Home = from, Local = in, Other = Country 1 to Country 2} {timeString} {eventString}Min D+{day}
       #Create event voice {MO_or_MT} {MO_MT_Choice} {String_MO_MT} {timeString} {EventString}Min D+{day}
       # eventLabel = f"Create event voice IDD to {countryName} using access code {vascodeUsed} {timeString} {eventString}Min D+{day}"
       eventLabel = f"Create event voice {MO_or_MT} {MO_MT_Choice} {String_MO_MT} {timeString} {eventString}Min D+{day}"

       step = [
              eventLabel,
              consumeOrCharged,
              restBonus
       ]

       stepsConsume.append(step)

       if len(additionalNegatifCase) > 0:
              showOrNot = random.randint(0,1)
              if showOrNot == 1:
                     selected_value = random.choice(additionalNegatifCase)
                     stepAdd = [
                            str(selected_value)+" "+str(timeString)+" D+"+str(day),
                            "Charged",
                            restBonus
                     ]
                     stepsConsume.append(stepAdd)
                     additionalNegatifCase.remove(selected_value)
       
       if count == len(merged_data):
              stepLast = [
                     f"Create event voice {MO_or_MT} {MO_MT_Choice} {firstCountryPos} {timeString} {eventString}Min D+{day}",
                     "Charged",
                     restBonus
              ]
              stepsConsume.append(stepLast)
                     
       getData += 1
       count += 1
       MO_MT_Out += 1

       return stepsConsume, QuotaVoice, getData, count, priorityOut, MO_MT_Out, additionalNegatifCase, negatifCaseOutTimeband

def validateStepNormalSMSInternational(QuotaVoice, QuotaSMS, day, mergedVascode, start_hour, end_hour, validity, bonDesc, additionalNegatifCase, negatifCaseOutTimeband):
       stepsConsume         = []

       vascode           = random.choice(mergedVascode)
       vascodeName       = vascode["name"]
       vascodeStatus     = vascode["status"]
       vascodeUsed       = vascodeName
       
       showOrNotOutTimeband = random.randint(0,1)
       if showOrNotOutTimeband == 1 and negatifCaseOutTimeband and end_hour <= 22:
              start_hour_out_timeband     = int(end_hour)+1
              timeNumber                  = random.randint(start_hour_out_timeband, 23)
              timeString                  = timeNumber
              negatifCaseOutTimeband      = False
       else:
              timeNumber    = random.randint(start_hour, end_hour)
              timeString    = timeNumber

       if vascodeStatus == 'Positif':
              if day < validity:
                     if start_hour <= end_hour:
                            # Time range does not span midnight
                            if start_hour <= timeNumber <= end_hour:
                                   # Number is within the time range Timeband
                                   consumeOrCharged     = 'Consume Bonus'
                                   reduceOrNot          = True
                            else:
                                   # Number is not within the time range Timeband
                                   consumeOrCharged     = 'Charged'
                                   reduceOrNot          = False
                     else:
                            # Time range spans midnight
                            if timeNumber >= start_hour or timeNumber <= end_hour:
                                   # Number is within the time range Timeband
                                   consumeOrCharged     = 'Consume Bonus'
                                   reduceOrNot          = True
                            else:
                                   # Number is not within the time range Timeband
                                   consumeOrCharged     = 'Charged'
                                   reduceOrNot          = False
              else:
                     consumeOrCharged     = 'Charged'
                     reduceOrNot          = False
       else:
              consumeOrCharged     = 'Charged'
              reduceOrNot          = False
       
       if QuotaSMS > 0 and reduceOrNot:
              decreasingQuotaSMS = 1
              QuotaSMS -= decreasingQuotaSMS
       else:
              consumeOrCharged     = 'Charged'

       if int(timeString) > 12:
              timeString = str(int(timeString) - 12) + 'PM'
       else:
              timeString = str(timeString) + "AM"
       
       QuotaSMS    = int(QuotaSMS)
       if QuotaVoice > 0 and QuotaSMS <= 0:
              restBonus = f"{QuotaVoice} Min {bonDesc}"
       elif QuotaVoice > 0 and QuotaSMS > 0:
              restBonus = f"{QuotaVoice} Min {bonDesc}, {QuotaSMS} SMS {bonDesc}"
       elif QuotaVoice <= 0 and QuotaSMS > 0:
              restBonus = f"{QuotaSMS} SMS {bonDesc}"
       elif QuotaVoice <= 0 and QuotaSMS <= 0:
              restBonus = "No Bonus" 

       if int(day) >= validity:
              restBonus = "No Bonus"

       eventLabel = f"Create event direct debit with vascode {vascodeUsed} {timeString} D+{day}"

       step = [
              eventLabel,
              consumeOrCharged,
              restBonus
       ]

       stepsConsume.append(step)

       if len(additionalNegatifCase) > 0:
              showOrNot = random.randint(0,1)
              if showOrNot == 1:
                     selected_value = random.choice(additionalNegatifCase)
                     stepAdd = [
                            str(selected_value)+" "+str(timeString)+" D+"+str(day),
                            "Charged",
                            restBonus
                     ]
                     stepsConsume.append(stepAdd)
                     additionalNegatifCase.remove(selected_value)
       
       if day == validity:
              stepLast = [
                     f"Create event direct debit with vascode {vascodeUsed} {timeString} D+{day}",
                     "Charged",
                     restBonus
              ]
              stepsConsume.append(stepLast)

       return stepsConsume, QuotaSMS, additionalNegatifCase, negatifCaseOutTimeband

def getStepReduceQuotaInternationalFlexiblePostpaid(QuotaVoice, QuotaSMS, bonDesc, MO, MT, vascodePositif, vascodeNegatif, countryPositif, countryNegatif, RCIndicator, startDate, endDate):
       stepsConsume         = []
       countryPositifSplit  = countryPositif.split(';')
       countryPositifData   = [{"name": name, "status": "Positif"} for name in countryPositifSplit]
       countryNegatifSplit  = countryNegatif.split(';') 
       countryNegatifData   = [{"name": name} for name in countryNegatifSplit] 
       mergedCountryData    = countryPositifData + countryNegatifData
       vascodePosSplit      = vascodePositif.split(";")
       vascodePosData       = [{"name": name, "status": "Positif"} for name in vascodePosSplit]
       vascodeNegSplit      = vascodeNegatif.split(";")
       vascodeNegData       = [{"name": name, "status": "Negatif"} for name in vascodeNegSplit]
       mergedVascode        = vascodePosData + vascodeNegData
       MOEligible           = MO
       MTEligible           = MT
       firstQuotaVoice      = QuotaVoice
       firstQuotaSMS        = QuotaSMS
       RCIndicator          = int(RCIndicator)
       
       # Generate a shuffled list of numbers from dayString to validity - 1
       random.shuffle(mergedVascode)
       random.shuffle(mergedCountryData)
       additionalNegatifCase = [
              "Create Voice Onnet 1 Min",
              "Create event 1 SMS onnet",
              "Create event GPRS 1MB RG 50",
              "Create Event Voice Offnet 5s",
              "Create Event Voice PSTN 5s",
              "Create Event Voice FWA 180s",
       ]
       #Steps for reduce quota voice
       getDataVoice                = 0
       countVoice                  = 1
       priorityOutVoice            = 0
       MO_MT_Data                  = [
              {"Type" : "MO", "Data" : "Home"},
              {"Type" : "MO", "Data" : "Local"},
              {"Type" : "MO", "Data" : "Other"},
              {"Type" : "MT", "Data" : "Home"},
              {"Type" : "MT", "Data" : "Local"},
              {"Type" : "MT", "Data" : "Other"}
       ]
       data_MO_MT           = 0

       param                = 0
       count                = 0
       
       if endDate != '1960-01-01 19:50:00':
              start_date_str = startDate
              start_date = datetime.strptime(start_date_str, "%Y%m%d")

              # End date
              end_date_str = endDate
              end_date = datetime.strptime(end_date_str, "%Y%m%d")

              # Loop through dates with an additional parameter
              current_date = start_date
              while current_date <= end_date:

                     if RCIndicator == -1:
                            if current_date.day == 1:
                                   QuotaVoice    = firstQuotaVoice
                                   QuotaSMS      = firstQuotaSMS
                     elif RCIndicator >= 0:
                            if count >= RCIndicator:
                                   QuotaVoice    = firstQuotaVoice
                                   QuotaSMS      = firstQuotaSMS

                     stepsConsumeVoice, QuotaVoice, getDataVoice, countVoice, priorityOutVoice, data_MO_MT, additionalNegatifCase = validateStepNormalVoiceInternationalFlexiblePostpaid(QuotaVoice, QuotaSMS, current_date, mergedCountryData, bonDesc, countryPositifData, MOEligible, MTEligible, getDataVoice, countVoice, priorityOutVoice, countryNegatifData, MO_MT_Data, data_MO_MT, additionalNegatifCase)
                     stepsConsume.extend(stepsConsumeVoice)
                     
                     stepsConsumeSMS, QuotaSMS, additionalNegatifCase = validateStepNormalSMSInternationalFlexiblePostpaid(QuotaVoice, QuotaSMS, current_date, mergedVascode, bonDesc, additionalNegatifCase)
                     stepsConsume.extend(stepsConsumeSMS)
                     
                     monthBefore   = current_date.month
                     plusDay       = random.randint(1,10)
                     
                     if RCIndicator > 0:
                            if plusDay > RCIndicator:
                                   plusDay = int(RCIndicator)

                     current_date += timedelta(days=plusDay)

                     if monthBefore != current_date.month:
                            current_date.replace(day=1)

                     count += plusDay
       else:
              start_date_str = startDate
              current_date = datetime.strptime(start_date_str, "%Y%m%d")
              
              while True:

                     if RCIndicator == -1:
                            if current_date.day == 1:
                                   QuotaVoice    = firstQuotaVoice
                                   QuotaSMS      = firstQuotaSMS
                     elif RCIndicator > 0:
                            if count == RCIndicator:
                                   QuotaVoice    = firstQuotaVoice
                                   QuotaSMS      = firstQuotaSMS
                                   
                     stepsConsumeVoice, QuotaVoice, getDataVoice, countVoice, priorityOutVoice, data_MO_MT, additionalNegatifCase = validateStepNormalVoiceInternationalFlexiblePostpaid(QuotaVoice, QuotaSMS, current_date, mergedCountryData, bonDesc, countryPositifData, MOEligible, MTEligible, getDataVoice, countVoice, priorityOutVoice, countryNegatifData, MO_MT_Data, data_MO_MT, additionalNegatifCase)
                     stepsConsume.extend(stepsConsumeVoice)
                     
                     stepsConsumeSMS, QuotaSMS, additionalNegatifCase = validateStepNormalSMSInternationalFlexiblePostpaid(QuotaVoice, QuotaSMS, current_date, mergedVascode, bonDesc, additionalNegatifCase)
                     stepsConsume.extend(stepsConsumeSMS)

                     monthBefore   = current_date.month
                     plusDay       = random.randint(1,10)

                     if RCIndicator > 0:
                            if plusDay > RCIndicator:
                                   plusDay = RCIndicator

                     current_date += timedelta(days=plusDay)

                     if monthBefore != current_date.month:
                            current_date.replace(day=1)

                     if RCIndicator == -1 or RCIndicator > 0:
                            count += plusDay

                            if count >= RCIndicator and RCIndicator > 0:
                                   param += 1
                                   count = 0
                            else:
                                   if current_date.day == 1:
                                          param += 1
                     
                            # Check if the parameter reaches a certain value (e.g., 5) to break out of the loop
                            if param >= 2:
                                   break
                     else:
                            if QuotaVoice <= 0 and QuotaSMS <= 0:
                                   break
                     

       return stepsConsume

def validateStepNormalVoiceInternationalFlexiblePostpaid(QuotaVoice, QuotaSMS, day, countryData, bonDesc, countryPosData, MOEligible, MTEligible, getDataVoice, countVoice, priorityOutVoice, countryNegatifData, MO_MT_Data, MO_MT_Out, additionalNegatifCase):
       stepsConsume  = []
       day           = day.strftime("%d-%m-%Y")
       getData       = getDataVoice
       count         = countVoice

       #Variable for count how much country positive and access code positive is out
       priorityOut   = priorityOutVoice

       if MO_MT_Out >= len(MO_MT_Data):
              MO_MT_Out = 0
       
       MO_or_MT = MO_MT_Data[MO_MT_Out]["Type"]

       if priorityOut >= len(countryPosData):
              if getData >= len(countryData):
                     getData = 0
              country       = countryData[getData]
              countryName   = country["name"]
       else:
              country       = countryPosData[priorityOut]
              countryName   = country["name"]

              priorityOut += 1

       if MO_or_MT == 'MO':
              MO_MT_Choice = MO_MT_Data[MO_MT_Out]["Data"]
              MO_ELigible = MOEligible.lower().split(";")
              if MO_MT_Choice.lower() in MO_ELigible:
                     checked_MO_MT = True
              else:
                     checked_MO_MT = False
       else:
              MO_MT_Choice = MO_MT_Data[MO_MT_Out]["Data"]
              MT_ELigible = MTEligible.lower().split(";")
              if MO_MT_Choice.lower() in MT_ELigible:
                     checked_MO_MT = True
              else:
                     checked_MO_MT = False

       if MO_MT_Choice == 'Home':
              String_MO_MT = f'From {countryName}'
       elif MO_MT_Choice == 'Local':
              String_MO_MT = f'In {countryName}'
       else:
              countryNeg    = random.choice(countryNegatifData)
              countryTo     = countryNeg["name"]
              if MO_or_MT == 'MO':
                     String_MO_MT  = f'In {countryName} To {countryTo}'
              else:
                     String_MO_MT  = f'In {countryName} From {countryTo}'
       
       if checked_MO_MT:
              if "status" in country:       
                     consumeOrCharged     = 'Consume Bonus'
                     reduceOrNot          = True
              else:
                     consumeOrCharged     = 'Charged'
                     reduceOrNot          = False
       else:
              consumeOrCharged     = 'Charged'
              reduceOrNot          = False
       
       if QuotaVoice > 0 and reduceOrNot:
              decreasingQuotaVoice = round((QuotaVoice * 0.5) / 4)
              if decreasingQuotaVoice == 0:
                     decreasingQuotaVoice = QuotaVoice
              QuotaVoice -= decreasingQuotaVoice
              eventString = decreasingQuotaVoice
       else:
              eventString          = '1'
              consumeOrCharged     = 'Charged'

       
       QuotaVoice    = int(QuotaVoice)
       if QuotaVoice > 0 and QuotaSMS <= 0:
              restBonus = f"{QuotaVoice} Min {bonDesc}"
       elif QuotaVoice > 0 and QuotaSMS > 0:
              restBonus = f"{QuotaVoice} Min {bonDesc}, {QuotaSMS} SMS {bonDesc}"
       elif QuotaVoice <= 0 and QuotaSMS > 0:
              restBonus = f"{QuotaSMS} SMS {bonDesc}"
       elif QuotaVoice <= 0 and QuotaSMS <= 0:
              restBonus = "No Bonus" 

       #{Create} {event} {voice} {MO/MT Home/Local/Other} {Home = from, Local = in, Other = Country 1 to Country 2} {timeString} {eventString}Min D+{day}
       #Create event voice {MO_or_MT} {MO_MT_Choice} {String_MO_MT} {timeString} {EventString}Min D+{day}
       # eventLabel = f"Create event voice IDD to {countryName} using access code {vascodeUsed} {timeString} {eventString}Min D+{day}"
       eventLabel = f"Create event voice {MO_or_MT} {MO_MT_Choice} {String_MO_MT} {eventString}Min {day}"

       step = [
              eventLabel,
              consumeOrCharged,
              restBonus
       ]

       stepsConsume.append(step)

       if len(additionalNegatifCase) > 0:
              showOrNot = random.randint(0,1)
              if showOrNot == 1:
                     selected_value = random.choice(additionalNegatifCase)
                     stepAdd = [
                            str(selected_value)+" "+str(day),
                            "Charged",
                            restBonus
                     ]
                     stepsConsume.append(stepAdd)
                     additionalNegatifCase.remove(selected_value)
                     
       getData += 1
       count += 1
       MO_MT_Out += 1

       return stepsConsume, QuotaVoice, getData, count, priorityOut, MO_MT_Out, additionalNegatifCase

def validateStepNormalSMSInternationalFlexiblePostpaid(QuotaVoice, QuotaSMS, day, mergedVascode, bonDesc, additionalNegatifCase):
       stepsConsume         = []
       day                  = day.strftime("%d-%m-%Y")
       vascode           = random.choice(mergedVascode)
       vascodeName       = vascode["name"]
       vascodeStatus     = vascode["status"]
       vascodeUsed       = vascodeName

       if vascodeStatus == 'Positif':
              consumeOrCharged     = 'Consume Bonus'
              reduceOrNot          = True
       else:
              consumeOrCharged     = 'Charged'
              reduceOrNot          = False
       
       if QuotaSMS > 0 and reduceOrNot:
              decreasingQuotaSMS = 1
              QuotaSMS -= decreasingQuotaSMS
       else:
              consumeOrCharged     = 'Charged'
       
       QuotaSMS    = int(QuotaSMS)
       if QuotaVoice > 0 and QuotaSMS <= 0:
              restBonus = f"{QuotaVoice} Min {bonDesc}"
       elif QuotaVoice > 0 and QuotaSMS > 0:
              restBonus = f"{QuotaVoice} Min {bonDesc}, {QuotaSMS} SMS {bonDesc}"
       elif QuotaVoice <= 0 and QuotaSMS > 0:
              restBonus = f"{QuotaSMS} SMS {bonDesc}"
       elif QuotaVoice <= 0 and QuotaSMS <= 0:
              restBonus = "No Bonus" 

       eventLabel = f"Create event direct debit with vascode {vascodeUsed} {day}"

       step = [
              eventLabel,
              consumeOrCharged,
              restBonus
       ]

       stepsConsume.append(step)

       if len(additionalNegatifCase) > 0:
              showOrNot = random.randint(0,1)
              if showOrNot == 1:
                     selected_value = random.choice(additionalNegatifCase)
                     stepAdd = [
                            str(selected_value)+" "+str(day),
                            "Charged",
                            restBonus
                     ]
                     stepsConsume.append(stepAdd)
                     additionalNegatifCase.remove(selected_value)
       

       return stepsConsume, QuotaSMS, additionalNegatifCase

def exportExcelANPS(eventName, params=None, neededParams = None):
       wb = Workbook()
       ws = wb.active

       offerName            = ''
       PPName               = ''
       preloadBonus         = ''
       wordingAddOffer      = ''
       wordingReachTreshold = ''
       dropCallOnnet        = 'N'
       dropCallOffnet       = 'N'
       threshold            = ''

       for params in params:
              if "Offer Name" in params:
                     offerName = params["Offer Name"]
              
              if "Price Plan Name" in params:
                     PPName = params["Price Plan Name"]
              
              if "Preload Bonus" in params:
                     preloadBonus = params["Preload Bonus"]
              
              if "Wording add offer" in params:
                     wordingAddOffer = params["Wording add offer"]

              if "Wording reach treshold" in params:
                     wordingReachTreshold = params["Wording reach treshold"]
              
              if "Drop Call Onnet" in params:
                     dropCallOnnet = params["Drop Call Onnet"]
              
              if "Drop Call Offnet" in params:
                     dropCallOffnet = params["Drop Call Offnet"]
              
              if "Threshold" in params:
                     threshold = params["Threshold"]
              
              steps = stepANPS(offerName, PPName, preloadBonus, wordingAddOffer, wordingReachTreshold, dropCallOnnet, dropCallOffnet, threshold)
              

              # Write Header Row
              header = [f'{eventName} | {offerName}']
              ws.append(header)

              # Merge Header Cells
              startColumnIndex = 3  # Example of a dynamic column index
              startColumn = chr(ord("A") + startColumnIndex)  # Calculate the start column dynamically
              endColumn = "E"
              startRow = 1
              endRow = 1
              cellRange = f"{startColumn}{startRow}:{endColumn}{endRow}"
              ws.merge_cells(cellRange)

              headerRow = ['No.', 'Steps:', 'Validation (per step)',	'*889#', 'Result']
              ws.append(headerRow)

              no = 1
              for num, step in enumerate(steps, start=1):
                     if isinstance(step, str):
                            row = [
                                   no,
                                   step,
                                   "Success",
                                   "No Bonus",
                                   "XYZ"
                            ]
                            no = no+1
                     else:
                            if step is None:
                                   continue
                            else:
                                   if len(step) == 5:
                                          row = [
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 step[3],
                                                 step[4]
                                          ]
                                   elif len(step) == 4:
                                          row = [
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 step[3],
                                                 "XYZ"
                                          ]
                                   elif len(step) == 3:
                                          row = [
                                                 no,
                                                 step[0],
                                                 step[1],
                                                 step[2],
                                                 "XYZ"
                                          ]
                                          no = no+1
                                   else:
                                          row = [
                                                 no,
                                                 step[0],
                                                 step[1],
                                                 "No Bonus",
                                                 "XYZ"
                                          ]
                                          no = no+1
                     ws.append(row)

       print("Testing Scenario Successfully Generated")
       
       # Save Excel File
       wb.save('Result/Scenario '+str(eventName)+' '+str(offerName)+'.xlsx')

def stepANPS(offerName, PPName, preloadBonus, wordingAddOffer, wordingReachTreshold, dropCallOnnet, dropCallOffnet, threshold):
       stepConsumePreload   = None

       if preloadBonus != '' and preloadBonus != 0 and preloadBonus != "0":
              stepConsumePreload   = ["Consume Bonus Preload","Consume Bonus","No Bonus"]
              preloadBonusString = preloadBonus
       else:
              preloadBonusString = "No Bonus"
       
       if dropCallOnnet == "Y":
              dropOnnet = f"Charged {threshold} IDR"
       else:
              dropOnnet = f"Charged Not {threshold} IDR"
       
       if dropCallOffnet == "Y":
              dropOffnet = f"Charged {threshold} IDR"
       else:
              dropOffnet = f"Charged Not {threshold} IDR"

       steps = [
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
              ["Update Balance 1000K","Balance Updated",preloadBonusString],
              ["Update exp date","Updated",preloadBonusString],
              [stepConsumePreload],
              ["Create event 80s voice onnet 9am D+0 zone ID =  2","Charged 2310 IDR","50 min Tsel"],
              [f"Attach Offer {offerName}","Offer Attached","50 min Tsel"],
              ["Check notifikasi & Wording",wordingAddOffer,"50 min Tsel"],
              ["Check Offername",offerName,"50 min Tsel"],
              ["Create Event Voice Onnet 1800s 1pm on Zone 19","Consume Bonus","20 min Tsel"],
              ["Create Event Voice Offnet 2s 1pm on Zone 18 Check rounded 15s","Charged 750 IDR","20 min Tsel"],
              ["Create Event 10 SMS Onnet 1pm on Zone 11","Charge 3500 IDR","20 min Tsel"],
              ["Create Event Voice PSTN 15s 3pm on Zone 48 Check rounded 15s","Charged 750 IDR","20 min Tsel"],
              ["Create Event Voice FWA 28s 1pm on Zone 6 Check rounded 30s","Charged 1500 IDR","20 min Tsel, 20000000 sec Tsel"],
              ["Check notifikasi & Wording",wordingReachTreshold,"20 min Tsel, 20000000 sec Tsel"],
              ["Create Event Voice Onnet 6000s 3pm on Zone 7","Consume Bonus","20 min Tsel, 19994000 sec Tsel"],
              ["Create Event Voice Offnet 10s 3pm on Zone 14","Charged 462/ Back to Rate PP","20 min Tsel, 19994000 sec Tsel"],
              ["Check Rounded Above Event Voice","Rounded Should be not 15s","20 min Tsel, 19994000 sec Tsel"],
              ["Create Event Voice Onnet 6000s 3pm on Zone 14","Consume Bonus","20 min Tsel, 19988000 sec Tsel"],
              ["Create Event Voice PSTN 10s 3pm on Zone 48","Charge 56 IDR/ Back to rate PP","20 min Tsel, 19988000 sec Tsel"],
              ["Create Event Voice FWA 16s 3pm on Zone 44","Charged 462/ Back to Rate PP","20 min Tsel, 19988000 sec Tsel"],
              ["Create event 10 sms onnet 10am on zone 46 D+1","Charge 3850 IDR","No Bonus"],
              ["Create Event Voice Offnet 7s 11am on Zone 36 D+1","Charged 750 IDR","No Bonus"],
              ["Create Event Voice Onnet 13s 1pm on Zone 46 D+1 Rate ANPS","Charged 750 IDR","No Bonus"],
              ["Create Event voice Offnet Initial 2pm D+1 on Zone 5 Rate ANPS","Initial Success","No bonus"],
              ["Create Event Voice Offnet Intermediate 180s 2pm D+1 on Zone 5","Intermediate Succes","No bonus"],
              ["Create Event Voice Offnet Terminate 0s 2pm D+1 on Zone 5 charge > threshold XgetY","Charged not 1500 IDR","20000000 sec Tsel"],
              ["Check Pricing Item ID and It's Name rate ANPS charge PP Voice Offnet - Flat rate","Checked","20000000 sec Tsel"],
              ["Create event 10 sms onnet 10am on zone 46 D+10","Charge 3850 IDR","No Bonus"],
              ["Create Event Voice FWA 7s 11am on Zone 36 D+10 Rate ANPS","Charged 750 IDR","No Bonus"],
              ["Create Event Voice Onnet 60s 1pm on Zone 46 D+10 Cross Threshold","Charged 3000 IDR","20000000 sec Tsel"],
              ["Create Event Voice Offnet 75s 1pm on Zone 30 D+10 Rate ANPS - Tarif flat tanpa rate 0","Charged not 2250 IDR","20000000 sec Tsel"],
              ["Create Event Voice Offnet 10s 3pm on Zone 14 D+10 Rounded should be not 15s","Charged 462/ Back to Rate PP","20000000 sec Tsel"],
              ["Create Event Voice Offnet Initial 11am D+30 on Zone 10","Initial Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success | Drop Call | Is Reversed = YES","No Bonus"],
              ["Create Event Voice Offnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Rejected 4012","No Bonus"],
              ["Create Event Voice Offnet Terminate 0s 11am D+30 on Zone 10",dropOffnet,"20000000 sec Tsel"],
              ["Create Event Voice Onnet 6000s 3pm D+30 on Zone 14","Consume Bonus","19994000 sec Tsel"],
              ["Create Event Voice PSTN 10s 3pm D+30 on Zone 48 | Check rounded should be not 15s","Charge 56 IDR/ Back to rate PP","19994000 sec Tsel"],
              ["Create Event Direct Debit with vascode cst_digi_350 3pm D+30","Charged","19994000 sec Tsel"],
              ["Create Event Voice PSTN 60s 7pm 2022-12-31 on Zone 51","Charged 3000 IDR","20000000 sec Tsel"],
              ["Create Event Voice Onnet 37s 1pm 2023-01-01 on Zone 20 Check rounded should be 45s","Charged 2250 IDR","No Bonus"],
              ["Create Event GPRS 1MB RG 50 7pm 2023-01-01 on Zone 2","Charged","No Bonus"],
              ["Create Event Voice Offnet 5s 10pm 2023-01-01 on Zone 27 Check rounded should be 15s","Charged 750 IDR","20000000 sec Tsel"],
              [f"Remove Offer {offerName}","Offer Removed","No Bonus"],
              ["Create Event Voice Onnet 60s 1pm on Zone 14 Rate PP Voice Onnet","Charged not 3000 IDR","60 Min Tsel"],
              ["Checkd Indira","Checked","60 Min Tsel"],
              [f"Create & Activate new subscriber PP {PPName}","Check active period",preloadBonusString],
              ["Update Balance 1000K","Balance Updated",preloadBonusString],
              ["Update exp date 2023-12-31","Updated",preloadBonusString],
              [stepConsumePreload],
              ["Create event 80s voice onnet 9am D+0 zone ID =  2","Charged 2310 IDR","50 min Tsel"],
              [f"Attach Offer {offerName}","Offer Attached","50 min Tsel"],
              ["Check notifikasi & Wording",wordingAddOffer,"50 min Tsel"],
              ["Create Event Voice Onnet 3000s 1pm on Zone 19","Consume Bonus","Akumulasi 2310 monetary"],
              ["Create Event Voice Onnet 2s 1pm on Zone 18 Check rounded should be 15s","Charged 750 IDR","Akumulasi 2310 monetary"],
              ["Create Event 10 SMS Onnet 1pm on Zone 11","Charge 3850 IDR","Akumulasi 2310 monetary"],
              ["Create Event Voice Offnet 18s 1pm on Zone 48 Check rounded should be 30s","Charged 1500 IDR","Akumulasi 2310 monetary"],
              ["Create Event Voice Onnet 5s 1pm on Zone 6 Check rounded should be 15s","Charged 750 IDR","20000000 sec Tsel, Akumulasi 2310 monetary"],
              ["Check notifikasi & Wording",wordingReachTreshold,"20000000 sec Tsel, Akumulasi 2310 monetary"],
              ["Create Event Voice Onnet 6000s 3pm on Zone 7","Consume Bonus","19994000 min Tsel, Akumulasi 2310 monetary"],
              ["Create Event Voice Onnet Initial 11am D+30 on Zone 10","Initial Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Success | Drop Call | Is Reversed = YES","No Bonus"],
              ["Create Event Voice Onnet Intermediate 180s 11am D+30 on Zone 10","Intermediate Rejected 4012","No Bonus"],
              ["Create Event Voice Onnet Terminate 0s 11am D+30 on Zone 10",dropOnnet,"20000000 sec Tsel"],
              ["Create Event voice Offnet Initial 2pm D+31 on Zone 6 Rate ANPS","Initial Success","No bonus"],
              ["Create Event Voice Offnet Intermediate 180s 2pm D+31 on Zone 6","Intermediate Succes","No bonus"],
              ["Create Event Voice Offnet Terminate 0s 2pm D+31 on Zone 6 charge > threshold XgetY",dropOffnet,"20000000 sec Tsel"],
              ["Create Event Voice Onnet 60s 7pm 2022-12-31 on Zone 51","Charged 3000 IDR","20000000 sec Tsel"],
              ["Create Event Voice Onnet 27s 1pm 2023-01-01 on Zone 19","Charged 1500 IDR","No Bonus"],
              ["Create Event GPRS 1MB RG 50 7pm 2023-01-01 on Zone 2","Charged","No Bonus"],
              ["Create Event Voice PSTN 18s 3pm 2023-01-01 on Zone 27","Charged 1500 IDR","20000000 sec Tsel"],
              [f"Remove Offer {offerName}","Offer Removed","No Bonus"],
              ["Create Event Voice Onnet 60s 1pm on Zone 14 Rate PP Voice","Charged not 3000 IDR","30 Min Tsel"],
              ["Create Event Voice Offnet 60s 1pm on Zone 10 Rate PP Voice","Charged not 3000 IDR","30 Min Tsel"],
              ["Create Event Voice PSTN 60s 1pm on Zone 23 Rate PP Voice","Charged not 3000 IDR","30 Min Tsel"],
              ["Create Event Voice FWA 60s 1pm on Zone 31 Rate PP Voice","Charged not 3000 IDR","30 Min Tsel"],
              ["Check Indira","Checked","No Bonus"]
       ]

       return steps
